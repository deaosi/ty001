# -*- coding: utf-8 -*-
"""
探域数据看板 - 本地服务后端

提供:
  GET  /api/shops            - 店铺列表(来自本地 shops.json)
  GET  /api/overview         - 平台维度汇总指标
  GET  /api/shop/{id}        - 单店汇总指标 + 按日明细
  GET  /api/refresh          - 强制刷新全部店铺数据(异步)
  GET  /api/tasks            - 刷新任务状态
  POST /api/cookies          - 更新 Cookie 配置

数据源: https://agent.tanyuai.com/api/data-service/business/compass/*
"""
import asyncio
import collections
import datetime
import json
import os
import re
import sys
import threading
import time
import urllib.parse
from io import BytesIO
from pathlib import Path

# Windows 控制台默认 GBK, 打印 emoji/生僻字符会抛 UnicodeEncodeError 打崩请求线程。
# 重配 stdout/stderr 为 UTF-8 并 errors=replace, 日志乱码可接受但绝不崩。
if sys.stdout and hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass
if sys.stderr and hasattr(sys.stderr, "reconfigure"):
    try:
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

import requests

# 跨进程存活探测用 psutil.pid_exists(Windows 的 os.kill(pid,0) 对已死进程误报存活)。
# psutil 是轻量纯查询(不联网), 每次 _nightly_fetch_active 调用 import 一次有开销,
# 提前到模块级只 import 一次; 未安装时不阻断(退化 os.kill+TTL 兜底)。
try:
    import psutil as _psutil
except ImportError:
    _psutil = None
import uvicorn
from fastapi import Body, Depends, FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, RedirectResponse, Response
from pydantic import BaseModel

import openpyxl  # 导入平台 Excel 解析(openpyxl 3.1.5, 环境已装; pandas/xlrd 未装勿依赖)

import auth        # 账号系统: 密码哈希 / token / 鉴权依赖
import trace_store  # SQLite 消息轨迹存储层(35 天滚动窗口, 全平台)
import log_store    # 独立轻量日志库(logs.db: 任务历史 + 系统日志, 与主库解耦)

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
CONFIG_FILE = BASE_DIR / "config.json"
SHOPS_FILE = BASE_DIR / "shops.json"

API_BASE = "https://agent.tanyuai.com/api/data-service/business/compass"

# 平台枚举(与探域前端源码一致): 0=淘宝 1=拼多多 2=有赞 4=快手 5=抖店 7=京东 8=视频号 9=得物 10=1688
# 本账号只用到 1/5/7; 7=京东(集团"京东"的店铺, 非枚举字面上的天猫), 2=有赞无店铺
# 本系统内 10/11 = 天猫1/天猫2(非 tanyu 枚举字面的 1688): 无抓取能力, 靠 Excel 文档导入
PLATFORM_NAMES = {0: "淘宝", 1: "拼多多", 2: "有赞", 4: "快手", 5: "抖音", 7: "京东", 10: "天猫1", 11: "天猫2"}

# 抓取配置: 拼多多(1)/京东(7, "京东"集团店铺)/抖音(5)抓取; 淘宝(0)/快手(4)保留接口但不抓取
# 注: 本账号店铺只涉及 1/5/7, 故 fetch 列表不含 2(有赞)等未用到平台
FETCH_PLATFORMS = [1, 5, 7]
# 导入平台(10/11 = 天猫1/天猫2): 无法从 tanyu 抓取, 数据靠 Excel 文档上传入库。
# 绝不加入 FETCH_PLATFORMS——否则 prefetch/核算 worker 会把它们当可抓取平台逐店发
# tanyu 请求(每次必失败+限速风控)。所有在线抓取入口必须对这些平台短路到纯 DB 聚合。
IMPORT_PLATFORMS = [10, 11]
KEEP_PLATFORMS = [0, 4]  # 保留接口, 不抓取不统计

# 导入平台店铺清单(天猫1 = platform 10, 9 店来自 RPA 登记的 Excel 文档)。
# thirdShopId 为确定性合成 ID(imp10_ + md5(店名)[:12]), 重复导入幂等。
# 只 upsert 进 SQLite shops 表(展示链路 load_all_shops→get_shops 读取),
# 绝不写 shops.json——sync_shops_from_tanyu 每次切集团全量覆写会冲掉它。
# 天猫1 = platform 10(9 店), 天猫2 = platform 11(4 店, 2026-08-13 注册自 RPA Excel 文档)。
IMPORT_SHOPS = [
    {"thirdShopId": "imp10_5f1ac7a4c2eb", "shopName": "魔鬼猫青橙专卖店", "platform": 10},
    {"thirdShopId": "imp10_28f77c920652", "shopName": "果时代数码旗舰店", "platform": 10},
    {"thirdShopId": "imp10_c27d29667f41", "shopName": "MONSTER魔声安诺专卖店", "platform": 10},
    {"thirdShopId": "imp10_70d7a1ab08cc", "shopName": "星科数码专营", "platform": 10},
    {"thirdShopId": "imp10_8edee4132ec3", "shopName": "联想华荣专卖店", "platform": 10},
    {"thirdShopId": "imp10_c9a83b825a6d", "shopName": "星诚影音专营店", "platform": 10},
    {"thirdShopId": "imp10_fe29a40e541c", "shopName": "联想博睿兴专卖店", "platform": 10},
    {"thirdShopId": "imp10_3d083ccf86c5", "shopName": "联想聚源专卖店", "platform": 10},
    {"thirdShopId": "imp10_447e9133401a", "shopName": "飞利浦昕屿专卖店", "platform": 10},
    {"thirdShopId": "imp11_2645e03ed6a3", "shopName": "华硕逸聆专卖店", "platform": 11},
    {"thirdShopId": "imp11_b0b5038b175b", "shopName": "联想博升专卖店", "platform": 11},
    {"thirdShopId": "imp11_64b9cf1743d4", "shopName": "飞利浦聚源专卖店", "platform": 11},
    {"thirdShopId": "imp11_ce38cb51504c", "shopName": "魔声佳越专卖店", "platform": 11},
]
_IMPORT_SHOPS_BY_NAME = {s["shopName"]: s for s in IMPORT_SHOPS}
# 大小写不敏感索引: RPA 文档里店铺名大小写不统一(如 "monster魔声安诺专卖店" 小写开头)
_IMPORT_SHOPS_BY_NAME_LOWER = {s["shopName"].lower(): s for s in IMPORT_SHOPS}
# 店铺行补平台名(供 /api/shops 展示"天猫1"而非数字): upsert_shops 写 platform_name 列,
# get_shops 读不到时回退 str(platform) 会显示 "10"。tanyu 抓取店铺由 sync_shops 带真名。
for _imp_s in IMPORT_SHOPS:
    _imp_s.setdefault("platformName", PLATFORM_NAMES.get(_imp_s["platform"], str(_imp_s["platform"])))


def _import_shop_id(shop_name):
    """按店名找导入平台合成 ID; 未登记的店名返回 None(导入时该行跳过并记 warning)

    优先精确匹配, 回退大小写不敏感匹配(RPA 文档大小写不统一)。
    """
    s = _IMPORT_SHOPS_BY_NAME.get(shop_name)
    if not s and shop_name:
        s = _IMPORT_SHOPS_BY_NAME_LOWER.get(shop_name.lower())
    return s["thirdShopId"] if s else None

# 汇总指标定义(标题 / 格式化 / 排序权重)
METRICS = [
    {"key": "service_3m_response_rate", "title": "3分钟回复率", "format": "percent"},
    {"key": "ai_consult_response_accept_rate", "title": "采纳率", "format": "percent"},
    {"key": "ai_consult_response_rate", "title": "生成率", "format": "percent"},
]
METRIC_BY_KEY = {m["key"]: m for m in METRICS}

# 刷新状态
_refresh_state = {
    "running": False,
    "progress": {"done": 0, "total": 0, "current": ""},
    "last_run": None,
    "error": None,
    "triggered_by": None,   # 谁发起的刷新(操作者用户名/昵称), 供状态展示
}
# 手动补抓数据状态(独立于 refresh: 用户选日期范围触发, 逐集团抓 trace 写库)
_prefetch_state = {
    "running": False,
    "progress": {"done": 0, "total": 0, "current": ""},
    "start_date": None,
    "end_date": None,
    "canceled": False,       # 取消标志(worker 每店边界检查)
    "started_at": None,
    "last_run": None,
    "error": None,
    "triggered_by": None,   # 谁发起的补抓(操作者用户名/昵称), 供状态展示
}
_lock = threading.Lock()
_rate_lock = threading.Lock()  # 限速窗口专用锁(独立于 _lock, 避免限速等待阻塞状态读写)
_config_lock = threading.Lock()  # config.json 读改写串行化(防多写者丢更新)

# ---------- 多用户任务队列(并发请求排队, 单并发串行执行) ----------
# 看板多人共用: 有人抓取时, 其他客服的抓取请求不再 409 拒绝, 而是入队显示"排队中",
# 当前任务结束后自动串行执行下一个。每个任务记录发起人/类型/进度/结果。
# 调度器只负责"何时启动"; 具体抓取仍走既有 worker(状态字典 running 标志),
# 队列与既有单并发约束(_assert_no_running_task)兼容, 不破坏现有语义。
import itertools as _itertools
_TASK_SEQ = _itertools.count(1)
_tasks_lock = threading.Lock()
_tasks = {}                # id -> 任务 dict(含历史, 供任务列表展示)
_tasks_order = []          # id 创建顺序(前端按时间倒序展示)
_task_queue = collections.deque()   # 排队中的 id
_running_task_id = None
_scheduler_started = False
MAX_TASK_HISTORY = 200   # 任务列表保留上限(防止长期运行无限增长)


def _enqueue_task(task_type, label, requested_by, params, requested_role=None):
    """创建任务入队; 返回任务 dict。当前无任务在跑时不入队(由调用方直接启动)。

    requested_role: 发起人角色('admin'/'user'), 供普通用户视角对管理员任务掩码。
    """
    global _scheduler_started
    with _tasks_lock:
        tid = next(_TASK_SEQ)
        task = {
            "id": tid, "type": task_type, "label": label,
            "requested_by": requested_by or "未登录", "status": "queued",
            "created_at": time.time(), "started_at": None, "finished_at": None,
            "progress": {"done": 0, "total": 0, "current": "排队中…"},
            "error": None, "result": None, "params": params,
            "requested_role": requested_role,
        }
        _tasks[tid] = task
        _tasks_order.append(tid)
        _task_queue.append(tid)
        # 列表保留上限: 长期运行不无限增长(前端按时间倒序取最近 N 条)
        while len(_tasks_order) > MAX_TASK_HISTORY:
            _old = _tasks_order.pop(0)
            _tasks.pop(_old, None)
        if not _scheduler_started:
            _scheduler_started = True
            threading.Thread(target=_task_scheduler, daemon=True).start()
    # 持久化到独立日志库(任务历史跨重启保留, 支撑池子浏览/日志中心/管理员删除)
    log_store.upsert_task(task)
    return task


def _queue_position(tid):
    """某任务在队列中的位置(0=队首, 即将执行); 不在队中返回 None"""
    with _tasks_lock:
        for i, qid in enumerate(_task_queue):
            if qid == tid:
                return i + 1
    return None


def _any_task_running():
    """任一抓取任务在跑(核算/今日/补抓/刷新/客服同步)"""
    return bool(_refresh_state.get("running") or _trace_state.get("running")
                or _today_state.get("running") or _prefetch_state.get("running")
                or _staff_sync_state.get("running"))


def _task_scheduler():
    """串行调度: 空闲时取出队首任务, 调用对应端点启动, 等其 worker 结束后下一个。"""
    global _running_task_id
    while True:
        # 等待无任何任务在跑(含交互请求抢占的空档)
        while _any_task_running():
            time.sleep(1.0)
        with _tasks_lock:
            if _running_task_id is not None or not _task_queue:
                time.sleep(1.0)
                continue
            tid = _task_queue.popleft()
            t = _tasks.get(tid)
            if not t or t["status"] != "queued":
                continue
            t["status"] = "running"
            t["started_at"] = time.time()
            _running_task_id = tid
        task = _tasks[tid]
        try:
            _run_task(task)
            task["status"] = "done"
        except _RequeueTask:
            # 空档被交互请求抢占: 放回队尾稍后重试, 不算失败
            task["status"] = "queued"
            task["started_at"] = None
            with _tasks_lock:
                _task_queue.append(task["id"])
        except Exception as e:
            task["status"] = "error"
            task["error"] = str(e)
            log_line("task", f"任务#{tid} {task['type']} 失败: {e}")
        finally:
            task["finished_at"] = time.time()
            with _tasks_lock:
                # 只清自己登记的任务 id: 交互请求可能刚登记了新任务(direct), 无条件
                # 清空会把它误判为无任务, 提前弹出队尾任务
                if _running_task_id == tid:
                    _running_task_id = None
            log_store.upsert_task(task)  # 调度器收尾持久化(状态已 done/queued/error)


class _RequeueTask(Exception):
    """调度器内部: 当前空档被抢占, 放回队尾重试"""


def _fake_request():
    """调度器调用端点用的最小 Request(无真实 HTTP; 绕过鉴权中间件)"""
    from starlette.requests import Request as _SR
    scope = {
        "type": "http", "asgi": {"version": "3.0", "spec_version": "2.3"},
        "http_version": "1.1", "method": "POST", "scheme": "http",
        "path": "/", "raw_path": b"/", "query_string": b"",
        "root_path": "", "headers": [(b"x-scheduler-task", b"1")],
        "client": ("127.0.0.1", 0), "server": ("127.0.0.1", 8080),
    }
    return _SR(scope)


def _run_task(task):
    """按任务类型调用对应端点启动, 并等待该类型 worker 结束, 进度镜像到任务"""
    req = _fake_request()
    typ, p = task["type"], task["params"]
    try:
        if typ == "overview":
            trace_overview(req, days=p.get("days", 7), platform=p.get("platform"),
                           force=p.get("force", 0), start=p.get("start"), end=p.get("end"),
                           from_cache=0, shop_ids=p.get("shop_ids"))
            state = _trace_state
        elif typ == "today":
            trace_today(req, platform=p.get("platform"), shop_ids=p.get("shop_ids"),
                        start_ts=p.get("start_ts", "00:00"), end_ts=p.get("end_ts"),
                        mode=p.get("mode", ""))
            state = _today_state
        elif typ == "prefetch":
            trace_prefetch_start(req, {"start": p.get("start"), "end": p.get("end")})
            state = _prefetch_state
        elif typ == "refresh":
            refresh(req)
            state = _refresh_state
        elif typ == "staff_sync":
            staff_names_sync(req)
            state = _staff_sync_state
        else:
            raise RuntimeError(f"未知任务类型: {typ}")
    except HTTPException as e:
        if e.status_code == 409:
            raise _RequeueTask()   # 空档被交互请求抢占, 重试
        raise
    # 调度器启动期间无人抢占: 记录发起人(端点内部用 fake request 拿不到真实用户名)
    try:
        state["triggered_by"] = task["requested_by"]
    except Exception:
        pass
    _await_state(state, task)


def _await_state(state, task):
    """等待某类型 worker 结束, 期间镜像进度到任务 dict。

    worker 由端点起线程、有启动延迟才置 running=True: 给 3s 宽限期判断是否
    "启动后结束"(running 从 True→False) 还是 "未启动直接返回"(缓存命中/无任务,
    running 全程 False → 宽限期后视为完成)。
    """
    grace = 3.0
    t0 = time.time()
    started = False
    while True:
        if state.get("running"):
            started = True
        try:
            task["progress"] = dict(state.get("progress") or {})
        except Exception:
            pass
        if started and not state.get("running"):
            break
        if not started and time.time() - t0 > grace:
            break
        time.sleep(0.5)
    try:
        task["progress"] = dict(state.get("progress") or {})
    except Exception:
        pass
    task["result"] = state.get("result")


def _task_public(task):
    """任务列表对外字段(含 params: 供前端"查看"按钮跳回对应平台+时间段视图)"""
    return {k: task.get(k) for k in ("id", "type", "label", "requested_by", "status",
                                     "created_at", "started_at", "finished_at",
                                     "progress", "error", "result", "params")}


def _queue_if_busy(request, task_type, label, params):
    """有任务在跑时入队并返回 queued 响应; 无任务在跑返回 None(调用方直接启动)。

    调度器调用(x-scheduler-task=1)且恰好被交互请求抢占时抛 409 → 调度器放回队尾重试。
    返回 queued 响应含 taskId/queuePosition, 前端据此轮询任务列表。
    """
    if not _any_task_running():
        return None
    if request.headers.get("x-scheduler-task") == "1":
        raise HTTPException(409, "任务进行中, 调度器稍后重试")
    task = _enqueue_task(task_type, label, _operator_label(request), params, _operator_role(request))
    return {"status": "queued", "taskId": task["id"],
            "queuePosition": _queue_position(task["id"]), "task": _task_public(task)}


def _register_running_task(state, task_type, label, requested_by, params=None, requested_role=None):
    """直接启动(未排队)的任务也登记进任务列表: 前端能一直看到"谁发起的 + 进度条"。

    只由交互请求(非调度器 x-scheduler-task)在端点直接启动 worker 时调用;
    调度器跑的排队任务已有记录, 由 _await_state 镜像进度, 不重复登记。
    - 该状态已空闲(直接启动前提), 若存在旧的 running 记录则标记 done(被新任务取代)
    - 占位 _running_task_id: 调度器在直接任务结束前不启动排队任务(保持单并发)
    requested_role: 发起人角色, 供普通用户视角对管理员任务掩码。
    """
    with _tasks_lock:
        global _running_task_id
        for t in list(_tasks.values()):
            if t.get("_state_ref") is state and t.get("status") == "running":
                t["status"] = "done"
                t["finished_at"] = time.time()
        tid = next(_TASK_SEQ)
        task = {
            "id": tid, "type": task_type, "label": label,
            "requested_by": requested_by or "未登录", "status": "running",
            "created_at": time.time(), "started_at": time.time(), "finished_at": None,
            "progress": dict(state.get("progress") or {}),
            "error": None, "result": None, "params": params or {},
            "requested_role": requested_role,
            "_state_ref": state, "_ever_started": False,
        }
        _tasks[tid] = task
        _tasks_order.append(tid)
        _running_task_id = tid
    # 持久化到独立日志库
    log_store.upsert_task(task)
    return task


def _sync_task_progress():
    """懒同步: 把 running 任务的进度/结果从对应状态字典镜像到记录; 状态结束则收尾。

    直接启动的任务由前端轮询 /api/tasks/list 触发这里, 无需额外线程。
    同一状态被新任务复用前, 旧的 running 记录已被 _register_running_task 标 done。
    _ever_started 区分"运行中→结束"(正常收尾)与"从未进入 running"(refresh 等
    worker 起线程后才置位, 轮询可能先于置位; 超时仍未启动按完成处理)。
    """
    with _tasks_lock:
        global _running_task_id
        for t in list(_tasks.values()):
            if t.get("status") != "running" or t.get("_state_ref") is None:
                continue
            state = t["_state_ref"]
            try:
                t["progress"] = dict(state.get("progress") or {})
                if state.get("result"):
                    t["result"] = state.get("result")
            except Exception:
                pass
            started = bool(state.get("running"))
            if started:
                t["_ever_started"] = True
            if started:
                log_store.upsert_task(t)  # 持久化进度(前端轮询触发, 频率低, 量小)
                continue
            if t.get("_ever_started") or (time.time() - t.get("created_at", 0) > 15):
                t["status"] = "error" if state.get("error") else "done"
                t["error"] = state.get("error")
                t["finished_at"] = time.time()
                if _running_task_id == t["id"]:
                    _running_task_id = None
                log_store.upsert_task(t)  # 任务收尾持久化
            else:
                log_store.upsert_task(t)


def _restore_task_history():
    """启动时从独立日志库恢复近期任务到内存(任务列表/池子跨重启可查)。

    运行中/排队中的重启后无法续跑, 标为"服务重启, 任务中断"(done); 已结束的保留原状。
    恢复后 _TASK_SEQ 从最大 id+1 起, 避免与 DB 已用 id 冲突。
    """
    global _TASK_SEQ, _running_task_id
    try:
        recent = log_store.query_tasks(limit=MAX_TASK_HISTORY)
    except Exception:
        recent = []
    max_id = 0
    with _tasks_lock:
        for t in recent:
            tid = int(t["id"])
            max_id = max(max_id, tid)
            if t.get("status") in ("running", "queued"):
                t["status"] = "done"
                t["error"] = t.get("error") or "服务重启, 任务中断"
                t["finished_at"] = time.time()
                try:
                    log_store.upsert_task(t)
                except Exception:
                    pass
            t["_state_ref"] = None  # 历史任务不再关联状态字典(懒同步跳过)
            _tasks[tid] = t
            _tasks_order.append(tid)
        _TASK_SEQ = _itertools.count(max_id + 1)
        _running_task_id = None


app = FastAPI(title="探域数据看板")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


# ---------- 操作日志中间件(会话可追溯 + 账号鉴权) ----------
# 鉴权: 除白名单外, 所有 /api/* 请求必须携带有效登录 token(Authorization: Bearer),
# 或内部服务 token(X-Service-Token, 定时任务 dingtalk_daily_push 用)。鉴权失败返回
# 401 JSON, 不进入业务逻辑。登录后把 user_id 关联进操作日志(管理员可追溯"谁干了什么")。
# 记录: 非轮询类 /api 请求写入 SQLite operation_log, 审计谁做了什么; 记录失败绝不阻塞请求。
# 跳过高频轮询 GET(前端 setInterval 定时拉取, 非用户主动操作, 刷屏无审计价值)。
_AUTH_WHITELIST = {
    "/api/auth/login", "/api/auth/register", "/api/auth/config", "/api/health",
}
_OPLOG_SKIP_GET = {
    "/api/tasks/refresh", "/api/tasks/list", "/api/groups/sync-status", "/api/trace/prefetch/status",
    "/api/trace/today/status", "/api/trace/overview/status",
    "/api/staff/names/status", "/api/login/status", "/api/risk", "/api/logs", "/api/logs/center", "/api/oplog",
}


@app.middleware("http")
async def _oplog_middleware(request: Request, call_next):
    user_id = None
    path = request.url.path
    method = request.method
    if path.startswith("/api/") and path not in _AUTH_WHITELIST:
        # 内部服务 token(定时任务)放行; 否则必须为有效登录用户
        if not auth.verify_service_token(request.headers.get("x-service-token")):
            user_id = auth.resolve_user_id(request)
            if user_id is None:
                return JSONResponse(
                    status_code=401,
                    content={"detail": "未登录或登录已过期, 请重新登录"},
                )
    response = await call_next(request)
    try:
        if path.startswith("/api/"):
            if method != "GET" or path not in _OPLOG_SKIP_GET:
                # 昵称走 URL 编码(header 规范是 latin-1, 中文直发会乱码)
                try:
                    client_name = urllib.parse.unquote(request.headers.get("x-client-name", ""))
                except Exception:
                    client_name = request.headers.get("x-client-name", "")
                trace_store.log_operation(
                    request.headers.get("x-client-id", ""),
                    client_name,
                    request.client.host if request.client else "",
                    method, path, request.url.query, response.status_code,
                    user_id=user_id,
                )
    except Exception:
        pass
    return response


# ---------- 进程内日志环形缓冲(供抽屉日志窗口展示) ----------
LOG_RING_MAX = 500
_log_ring = collections.deque(maxlen=LOG_RING_MAX)
_log_lock = threading.Lock()


def log_line(tag, msg):
    """写一条带时间戳的日志: 进环形缓冲(抽屉可查) + 打印到 stdout(server.log 追加)
    + 持久化到独立日志库(logs.db, 供管理员日志中心分类查看/删除)。

    绝不写入 cookie 值——所有日志内容只允许业务描述, 禁止任何凭据/密钥字段。
    """
    entry = {"ts": time.strftime("%Y-%m-%d %H:%M:%S"), "tag": tag, "msg": msg}
    try:
        with _log_lock:
            _log_ring.append(entry)
    except Exception:
        pass
    try:
        log_store.append_system_log(tag, msg)
    except Exception:
        pass
    print(f"[{tag}] {msg}")


# ---------- 配置 ----------
def default_config():
    return {
        "cookies": {
            "tanyu-account-id": "2656113728446465571",
            "tanyu-agent-account": "fM_VS4GirTjMlPPJx_llv5kWStXKTMrRvW__",
            "tanyu-group-account": "le_FjY2F9WFwuBxC2_ISmD0LfcZkkuHaG9__",
            "tanyu-group-id": "1901419852011174006",
        },
        "groups": [],  # 已发现的集团列表(来自微信扫码登录 localStorage.groupList)
        # 夜间预抓多集团配置: 平台顺序(1拼多多 5抖音 7京东)、
        # 默认窗口天数(未在 prefetch_windows 列出的平台用此值)、
        # 各平台窗口覆盖: 全平台保留 35 天(与 config.json 的 prefetch_windows 保持一致)。
        #   窗口=各平台保留的历史天数上限: 每晚滚动(新增昨天、超过窗口的最老一天被裁剪)。
        # prefetch_force_days: 预抓时窗口最近 N 天强制重抓(默认7=近7天),
        #   防 tanyu 回溯更新 sendType 造成的采纳口径漂移(近7天核算对齐 tanyu 后台);
        #   0=关闭。2026-08-10 由 1 调为 7(用户反馈核算近7天与后台对比有误差,
        #   实测 tanyu 对已抓消息回溯更新 sendType + 补漏消息, 只重抓昨天不够)。
        "prefetch_platforms": [1, 5, 7],
        "prefetch_days": 35,
        "prefetch_windows": {1: 35, 5: 35, 7: 35},
        "prefetch_force_days": 7,
    }


# groupName 含平台名的关键词 → 平台号。探域集团名形如「星科数码专营-拼多多」,
# 登录写 groups 时若漏带 platform(accountType=3 子账号场景), 靠名字推断补上,
# 否则 _today_target_shops / 客服同步的 s.platform == g.platform 永假, 全部店铺被过滤空转。
# 关键词必须比枚举数字名(天猫1/天猫2)更精确: 先匹配 "天猫" 再匹配 "京东",
# 避免「京东」前缀误伤「天猫京东店」类店名; 淘宝/快手/有赞仅保留接口不抓取。
_GROUP_PLATFORM_KEYWORDS = [
    ("拼多多", 1), ("抖音", 5), ("京东", 7), ("天猫", 10), ("淘宝", 0), ("快手", 4), ("有赞", 2),
]


def _infer_group_platform(group):
    """按 groupName 关键词推断平台号; 已有显式 platform 原样返回"""
    plat = group.get("platform")
    if plat is not None:
        return plat
    name = group.get("groupName") or ""
    for kw, p in _GROUP_PLATFORM_KEYWORDS:
        if kw in name:
            return p
    return None


def _load_config_unlocked():
    """不加锁读取 config(调用方必须已持有 _config_lock 或确信无并发写)"""
    if CONFIG_FILE.exists():
        try:
            cfg = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
            # 内存补齐 groups 缺失的 platform(不下盘: 避免与夜间进程的
            # config 写者争用, 且推断值可能不精确, 落盘会污染真实配置)
            for g in cfg.get("groups") or []:
                if g.get("groupId") and g.get("platform") is None:
                    g["platform"] = _infer_group_platform(g)
            return cfg
        except Exception as e:
            # 文件存在但解析失败: 绝不静默覆盖真实配置, 先备份到 .bak 再重建默认
            print(f"[config] config.json 解析失败({e}), 已备份为 config.json.bak")
            try:
                bak = CONFIG_FILE.with_name("config.json.bak")
                os.replace(CONFIG_FILE, bak)
            except Exception:
                pass
    cfg = default_config()
    return cfg


def load_config():
    if not CONFIG_FILE.exists():
        cfg = _load_config_unlocked()
        save_config(cfg)
        return cfg
    return _load_config_unlocked()


def save_config(cfg):
    # 原子写入: 先写临时文件再 os.replace, 读取方不会读到半截 JSON。
    # 锁内写: 多个 config 写者(switch_group/登录/扫码/cookie 更新)串行化,
    # 配合 mutate_config 的读改写整体持锁, 消除丢更新。
    with _config_lock:
        _atomic_write_text(CONFIG_FILE, json.dumps(cfg, ensure_ascii=False, indent=2))


def mutate_config(fn, default=True):
    """config 读-改-写整体持锁: fn(cfg) 修改后原子落盘, 返回 cfg。

    所有"load_config → 修改 → save_config"路径都应改用本函数, 防止并发写者
    互相覆盖(登录 worker 保存 cookie 与 switch_group 保存集团 cookie 竞争)。
    """
    with _config_lock:
        cfg = _load_config_unlocked()
        fn(cfg)
        _atomic_write_text(CONFIG_FILE, json.dumps(cfg, ensure_ascii=False, indent=2))
        return cfg


# ---------- 夜间抓取跨进程标志 ----------
# nightly_fetch.py 是独立进程, 逐集团改写 config.json 的 tanyu-group-id cookie
# (切换激活集团, 10~40 分钟/晚)。期间 8080 常驻的任何直连 tanyu 请求若带错误
# 激活集团 cookie 出网, 会返回错集团数据(跨集团串数据)并被 6h TTL 缓存放大。
# 缓解: 夜间进程写独立标志文件 data/nightly_fetch_active.json{pid,at}, 常驻检测到
# (且非本进程发起)时, 直连请求立即返回"繁忙"(503/请稍后), 不向 tanyu 出网。
# 独立小文件而非 config.json 键: 标志 set/clear 不再与 switch_group/登录回调等
# config 写者共享 read-modify-write 路径, 从根上消除写侧 TOCTOU(标志被并发覆盖丢)。

NIGHTLY_FETCH_FLAG_TTL = 7 * 3600  # 标志最长有效期(计划任务执行时限 6h 上限的裕量)
NIGHTLY_FETCH_FLAG_FILE = DATA_DIR / "nightly_fetch_active.json"


def set_nightly_fetch_flag():
    """夜间抓取进程在抓取开始前调用: 原子写独立标志文件(pid+时间戳)"""
    try:
        _atomic_write_text(NIGHTLY_FETCH_FLAG_FILE, json.dumps(
            {"pid": os.getpid(), "at": time.time()}, ensure_ascii=False))
    except Exception as e:
        print(f"[prefetch] ⚠️ 写夜间抓取标志失败: {e}")


def clear_nightly_fetch_flag():
    """夜间抓取结束/异常时调用: 删除标志文件(仅限本进程写入的标志)"""
    try:
        data = json.loads(NIGHTLY_FETCH_FLAG_FILE.read_text(encoding="utf-8"))
        if data.get("pid") == os.getpid():
            NIGHTLY_FETCH_FLAG_FILE.unlink(missing_ok=True)
    except FileNotFoundError:
        pass
    except Exception:
        pass


def _nightly_fetch_active():
    """8080 常驻判断夜间抓取是否正在占用激活集团 cookie(阻塞直连请求)

    三重失效兜底, 缺一不可:
      1) pid==os.getpid(): 本进程就是夜间抓取, 不阻塞自己
      2) 进程存活探测(os.kill(pid,0)): 硬杀(taskkill /F/断电/蓝屏/os._exit)
         不走 prefetch 的 finally 清除标志, 死进程残留标志立即失效, 无需等 7h TTL
         (与 nightly_fetch._lock 的存活探测同模式)
      3) TTL: 7h 硬上界(pid 复用可能让存活探测误报存活, 时间兜底保证绝对自愈)
    标志是独立小文件(data/nightly_fetch_active.json), 与 config.json 解耦,
    不与 switch_group 等 config 写者共享读改写路径(无写侧 TOCTOU 覆盖丢失)。
    """
    try:
        data = json.loads(NIGHTLY_FETCH_FLAG_FILE.read_text(encoding="utf-8"))
    except Exception:
        return False
    pid = data.get("pid")
    at = data.get("at")
    if not pid or not at:
        return False
    if pid == os.getpid():
        return False  # 本进程就是夜间抓取, 不阻塞自己
    if time.time() - float(at) > NIGHTLY_FETCH_FLAG_TTL:
        return False  # 残留标志超时, 不再阻塞
    # 写入方进程存活探测: 进程已死视为标志失效, 立即解除 8080 阻塞(不等 7h TTL)。
    # Windows 的 os.kill(pid,0) 对已死进程仍返回成功(误报存活), 故优先 psutil。
    # pid_exists 精确(查进程是否真实存在); 未装 psutil 时退化 os.kill+TTL 兜底。
    if _psutil is not None:
        try:
            if not _psutil.pid_exists(int(pid)):
                return False
        except Exception:
            return False
    else:
        try:
            os.kill(int(pid), 0)
        except OSError:
            return False
    return True


def _nightly_flag_info():
    """夜间抓取标志详情(供前端展示"谁在抓取/为何阻塞"): 未激活返回 {"active": False}

    active=True 时附带: 来源(独立进程=系统定时任务, 本进程=手动补抓/今日抓取)、
    启动时间、已运行分钟数。与 _nightly_fetch_active 同一份标志文件的同一套
    失效兜底(pid 存活 + TTL), 保证展示与阻塞判定一致。
    """
    try:
        data = json.loads(NIGHTLY_FETCH_FLAG_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {"active": False}
    pid = data.get("pid")
    at = data.get("at")
    if not pid or not at:
        return {"active": False}
    if time.time() - float(at) > NIGHTLY_FETCH_FLAG_TTL:
        return {"active": False}
    if _psutil is not None:
        try:
            if not _psutil.pid_exists(int(pid)):
                return {"active": False}
        except Exception:
            return {"active": False}
    return {
        "active": True,
        "pid": pid,
        "at": at,
        "since": time.strftime("%H:%M:%S", time.localtime(at)),
        "minutes": int((time.time() - at) / 60),
        "source": "系统定时任务(夜间预抓)" if int(pid) != os.getpid() else "手动补抓(本看板)",
    }


def _operator_label(request):
    """当前请求操作者标签(供任务状态展示"这个抓取是哪个用户发起的")。

    登录用户返回「用户名(#uid)」, 便于任务列表/日志中心追溯是哪个账号;
    未登录时回退浏览器昵称(x-client-name, URL 编码)。解析失败返回 "未登录"。
    绝不返回 cookie/凭据。
    """
    try:
        uid = auth.resolve_user_id(request)
        if uid:
            u = trace_store.get_user_by_id(uid)
            if u and u.get("username"):
                return f"{u['username']} (#{uid})"
    except Exception:
        pass
    try:
        nick = urllib.parse.unquote(request.headers.get("x-client-name", ""))
        if nick:
            return f"{nick}(未登录)"
    except Exception:
        pass
    return "未登录"


def _operator_role(request):
    """当前请求操作者的角色('admin'/'user'/None), 供任务创建时记录发起人角色。

    掩码基于**任务创建时的角色**(而非查看时角色): 管理员中途降级后, 其旧任务
    仍对普通用户隐藏真实身份。
    """
    try:
        uid = auth.resolve_user_id(request)
        if uid:
            u = trace_store.get_user_by_id(uid)
            if u:
                return u.get("role")
    except Exception:
        pass
    return None


# ---------- 管理员身份掩码(普通用户看不到管理员的真实账号) ----------
def _viewer_is_admin(request) -> bool:
    """当前请求者是否为管理员(决定是否对管理员发起的任务做身份掩码)。"""
    try:
        uid = auth.resolve_user_id(request)
        if not uid:
            return False
        u = trace_store.get_user_by_id(uid)
        return bool(u and u.get("role") == "admin")
    except Exception:
        return False


def _label_is_admin(label) -> bool:
    """解析操作者标签(用户名(#uid) / 用户名 / 昵称), 判断是否属于管理员账号。"""
    s = label or ""
    m = re.search(r"\(#(\d+)\)", s)
    if m:
        try:
            u = trace_store.get_user_by_id(int(m.group(1)))
            return bool(u and u.get("role") == "admin")
        except Exception:
            pass
    name = s.split("(")[0].strip() if "(" in s else s
    if not name or name in ("未登录", "admin"):
        return False
    try:
        u = trace_store.get_user_by_username(name)
        return bool(u and u.get("role") == "admin")
    except Exception:
        return False


def _mask_admin_label(request, label, role=None):
    """非管理员视角: 把管理员发起的操作者标签抹成 'admin', 隐藏真实账号。

    管理员自己看(日志中心/自己的任务列表)不受影响, 显示真实昵称。
    优先用任务创建时记录的角色(role='admin'), 保证管理员中途降级后旧任务仍隐藏;
    无 role 时按标签解析 uid/用户名判断。
    """
    if _viewer_is_admin(request):
        return label
    if role == "admin":
        return "admin"
    return "admin" if _label_is_admin(label) else label


# ---------- 数据访问 ----------
def get_headers():
    cfg = load_config()
    cookies = cfg.get("cookies", {})
    cookie_str = "; ".join(f"{k}={v}" for k, v in cookies.items())
    return {
        "Cookie": cookie_str,
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36",
        "Content-Type": "application/json",
        "Accept": "application/json, text/plain, */*",
        "Referer": "https://agent.tanyuai.com/",
        "Origin": "https://agent.tanyuai.com",
    }


def post_api(path, payload, timeout=20):
    """POST 请求探域接口, 返回解析后的 JSON; 带限速+风控检测"""
    _assert_no_risk()
    # 夜间抓取进程正在逐集团切激活集团 cookie, 此刻出网会串集团数据 → 直接阻塞
    if _nightly_fetch_active():
        raise BusyQueueError("夜间抓取进行中, 请稍后重试")
    _rate_limit()
    resp = requests.post(API_BASE + path, json=payload, headers=get_headers(), timeout=timeout)
    data = resp.json()
    if _check_risk(resp, data):
        # 抛 RiskTriggered(继承 RuntimeError)而非裸 RuntimeError: 首次命中即让调用方
        # 的 except RiskTriggered 分支统一走"风控/登录失效停止", 而不是被通用 except
        # 吞成局部回落/假零(否则要到下一次 _assert_no_risk 才抛 RiskTriggered 停表)
        raise RiskTriggered(f"风控/登录失效: {data.get('msg', resp.status_code)}")
    if data.get("code") != 0:
        raise RuntimeError(data.get("msg", "接口错误"))
    return data.get("data")


# ---------- 交互式请求快通道 ----------
# 看板页面加载/切换时间段的请求与后台抓取任务共用同一套全局限速。
# 若限速器队列已满(后台任务在跑), 这些交互请求会排队 30~60s, 前端体验极差。
# 方案: 交互请求不排队等待 —— 队列繁忙时立即返回 503(不发任何请求),
# 由前端展示"稍后重试"并在后台任务结束后自动补拉。
INTERACTIVE_CACHE_TTL = 10          # 快通道缓存秒数(交互场景重复查询零上游请求)
_interactive_cache = {}             # { (path, json_str): (ts, data) }


def _interactive_cache_get(path, payload):
    key = (path, json.dumps(payload, ensure_ascii=False, sort_keys=True))
    hit = _interactive_cache.get(key)
    if hit and time.time() - hit[0] < INTERACTIVE_CACHE_TTL:
        return hit[1]
    return None


def _interactive_cache_put(path, payload, data):
    global _interactive_cache
    key = (path, json.dumps(payload, ensure_ascii=False, sort_keys=True))
    _interactive_cache[key] = (time.time(), data)
    if len(_interactive_cache) > 64:  # 防内存无限增长
        cutoff = time.time() - INTERACTIVE_CACHE_TTL
        _interactive_cache = {k: v for k, v in _interactive_cache.items() if v[0] > cutoff}


def post_api_interactive(path, payload, timeout=20):
    """交互式请求: 不排队, 限速繁忙时立即返回 503; 10s 内同参命中内存缓存零请求

    缓存与首次返回保持同一形状: 都返回 data.data(内层业务数据)。
    注意: 绝不能把完整响应 {code,msg,data:{...}} 存进缓存——命中后返回完整响应,
    调用方 `.get("items")/.get("dates")` 会取不到内层字段, 表现为"间歇性空数据"。
    """
    cached = _interactive_cache_get(path, payload)
    if cached is not None:
        return cached
    _assert_no_risk()
    # 夜间抓取进程正在切激活集团 cookie → 交互直连请求不发, 前端显示"请稍后"
    if _nightly_fetch_active():
        raise BusyQueueError("夜间抓取进行中, 请稍后重试")
    if not _rate_limit_available():
        raise BusyQueueError("后台抓取任务进行中, 请稍后重试")
    resp = requests.post(API_BASE + path, json=payload, headers=get_headers(), timeout=timeout)
    data = resp.json()
    if _check_risk(resp, data):
        # 抛 RiskTriggered(继承 RuntimeError)而非裸 RuntimeError: 首次命中即让调用方
        # 的 except RiskTriggered 分支统一走"风控/登录失效停止", 而不是被通用 except
        # 吞成局部回落/假零(否则要到下一次 _assert_no_risk 才抛 RiskTriggered 停表)
        raise RiskTriggered(f"风控/登录失效: {data.get('msg', resp.status_code)}")
    if data.get("code") != 0:
        raise RuntimeError(data.get("msg", "接口错误"))
    inner = data.get("data") or {}
    _interactive_cache_put(path, payload, inner)
    return inner


class BusyQueueError(RuntimeError):
    """限速队列繁忙(后台任务占用), 交互请求不排队直接返回"""


def load_shops(platform=None):
    """加载本地店铺列表, 合并平台名; 可按平台过滤"""
    if not SHOPS_FILE.exists():
        return []
    shops = json.loads(SHOPS_FILE.read_text(encoding="utf-8")).get("data", [])
    for s in shops:
        s["platformName"] = PLATFORM_NAMES.get(s.get("platform"), str(s.get("platform")))
    if platform is not None:
        shops = [s for s in shops if s.get("platform") == platform]
    return shops


def load_all_shops(platform=None):
    """加载跨集团全量店铺(SQLite 优先, shops.json 兜底), 可按平台过滤

    与 load_shops(shops.json, 仅当前激活集团)不同: 这里覆盖三个集团的店铺,
    平台切换后任何平台的店铺列表都有数据(客服池/总览同源)。SQLite 为空
    (如首次运行未回填)时回退 shops.json, 保证不因新数据源引入空列表。
    """
    try:
        shops = trace_store.get_shops(platform)
        if shops:
            return shops
    except Exception:
        pass
    # 兜底: shops.json 仅有当前激活集团店铺, 过滤后可能为空(切平台场景)
    return load_shops(platform)


def date_range(days=7, end=None):
    """生成近 N 天日期区间(截至昨天); days 统一钳制 1~35(滚动窗口上限, 防滥用)"""
    import datetime

    days = max(1, min(int(days), 35))
    end = end or (datetime.date.today() - datetime.timedelta(days=1))
    start = end - datetime.timedelta(days=days - 1)
    return start.isoformat(), end.isoformat()


def fetch_summary(payload):
    """拉取汇总指标, 返回 {key: {current, previous, comparePercent}}"""
    data = post_api("/summary", payload)
    items = data.get("items", []) if data else []
    return {it["key"]: it for it in items}


def fetch_summary_interactive(payload):
    """交互式汇总: 走快通道(不排队), 10s 内同参命中缓存零上游请求"""
    data = post_api_interactive("/summary", payload)
    items = data.get("items", []) if data else []
    return {it["key"]: it for it in items}


def fetch_section_table(payload):
    """拉取 section 明细表, 返回 {dates: [...], rows: [...]}"""
    data = post_api("/section/table", payload)
    if not data:
        return {"dates": [], "rows": []}
    return {"dates": data.get("dates", []), "rows": data.get("rows", [])}


def fetch_section_table_interactive(payload):
    """交互式 section 明细: 走快通道(不排队), 10s 内同参命中缓存零上游请求"""
    data = post_api_interactive("/section/table", payload)
    if not data:
        return {"dates": [], "rows": []}
    return {"dates": data.get("dates", []), "rows": data.get("rows", [])}


# ---------- 罗盘扩展: 趋势 & 客服经营 ----------
# 接口来自探域"数据罗盘"前端(web-cdn.tanyuai.com agent-center bundle), 已实测可用:
#   POST /api/data-service/business/compass/section/trend            分节趋势(按天序列)
#   POST /api/data-service/business/compass/customer-service/summary 客服经营汇总
#   POST /api/data-service/business/compass/customer-service/detail/table  客服明细表
#   POST /api/data-service/business/compass/customer-service/detail/trend  客服明细趋势
# 默认指标键(与罗盘前端定义一致):
#   operations: 销售额/支付人数/退款金额;  service: 咨询量/采纳数/客服销售额
#   ai: 咨询消息数/采纳数/采纳率
TREND_SECTION_METRICS = {
    "operations": ["ops_order_payment_amount", "ops_order_payment_uv",
                   "ops_order_refund_amount_by_pay_time"],
    "service": ["service_consult_cnt", "service_accept_cnt", "service_sale_amount"],
    "ai": ["ai_consult_msg_cnt", "ai_consult_response_accept_cnt", "ai_consult_response_rate"],
}
CS_DETAIL_METRICS = {
    "service": ["service_consult_cnt", "service_accept_cnt", "service_sale_amount"],
    "ai": ["ai_consult_msg_cnt", "ai_consult_response_accept_cnt"],
}
CS_SUMMARY_LABELS = {
    "service_sale_amount": "客服销售额",
    "service_payment_conversion_rate": "支付转化率",
    "service_consult_cnt": "咨询量",
    "service_accept_cnt": "采纳数",
    "service_3m_response_rate": "3分钟响应率",
    "service_response_avg_sec": "平均响应秒数",
    "service_first_response_avg_sec": "首次响应秒数",
    "service_payment_cnt": "支付订单数",
}
TREND_METRIC_NAMES = {
    "ops_order_payment_amount": "销售额",
    "ops_order_payment_uv": "支付人数",
    "ops_order_refund_amount_by_pay_time": "退款金额",
    "service_consult_cnt": "咨询量",
    "service_accept_cnt": "采纳数",
    "service_sale_amount": "客服销售额",
    "ai_consult_msg_cnt": "咨询消息数",
    "ai_consult_response_accept_cnt": "采纳数",
    "ai_consult_response_rate": "采纳率",
}


def fetch_section_trend_interactive(payload):
    """交互式 section 趋势(按天序列): 走快通道(不排队), 10s 内同参命中缓存零上游请求"""
    data = post_api_interactive("/section/trend", payload)
    if not data:
        return {"dates": [], "series": []}
    return {"dates": data.get("dates", []), "series": data.get("series", [])}


def fetch_cs_summary_interactive(payload):
    """交互式客服经营汇总, 返回 {key: {current, previous, comparePercent}}"""
    data = post_api_interactive("/customer-service/summary", payload)
    items = data.get("items", []) if data else []
    return {it["key"]: it for it in items}


def fetch_cs_detail_table_interactive(payload):
    """交互式客服明细表(customer-service/detail/table)"""
    data = post_api_interactive("/customer-service/detail/table", payload)
    if not data:
        return {"dates": [], "rows": []}
    return {"dates": data.get("dates", []), "rows": data.get("rows", [])}


def fetch_cs_detail_trend_interactive(payload):
    """交互式客服明细趋势(customer-service/detail/trend)"""
    data = post_api_interactive("/customer-service/detail/trend", payload)
    if not data:
        return {"dates": [], "series": []}
    return {"dates": data.get("dates", []), "series": data.get("series", [])}


def cache_path(kind, target_id):
    """按类型+ID 定位缓存文件: data/{kind}/{id}.json"""
    d = DATA_DIR / kind
    d.mkdir(parents=True, exist_ok=True)
    return d / f"{target_id}.json"


TRACE_DAYS_CACHE_TTL = 60          # 内存里已解析的 trace_days 缓存秒数
_trace_days_cache = {}             # {shop_id: (ts, parsed_cache)}, 跨请求共享避免重复读盘+json 解析

# 按日缓存有效期(探域数据每日更新, 已抓取的天不能永久复用)
#   今天/昨天: 6 小时 TTL(当天数据随时可能更新, 短期窗口内也是新口径)
#   更早的天:  7 天 TTL(历史天数据已定型, 隔周刷新一次即可)
DAY_TTL = 6 * 3600          # 最近两天(含昨天)的有效期
HISTORY_DAY_TTL = 7 * 24 * 3600   # 更早历史天的有效期


def _day_cache_ttl(day_str):
    """按天返回缓存有效期秒数: 今天/昨天用短 TTL, 更早的历史用长 TTL"""
    try:
        d = datetime.date.fromisoformat(day_str)
        today = datetime.date.today()
        days_ago = (today - d).days
    except Exception:
        days_ago = 999
    return DAY_TTL if days_ago <= 1 else HISTORY_DAY_TTL


def _load_trace_days_cache(shop_id, force=False):
    """读取店铺按日缓存(带进程内共享缓存), force=True 时跳过内存缓存直接读盘"""
    _prune_trace_days_cache()
    if not force:
        hit = _trace_days_cache.get(shop_id)
        if hit and time.time() - hit[0] < TRACE_DAYS_CACHE_TTL:
            return hit[1]
    data = load_cache("trace_days", shop_id, max_age=HISTORY_DAY_TTL)
    _trace_days_cache[shop_id] = (time.time(), data)
    return data


def _cached_days_usable(cache, day_str):
    """判断某一天是否可直接复用: 在该天自己的有效期内则命中, 过期则视为缺失

    TTL 按天龄分级; 存量文件未按日记录抓取时间时, 回退到文件级 fetched_at。
    空天(0 条消息)只要在有效期内同样复用, 避免每次都重复抓取。
    """
    if not cache or not cache.get("days"):
        return False
    if day_str not in cache["days"]:
        return False
    fetched = (cache.get("day_fetched_at") or {}).get(day_str)
    if fetched is None:
        fetched = cache.get("fetched_at")
    if not fetched:
        return False
    return (time.time() - fetched) < _day_cache_ttl(day_str)


def _prune_trace_days_cache():
    """清理进程内共享缓存中超过 TTL 的条目(防无限增长)"""
    now = time.time()
    for shop_id, (ts, _) in list(_trace_days_cache.items()):
        if now - ts > TRACE_DAYS_CACHE_TTL:
            _trace_days_cache.pop(shop_id, None)


def load_cache(kind, target_id, max_age=21600):
    # 默认 TTL 6 小时: 平台/店铺汇总与明细是"自然日"数据, 一天内基本不变。
    # 之前的 30 分钟 TTL 导致切换平台时几乎必然全部缓存过期、重新抓取, 是卡顿主因。
    p = cache_path(kind, target_id)
    if not p.exists():
        return None
    if time.time() - p.stat().st_mtime > max_age:
        return None
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return None


def save_cache(kind, target_id, data):
    p = cache_path(kind, target_id)
    _atomic_write_text(p, json.dumps(data, ensure_ascii=False, separators=(",", ":")))


def _atomic_write_text(path: Path, text: str):
    """原子写入: 唯一临时名(pid+线程+随机)写后再 os.replace, 并发/崩溃不互相踩踏"""
    import threading
    tmp = path.with_name(
        f".{path.name}.{os.getpid()}.{threading.get_ident()}.{random.randint(0, 2 ** 31 - 1)}.tmp"
    )
    try:
        tmp.write_text(text, encoding="utf-8")
        os.replace(tmp, path)
    finally:
        if tmp.exists():
            try:
                tmp.unlink()
            except OSError:
                pass


# tanyu /summary 只支持三种统计口径, 全部忽略 startDate/endDate(已实测):
#   natural_day   = 值=昨天, 环比=前天
#   natural_week  = 值=本周, 环比=上周
#   natural_month = 值=本月, 环比=上月
# 店铺表现筛选因此按口径切换, 而非天数/自定义区间(那些对 summary 是空操作)。
SUMMARY_STAT_TYPES = ("natural_day", "natural_week", "natural_month")


def _valid_stat_type(stat_type):
    return stat_type in SUMMARY_STAT_TYPES


def _valid_date_iso(value):
    """校验 YYYY-MM-DD 格式, 合法返回原串, 非法返回 None"""
    try:
        import datetime as _dt
        _dt.date.fromisoformat(value)
        return value
    except (TypeError, ValueError):
        return None


def _valid_datetime_iso(value):
    """校验 YYYY-MM-DD 或 YYYY-MM-DDTHH:MM(秒级核算用), 合法返回原串, 非法返回 None"""
    try:
        import datetime as _dt
        _dt.datetime.fromisoformat(value)
        return value
    except (TypeError, ValueError):
        return None


def _split_bounds(start, end):
    """把核算区间归一化成 (start_ms, end_ms, start_day, end_day, has_time)

    start/end 接受 YYYY-MM-DD(纯日期) 或 YYYY-MM-DDTHH:MM(带时间, datetime-local 值)。
    end_ms 语义(唯一事实源, 下游禁止自己 +1day):
      纯日期 = 当天 23:59:59.999; 带时间 = 该时刻 +999ms。
    """
    import datetime as dt
    s = dt.datetime.fromisoformat(start)
    e = dt.datetime.fromisoformat(end)
    has_time = ("T" in start) or ("T" in end)
    start_ms = int(s.timestamp() * 1000)
    if has_time:
        end_ms = int(e.timestamp() * 1000) + 999
    else:
        end_ms = int((e + dt.timedelta(days=1)).timestamp() * 1000) - 1
    return start_ms, end_ms, s.date().isoformat(), e.date().isoformat(), has_time


def _check_span_limit(start, end, max_days=35):
    """校验自定义区间跨度 ≤ max_days(滚动窗口上限, 防参数滥用触发海量抓取)"""
    import datetime as dt
    try:
        s = dt.datetime.fromisoformat(start)
        e = dt.datetime.fromisoformat(end)
    except (TypeError, ValueError):
        return
    if (e - s).days > max_days:
        raise HTTPException(400, f"区间跨度不能超过 {max_days} 天: {start} ~ {end}")


def _stat_type_range(stat_type, ref=None):
    """统计口径对应的主值区间(与前端 statTypeRange 语义一致)。

    tanyu summary 三种口径忽略 startDate/endDate, 但明细表(section/table)
    对日期敏感; 未显式传 start/end 时按口径给默认区间, 使回显的
    startDate/endDate 不再是误导的 昨天~昨天:
      natural_day   = 昨天 ~ 昨天
      natural_week  = 本周一 ~ 周日(完整自然周)
      natural_month = 本月1日 ~ 今天
    ref 为基准日(默认今天), 便于测试/回溯历史周期。
    注意: 当前周 ref 为今天时 end 可能指向未来(本周日, 如周五的 08-09);
    该区间用于展示/缓存键, 出网 payload 的 endDate 须 clamp 到今天
    (见 shop_detail 的 table payload 构造处)。
    """
    ref = ref or datetime.date.today()
    one_day = datetime.timedelta(days=1)
    if stat_type == "natural_week":
        start = ref - datetime.timedelta(days=ref.weekday())  # weekday(): 周一=0
        end = ref + datetime.timedelta(days=6 - ref.weekday())  # 周日(完整自然周)
        return start.isoformat(), end.isoformat()
    if stat_type == "natural_month":
        start = ref.replace(day=1)
        return start.isoformat(), ref.isoformat()
    yest = ref - one_day
    return yest.isoformat(), yest.isoformat()


# ---------- 批量刷新 ----------
# 网络瞬时错误(SSL 断连/连接重置/超时): 刷新遇到时退避重试, 避免连打撞限速
_NETWORK_RETRIES = 3
_NETWORK_BASE_DELAY = 5.0


def _with_network_retry(fn, label, tag="refresh", on_retry=None):
    """网络瞬时错误指数退避重试(最多 3 次)

    仅对 requests 网络异常(SSLError/ConnectionError/Timeout)重试; 业务错误
    (风控/登录失效/接口错误码)不是 requests 网络异常, 原样抛出不重试。
    on_retry(attempt, delay): 每次重试前回调(供前端加载读条展示重试阶段)。
    """
    delay = _NETWORK_BASE_DELAY
    for attempt in range(_NETWORK_RETRIES):
        try:
            return fn()
        except (requests.exceptions.ConnectionError, requests.exceptions.Timeout) as e:
            if attempt == _NETWORK_RETRIES - 1:
                raise
            if on_retry:
                try:
                    on_retry(attempt + 1, delay)
                except Exception:
                    pass
            log_line(tag, f"{label} 网络错误(第 {attempt+1}/{_NETWORK_RETRIES} 次, "
                          f"{type(e).__name__}), {delay:.0f}s 后重试")
            time.sleep(delay)
            delay *= 2


def refresh_one(shop, start, end):
    """刷新单个店铺: summary + 3 个 section 明细, 返回汇总"""
    shop_id = shop["thirdShopId"]
    platform = shop.get("platform", 1)
    base = {
        "statType": "natural_day",
        "startDate": start,
        "endDate": end,
        "platform": platform,
        "dimension": "shop",
        "targetId": shop_id,
    }
    summary = _with_network_retry(lambda: fetch_summary(base),
                                  f"{shop['shopName']} summary")
    # 缓存键带口径后缀(与 shop_detail 一致): 批量刷新只刷 natural_day,
    # 写 {shop_id}__natural_day 键, 避免与周/月口径缓存互相覆盖
    save_cache("summary", f"{shop_id}__natural_day", {"fetched_at": time.time(), "data": summary})

    for section in ["operations", "service", "ai"]:
        payload = {**base, "section": section}
        try:
            table = _with_network_retry(lambda: fetch_section_table(payload),
                                        f"{shop['shopName']} section={section}")
            # 明细表按日期区间缓存(键带起止日期): 与 shop_detail 的读键严格一致,
            # 避免批量刷新(近7天)与单店自然日(昨天)写读同一键串日期范围。
            save_cache("table", f"{shop_id}__natural_day__{section}__{start}__{end}",
                       {"fetched_at": time.time(), "data": table})
        except Exception as e:
            log_line("refresh", f"{shop['shopName']} section={section} 失败: {e}")
    return shop_id


def refresh_all_async(start, end, shops=None):
    """后台线程: 刷新给定店铺列表(默认当前激活集团的店铺, 限抓取平台)

    shops: 可选店铺列表(跨集团刷新时由 refresh_all_groups_async 逐集团传入),
           为空时用 load_shops()(当前激活集团)过滤抓取平台。
    """

    def worker():
        with _lock:
            if _refresh_state["running"]:
                return
            _refresh_state["running"] = True
            _refresh_state["error"] = None
        try:
            # 只抓取启用的平台
            if shops is None:
                shops = [s for s in load_shops() if s.get("platform") in FETCH_PLATFORMS]
            total = len(shops)
            _refresh_state["progress"] = {"done": 0, "total": total, "current": ""}
            ok = fail = 0
            consecutive_fail = 0
            for i, shop in enumerate(shops, 1):
                _refresh_state["progress"]["current"] = f"{shop['platformName']} · {shop['shopName']} ({i}/{total})"
                if i > 1:
                    sleep_trace_shop()
                try:
                    refresh_one(shop, start, end)
                    ok += 1
                    consecutive_fail = 0
                except RiskTriggered as e:
                    # 风控/登录失效: 立即停止
                    log_line("refresh", f"⛔ 风控触发, 停止剩余 {total - i} 家店铺: {e}")
                    _refresh_state["error"] = f"风控/登录失效, 已停止: {e}"
                    break
                except BusyQueueError as e:
                    # 夜间抓取占用激活集团 cookie: 停止, 不写零值缓存
                    log_line("refresh", f"⏸ 夜间抓取进行中, 停止刷新: {e}")
                    _refresh_state["error"] = f"夜间抓取进行中, 请稍后重试: {e}"
                    break
                except Exception as e:
                    fail += 1
                    consecutive_fail += 1
                    log_line("refresh", f"{shop['shopName']} 失败: {e}")
                    if consecutive_fail >= 3:
                        # 连续失败说明 tanyu 侧异常(SSL 断连/限流), 整体退避 30s 再继续,
                        # 避免密集连打失败请求加剧限流
                        log_line("refresh", f"⏸ 连续 {consecutive_fail} 家失败, 暂停 30s 退避")
                        time.sleep(30)
                        consecutive_fail = 0
                _refresh_state["progress"]["done"] = i
            log_line("refresh", f"刷新完成: 成功 {ok} 家, 失败 {fail} 家")
            _refresh_state["last_run"] = time.time()
        except Exception as e:
            _refresh_state["error"] = str(e)
        finally:
            _refresh_state["running"] = False

    threading.Thread(target=worker, daemon=True).start()


def refresh_all_groups_async(start, end):
    """后台线程: 按集团切换轮询刷新全部三个集团(pdd/jd/douyin)的店铺 summary

    与夜间预抓同模式: 逐集团 switch_group → sync_shops_from_tanyu →
    refresh_all_async(该集团店铺) → 下一集团。结束后恢复进入前的激活集团,
    让常驻看板继续服务原集团。风控/登录失效立即整体停止。
    """
    cfg = load_config()
    platform_order = cfg.get("prefetch_platforms") or [1, 5, 7]
    plat_rank = {p: i for i, p in enumerate(platform_order)}
    ordered = sorted(
        [g for g in cfg.get("groups", []) if g.get("groupId") and g.get("accountId")],
        key=lambda g: plat_rank.get(g.get("platform"), 99),
    )
    if not ordered:
        log_line("refresh", "⚠️ config.groups 为空或无可用集团, 仅刷新当前激活集团")
        refresh_all_async(start, end)
        return
    # 记录进入前的激活集团, 结束后切回
    orig_gid = cfg.get("cookies", {}).get("tanyu-group-id")

    def worker():
        try:
            for g in ordered:
                if _risk_state.get("triggered"):
                    log_line("refresh", "⛔ 风控/登录失效, 整体停止")
                    break
                try:
                    switch_group(g["groupId"])
                except RiskTriggered:
                    log_line("refresh", "⛔ 风控触发于切换集团, 整体停止")
                    _refresh_state["error"] = _risk_state.get("reason") or "风控/登录失效"
                    return
                except Exception as e:
                    log_line("refresh", f"⚠️ 集团「{g.get('groupName')}」切换失败, 跳过: {e}")
                    continue
                group_shops = [s for s in load_shops() if s.get("platform") in FETCH_PLATFORMS]
                if not group_shops:
                    log_line("refresh", f"集团「{g.get('groupName')}」无抓取平台店铺, 跳过")
                    continue
                # 用该集团店铺子集跑一次刷新(不带新线程, 直接在当前 worker 内循环)
                # running 已由端点同步置位; 单并发由队列保证, worker 内不再自检
                try:
                    total = len(group_shops)
                    _refresh_state["progress"] = {"done": 0, "total": total, "current": ""}
                    ok = fail = 0
                    consecutive_fail = 0
                    for i, shop in enumerate(group_shops, 1):
                        _refresh_state["progress"]["current"] = (
                            f"集团[{g.get('groupName')}] {shop['platformName']} · "
                            f"{shop['shopName']} ({i}/{total})")
                        if i > 1:
                            sleep_trace_shop()
                        try:
                            refresh_one(shop, start, end)
                            ok += 1
                            consecutive_fail = 0
                        except RiskTriggered as e:
                            log_line("refresh", f"⛔ 风控触发于集团[{g.get('groupName')}], 停止: {e}")
                            _refresh_state["error"] = f"风控/登录失效, 已停止: {e}"
                            return
                        except BusyQueueError as e:
                            # 夜间抓取占用激活集团 cookie: 停止, 不写零值缓存
                            log_line("refresh", f"⏸ 夜间抓取进行中, 停止刷新: {e}")
                            _refresh_state["error"] = f"夜间抓取进行中, 请稍后重试: {e}"
                            return
                        except Exception as e:
                            fail += 1
                            consecutive_fail += 1
                            log_line("refresh", f"{shop['shopName']} 失败: {e}")
                            if consecutive_fail >= 3:
                                # 连续失败说明 tanyu 侧异常(SSL 断连/限流), 整体退避 30s 再继续,
                                # 避免密集连打失败请求加剧限流
                                log_line("refresh", f"⏸ 连续 {consecutive_fail} 家失败, 暂停 30s 退避")
                                time.sleep(30)
                                consecutive_fail = 0
                        _refresh_state["progress"]["done"] = i
                    log_line("refresh", f"集团[{g.get('groupName')}] 刷新完成: 成功 {ok} 家, 失败 {fail} 家")
                    _refresh_state["last_run"] = time.time()
                except Exception as e:
                    log_line("refresh", f"集团[{g.get('groupName')}] 刷新异常: {e}")
            log_line("refresh", "全部集团刷新完成")
        finally:
            # 先恢复原激活集团(耗时数秒), 完成后再清 running —— 恢复期间保持
            # running=True, 防止新 worker 在恢复窗口内启动互踩集团 cookie
            try:
                if orig_gid and not _risk_state.get("triggered"):
                    if orig_gid != load_config().get("cookies", {}).get("tanyu-group-id"):
                        # invalidate=False: 恢复集团不清核算缓存(刷新任务不影响已算好的结果)
                        switch_group(orig_gid, invalidate=False)
                        log_line("refresh", f"已恢复原激活集团 {orig_gid}")
                elif _risk_state.get("triggered"):
                    log_line("refresh", "⚠️ 风控触发, 未恢复原集团(需重新登录后手工切换)")
            except Exception as e:
                log_line("refresh", f"⚠️ 恢复原集团失败: {e}")
            finally:
                with _lock:
                    _refresh_state["running"] = False

    threading.Thread(target=worker, daemon=True).start()


# ---------- 集团管理 ----------
GC_API = "https://agent.tanyuai.com/api/gc"

# accountType -> 平台/类型名(集团标签)
ACCOUNT_TYPE_NAMES = {3: "客服智能体"}


def get_current_group():
    """获取当前激活集团(来自 detail 接口)"""
    try:
        _assert_no_risk()
        _rate_limit()
        r = requests.get(f"{GC_API}/agent-personal/detail", headers=get_headers(), timeout=15)
        d = r.json()
        if _check_risk(r, d):
            return None
        if d.get("success") is False:
            return None
        g = (d.get("data") or {}).get("group") or {}
        return {"id": g.get("id"), "name": g.get("name"), "chatbotGroupId": g.get("chatbotGroupId")}
    except Exception:
        return None


def get_group_list():
    """已发现的集团列表(config.groups) + 当前集团"""
    cfg = load_config()
    groups = cfg.get("groups") or []
    cur = get_current_group()
    if cur and cur["id"]:
        # 当前集团不在列表时自动补上(锁内读改写, 防并发写者丢更新)
        if not any(g.get("groupId") == cur["id"] for g in groups):
            new_g = {"groupId": cur["id"], "groupName": cur["name"],
                     "accountId": None, "accountType": None, "current": True}
            new_g["platform"] = _infer_group_platform(new_g)  # 新集团落盘带平台, 避免再走推断
            groups = [new_g] + groups
            mutate_config(lambda c: c.__setitem__("groups", groups))
        for g in groups:
            g["current"] = g.get("groupId") == cur["id"]
        # 给 current 对象补上平台字段(前端集团→平台联动不再需要从列表猜测)
        cur_g = next((g for g in groups if g.get("current")), None)
        if cur_g and "platform" in cur_g and cur["id"] == cur_g.get("groupId"):
            cur["platform"] = cur_g["platform"]
    return {"groups": groups, "current": cur}


def _parse_cookie_expiry_from_header(set_cookie_str, now_ts):
    """从单条 Set-Cookie 头串解析 cookie 到期日

    格式形如: 'tanyu-agent-account=xxx; Max-Age=2592000; Expires=Sun, 6 Sep 2026 11:27:42 +0800; ...'
    优先取 Max-Age(相对当前时刻), 回退 Expires(RFC1123)。返回 (cookie名, 'YYYY-MM-DD') 或 None。
    """
    if not set_cookie_str:
        return None
    parts = set_cookie_str.split(";")
    if not parts:
        return None
    first = parts[0].strip()
    if "=" not in first:
        return None
    name = first.split("=", 1)[0].strip()
    if not name.startswith("tanyu-"):
        return None
    max_age = None
    expires = None
    for p in parts[1:]:
        p = p.strip()
        if p.lower().startswith("max-age="):
            try:
                max_age = float(p.split("=", 1)[1].strip())
            except Exception:
                pass
        elif p.lower().startswith("expires="):
            expires = p.split("=", 1)[1].strip()
    when = None
    if max_age is not None:
        when = now_ts + max_age
    elif expires:
        # RFC1123 / 变体(含时区偏移, 如 'Sun, 6 Sep 2026 11:27:42 +0800')
        try:
            import email.utils
            parsed = email.utils.parsedate_to_datetime(expires)
            when = parsed.timestamp()
        except Exception:
            try:
                from datetime import datetime
                parsed = datetime.strptime(expires, "%a, %d %b %Y %H:%M:%S GMT")
                when = parsed.timestamp()
            except Exception:
                when = None
    if when is None:
        return None
    import datetime as _dt
    return name, _dt.date.fromtimestamp(when).isoformat()


def _capture_cookie_expiry(resp):
    """从 Set-Cookie 响应头解析各 cookie 到期日, 写入 config.cookie_expires(仅日期, 无 cookie 值)

    switch-group-by-wx 对 tanyu-agent-account 带 Max-Age=2592000(30天), tanyu-account-id/
    tanyu-group-id 是 SESSION(无 Max-Age/Expires)不记录。requests 的 cookie jar 会丢弃
    Max-Age/Expires 属性, 故直接解析原始 Set-Cookie 响应头。
    只存 'YYYY-MM-DD' 日期字符串, 绝不存 cookie 值/凭据。
    """
    try:
        expires_map = dict((load_config().get("cookie_expires") or {}))
        changed = False
        now = time.time()
        # 可能有多条 Set-Cookie 头: 用底层 getlist 逐个解析
        try:
            headers_list = resp.raw.headers.getlist("Set-Cookie")
        except Exception:
            headers_list = [resp.headers.get("Set-Cookie", "")] if resp.headers.get("Set-Cookie") else []
        for sc in headers_list:
            parsed = _parse_cookie_expiry_from_header(sc, now)
            if parsed:
                name, day = parsed
                if expires_map.get(name) != day:
                    expires_map[name] = day
                    changed = True
        if changed:
            # 锁内读改写落盘, 不覆盖其他写者的更新
            mutate_config(lambda c: c.update({"cookie_expires": expires_map}))
    except Exception as e:
        log_line("auth", f"cookie 到期时间解析失败(不影响使用): {e}")


def switch_group(group_id, invalidate=True):
    """切换到指定集团: 调 switch-group-by-wx 捕获 Set-Cookie 更新 config

    invalidate=False: 跳过同步对核算缓存的失效(worker 自己恢复原集团时用,
    否则刚算好的核算结果会在收尾时被 _invalidate_audit_caches 清空)。
    """
    cfg = load_config()
    groups = cfg.get("groups") or []
    g = next((x for x in groups if x.get("groupId") == group_id), None)
    if not g:
        raise ValueError(f"集团不存在: {group_id}")
    if not g.get("accountId"):
        raise ValueError(f"集团「{g.get('groupName')}」缺少 accountId, 无法切换(请通过扫码登录重新发现集团)")
    _assert_no_risk()
    _rate_limit()
    resp = requests.post(
        f"{GC_API}/agent/auth/switch-group-by-wx",
        json={"accountId": g["accountId"], "accountType": g.get("accountType", 3)},
        headers=get_headers(), timeout=15,
    )
    d = resp.json()
    if _check_risk(resp, d):
        raise RuntimeError(f"风控/登录失效: {d.get('msg', resp.status_code)}")
    if d.get("success") is False:
        raise RuntimeError(d.get("msg", "切换集团失败"))
    # 捕获 Set-Cookie 并保存(锁内读改写, 防与其他 config 写者丢更新)
    mutate_config(lambda c: c["cookies"].update(
        {n: v for n, v in resp.cookies.items() if n.startswith("tanyu-")}))
    # 记录各 cookie 到期日(仅日期, 无 cookie 值)
    _capture_cookie_expiry(resp)
    # 同步店铺列表
    sync_shops_from_tanyu(invalidate=invalidate)
    log_line("group", f"已切换集团「{g.get('groupName')}」")
    return {"ok": True, "group": get_current_group(), "updated": True}


def _invalidate_audit_caches():
    """店铺列表变化后清空核算相关缓存(总览磁盘缓存 + 内存态), 防止跨集团串数据"""
    global _trace_state
    try:
        if TRACE_OVERVIEW_CACHE_FILE.exists():
            TRACE_OVERVIEW_CACHE_FILE.unlink()
    except Exception:
        pass
    with _lock:
        _trace_state["result"] = None
        _trace_state["subset_result"] = None  # 导入/换集团后子集核算快照一并失效
        _trace_state["start_date"] = None
        _trace_state["end_date"] = None
        _trace_state["platform"] = None
        _trace_state["last_run"] = None
        _trace_state["error"] = None
        _trace_state["partial_list"] = []
    # 注意: 不重置 paused/canceled。若 sync 已把运行中的核算置为 canceled=True,
    # 这里保留该标志, 让旧 worker 在下个店铺边界退出(取消是终态, 下次开始核算时复位);
    # paused 由 worker 的 finally 清掉。event.set() 唤醒阻塞在暂停等待中的 worker 使其退出。
    _trace_resume_evt.set()
    # 进程内共享的 trace_days 解析缓存按 shop_id 存, 新集团店铺 ID 不冲突, 无需清空;
    # 但旧集团文件留在磁盘无害(不会被新店铺 ID 读到)。如需回收可删 trace_days 目录。


# 店铺同步可靠性: 同集团短时间重复切换复用上次结果, 减少多人频繁切换对 tanyu brief 的请求
SHOPS_SYNC_TTL = 60            # 秒
_shops_sync_cache = {}         # {groupId: {"ts": float, "count": int}}
# 切换集团时店铺同步阶段(供前端加载读条): phase=syncing/retry/fallback/done
_group_sync_state = {"phase": "idle", "attempt": 0, "detail": "", "updated": 0}


def _set_group_sync(phase, attempt=0, detail=""):
    """更新店铺同步阶段(前端轮询 /api/groups/sync-status 展示加载读条)"""
    global _group_sync_state
    _group_sync_state = {"phase": phase, "attempt": attempt,
                         "detail": detail, "updated": time.time()}


def sync_shops_from_tanyu(invalidate=True):
    """从探域 brief 接口抓取当前集团的全部店铺, 写入 shops.json

    可靠性: brief 请求带网络退避重试; 重试耗尽/接口异常时用本地 SQLite 该集团
    平台店铺兜底(shops.json 切到该集团店铺), 不抛异常 —— 多人频繁切换 + tanyu
    限流/断连下切换平台也不会有"加载失败"。同集团 SHOPS_SYNC_TTL 内重复切换
    复用上次结果不发请求(降限速)。风控/登录失效/夜间抓取占用不兜底, 仍抛异常。

    invalidate=False: 跳过"取消运行中核算 + 清核算缓存"副作用(worker 内部
    恢复原集团时用, 保留刚完成的核算结果)。用户/其他任务发起的切换保持默认
    True, 取消旧 worker 防跨集团串数据。
    """
    # 集团/店铺集合将变化: 先取消进行中的核算(店铺数据即将失效, 旧 worker 结果会串组),
    # 避免旧 worker 继续抓取旧集团店铺。取消保留已完成店铺的部分结果供前端展示。
    if invalidate:
        with _lock:
            if _trace_state["running"]:
                _trace_state["canceled"] = True
                _trace_state["paused"] = False
        if _trace_state.get("canceled"):
            _trace_resume_evt.set()  # 若 worker 阻塞在暂停等待中, 唤醒使其退出
    cfg = load_config()
    gid = cfg.get("cookies", {}).get("tanyu-group-id")
    g = next((x for x in cfg.get("groups", []) if x.get("groupId") == gid), None)
    plat = g.get("platform") if g else None
    # TTL 缓存: 同集团短时间重复切换不发 brief 请求(多人切换降限速)
    hit = _shops_sync_cache.get(gid)
    if hit and time.time() - hit["ts"] < SHOPS_SYNC_TTL:
        shops = load_all_shops(plat)
        if shops:
            _atomic_write_text(SHOPS_FILE, json.dumps({"data": shops}, ensure_ascii=False, indent=2))
            if invalidate:
                _invalidate_audit_caches()
            _set_group_sync("done", 0, f"TTL 缓存 {len(shops)} 家")
            log_line("group", f"店铺同步(TTL 缓存): {len(shops)} 家")
            return {"count": len(shops),
                    "platforms": {p: sum(1 for s in shops if s["platform"] == p) for p in set(s["platform"] for s in shops)},
                    "source": "cache"}
    try:
        _assert_no_risk()
        _rate_limit()
        _set_group_sync("syncing", 0, "正在同步店铺列表…")
        try:
            r = _with_network_retry(
                lambda: requests.get(f"{GC_API}/agent-personal/brief?searchValid=true",
                                     headers=get_headers(), timeout=20),
                "brief 店铺列表", tag="group",
                on_retry=lambda n, d: _set_group_sync(
                    "retry", n, f"网络波动, 正在自动重试(第 {n}/{_NETWORK_RETRIES} 次, {d:.0f}s 后)…"))
            d = r.json()
        except (requests.exceptions.RequestException, ValueError) as e:
            return _sync_shops_fallback(e, invalidate=invalidate)
        if _check_risk(r, d):
            raise RuntimeError(f"风控/登录失效: {d.get('msg', r.status_code)}")
        if d.get("success") is False:
            return _sync_shops_fallback(RuntimeError(d.get("msg", "店铺列表获取失败")), invalidate=invalidate)
        data = d.get("data") or []
        if not isinstance(data, list):
            data = data.get("chatbotShops") or []
        shops = []
        for s in data:
            if not s.get("thirdShopId"):
                continue
            shops.append({
                "thirdShopId": s["thirdShopId"],
                "shopName": s.get("shopName") or "未知",
                "platform": s.get("platform", 0),
                "sellerId": s.get("sellerId", ""),
                "ifAgentReceipt": s.get("ifAgentReceipt", False),
            })
        if not shops:
            return _sync_shops_fallback(RuntimeError("接口返回空店铺列表"))
        for s in shops:
            s["platformName"] = PLATFORM_NAMES.get(s.get("platform"), str(s.get("platform")))
        # 原子写 shops.json: 与 8080 常驻并发写时互不读到半截 JSON(直接 write_text
        # 会截断写, 另一进程 load_shops 可能读到空/半截列表)
        _atomic_write_text(SHOPS_FILE, json.dumps({"data": shops}, ensure_ascii=False, indent=2))
        # 店铺集合已变: 同步 SQLite 店铺维度表, 并清掉核算缓存, 避免旧集团结果串到新集团
        try:
            trace_store.upsert_shops(shops)
        except Exception as e:
            log_line("db", f"shops 同步失败: {e}")
        if invalidate:
            _invalidate_audit_caches()
        _shops_sync_cache[gid] = {"ts": time.time(), "count": len(shops)}
        _set_group_sync("done", 0, f"同步完成 {len(shops)} 家")
        log_line("group", f"店铺同步完成: {len(shops)} 家")
        return {"count": len(shops),
                "platforms": {p: sum(1 for s in shops if s["platform"] == p) for p in set(s["platform"] for s in shops)}}
    except Exception as e:
        raise RuntimeError(f"同步店铺失败: {e}")


def _sync_shops_fallback(reason, invalidate=True):
    """brief 同步失败降级: 用本地 SQLite 该集团平台店铺兜底, 切换不失败

    仅用于网络瞬时错误/接口返回空等可降级场景(风控/夜间占用不进来)。
    兜底同样写 shops.json, 保证 load_shops 读到当前集团的店铺, 且核算缓存失效
    不串旧集团结果。新增店铺延迟到下次成功同步出现。本地也无该集团店铺时真失败。
    """
    try:
        cfg = load_config()
        gid = cfg.get("cookies", {}).get("tanyu-group-id")
        g = next((x for x in cfg.get("groups", []) if x.get("groupId") == gid), None)
        plat = g.get("platform") if g else None
        shops = load_all_shops(plat)
        if not shops:
            raise RuntimeError(f"同步店铺失败且本地无该集团缓存店铺: {reason}")
        _atomic_write_text(SHOPS_FILE, json.dumps({"data": shops}, ensure_ascii=False, indent=2))
        if invalidate:
            _invalidate_audit_caches()
        _set_group_sync("done", 0, f"网络异常, 已用本地缓存 {len(shops)} 家")
        log_line("group", f"⚠️ 同步店铺失败({type(reason).__name__}), 用本地缓存 {len(shops)} 家兜底"
                          f"(新增店铺将在下次成功同步出现)")
        return {"count": len(shops),
                "platforms": {p: sum(1 for s in shops if s["platform"] == p) for p in set(s["platform"] for s in shops)},
                "source": "fallback"}
    except RuntimeError:
        raise
    except Exception as e:
        raise RuntimeError(f"同步店铺失败: {reason}; 本地兜底也失败: {e}")


# ---------- 客服昵称同步 ----------
def _sync_staff_names_worker(start_day, end_day):
    """后台线程: 逐集团逐店抓 customer-service/table, 提取 serviceAccount → 昵称/店铺名

    时序与 refresh_all_groups_async 相同: 逐集团 switch_group → 该集团店铺 →
    逐店 POST customer-service/table(section=service) → 下一集团。结束后恢复进入前的
    激活集团。风控/登录失效立即整体停止。结果写 data/staff_names.json(无 cookie 值)。
    """
    cfg = load_config()
    platform_order = cfg.get("prefetch_platforms") or [1, 5, 7]
    plat_rank = {p: i for i, p in enumerate(platform_order)}
    ordered = sorted(
        [g for g in cfg.get("groups", []) if g.get("groupId") and g.get("accountId")],
        key=lambda g: plat_rank.get(g.get("platform"), 99),
    )
    if not ordered:
        log_line("staff", "⚠️ config.groups 为空, 无可同步的客服昵称")
        _staff_sync_state["error"] = "无可用集团"
        _staff_sync_state["running"] = False
        return
    orig_gid = cfg.get("cookies", {}).get("tanyu-group-id")
    name_map = {}
    total_plan = 0
    group_plan = []
    try:
        for g in ordered:
            if _risk_state.get("triggered"):
                log_line("staff", "⛔ 风控/登录失效, 客服昵称同步整体停止")
                _staff_sync_state["error"] = _risk_state.get("reason") or "风控/登录失效"
                return
            try:
                switch_group(g["groupId"])
            except RiskTriggered:
                log_line("staff", "⛔ 风控触发于切换集团, 整体停止")
                _staff_sync_state["error"] = _risk_state.get("reason") or "风控/登录失效"
                return
            except Exception as e:
                log_line("staff", f"⚠️ 集团「{g.get('groupName')}」切换失败, 跳过: {e}")
                continue
            # 该集团店铺(SQLite 表含全部集团店铺, 按集团平台过滤, 复用 _today_target_shops 修复逻辑)
            try:
                shops = [s for s in trace_store.get_shops() if s.get("platform") == g.get("platform")]
            except Exception:
                shops = []
            if not shops:
                log_line("staff", f"集团「{g.get('groupName')}」无抓取平台店铺, 跳过")
                continue
            group_plan.append((g, shops))
            total_plan += len(shops)
        if not group_plan:
            # 全集团无店铺可同步: 打印 SQLite shops 平台分布 + 配置期望平台,
            # 便于区分"shops 表确实空/缺平台" vs "platform 类型不匹配"(如手改 config 成字符串)
            try:
                dist = {}
                for s in trace_store.get_shops():
                    dist[s.get("platform")] = dist.get(s.get("platform"), 0) + 1
                expect = [(g.get("groupName"), g.get("platform"), type(g.get("platform")).__name__)
                          for g in ordered]
                log_line("staff", f"⚠️ 无任何集团店铺可同步 (SQLite shops 平台分布: {dist}, "
                                  f"配置期望: {expect})")
            except Exception:
                log_line("staff", "⚠️ 无任何集团店铺可同步")
            _staff_sync_state["error"] = "无店铺可同步"
            return
        _staff_sync_state["progress"] = {"done": 0, "total": total_plan, "current": ""}
        done = 0
        ok = fail = 0
        consecutive_fail = 0
        for g, shops in group_plan:
            if _risk_state.get("triggered"):
                log_line("staff", "⛔ 风控/登录失效, 同步停止")
                _staff_sync_state["error"] = _risk_state.get("reason") or "风控/登录失效"
                return
            for i, shop in enumerate(shops, 1):
                if _risk_state.get("triggered"):
                    log_line("staff", "⛔ 风控/登录失效, 同步停止")
                    _staff_sync_state["error"] = _risk_state.get("reason") or "风控/登录失效"
                    return
                _staff_sync_state["progress"]["current"] = (
                    f"集团[{g.get('groupName')}] {shop.get('platformName', '')} · "
                    f"{shop.get('shopName', '')} ({done + 1}/{total_plan})")
                if done > 0 or i > 1:
                    sleep_trace_shop()
                try:
                    # 网络瞬时错误(SSL 断连/限流)指数退避重试, 与刷新/补抓同策略;
                    # on_retry 把"自动重试中"写回进度, 前端可见
                    rows = _with_network_retry(
                        lambda: _fetch_staff_table_rows(shop, start_day, end_day),
                        f"{shop.get('shopName', '')} 客服表", tag="staff",
                        on_retry=lambda n, d: _staff_sync_state["progress"].update(
                            current=f"集团[{g.get('groupName')}] {shop.get('shopName','')} · "
                                    f"网络波动, 自动重试(第 {n}/{_NETWORK_RETRIES} 次)…"),
                    )
                    plat = shop.get("platform")
                    for r in rows:
                        acct = r.get("serviceAccount")
                        # 跳过汇总行与无冒号的异常条目(客服账号必有 ':' 分隔,
                        # 无冒号的 'cs_<sellerId>' 是接口返回的非客服行, 永不匹配 byStaff)
                        if not acct or acct == "__SUMMARY__" or ":" not in acct:
                            continue
                        sv = r.get("serviceName") or ""
                        nick, shop_nm = _parse_staff_service_name(sv, plat)
                        name_map[acct] = {
                            "nick": nick,
                            "shopName": shop_nm,
                            "serviceName": sv,
                            "platform": plat,
                        }
                    ok += 1
                    consecutive_fail = 0
                except RiskTriggered as e:
                    log_line("staff", f"⛔ 风控触发于 {shop.get('shopName', '')}: {e}")
                    _staff_sync_state["error"] = f"风控/登录失效, 已停止: {e}"
                    return
                except Exception as e:
                    fail += 1
                    consecutive_fail += 1
                    log_line("staff", f"{shop.get('shopName', '')} 客服昵称抓取失败: {e}")
                    if consecutive_fail >= 3:
                        # 连续失败说明 tanyu 侧异常(SSL 断连/限流), 整体退避 30s 再继续,
                        # 避免密集连打失败请求加剧限流
                        log_line("staff", f"⏸ 连续 {consecutive_fail} 家失败, 暂停 30s 退避")
                        time.sleep(30)
                        consecutive_fail = 0
                done += 1
                _staff_sync_state["progress"]["done"] = done
        # 落盘(带抓取时间戳; 绝不写 cookie)
        _atomic_write_text(STAFF_NAMES_FILE, json.dumps(
            {"fetched_at": time.time(), "map": name_map}, ensure_ascii=False, indent=2))
        _staff_sync_state["last_run"] = time.time()
        log_line("staff", f"客服昵称同步完成: 成功 {ok} 家, 失败 {fail} 家, 共 {len(name_map)} 个客服账号")
    except Exception as e:
        _staff_sync_state["error"] = str(e)
    finally:
        # 先恢复原激活集团(耗时数秒), 完成后再清 running(防止恢复窗口内双启动)
        try:
            if orig_gid and not _risk_state.get("triggered"):
                try:
                    if orig_gid != load_config().get("cookies", {}).get("tanyu-group-id"):
                        # invalidate=False: 恢复集团不清核算缓存(客服同步不影响已算好的结果)
                        switch_group(orig_gid, invalidate=False)
                        log_line("staff", f"已恢复原激活集团 {orig_gid}")
                except Exception as e:
                    log_line("staff", f"⚠️ 恢复原集团失败: {e}")
        finally:
            _staff_sync_state["running"] = False


def _fetch_staff_table_rows(shop, start_day, end_day):
    """拉单个店铺的客服表(section=service, 实际客服数据; operations 全 0 无数据)"""
    payload = {
        "startDate": start_day,
        "endDate": end_day,
        "platform": shop.get("platform"),
        "shopId": shop["thirdShopId"],
        "section": "service",
        "metricKeys": ["service_consult_cnt"],
    }
    data = post_api("/customer-service/table", payload)
    rows = (data or {}).get("rows") or []
    return rows


def _parse_staff_service_name(sv, platform):
    """从 serviceName 拆出客服昵称与店铺展示名

    拼多多(1): '易方配件专营店:哆啦A梦' -> (哆啦A梦, 易方配件专营店)
    抖音(5)/京东(7): '小罗' -> (小罗, '')  (serviceName 仅昵称)
    """
    if not sv:
        return "", ""
    if ":" in sv:
        a, b = sv.split(":", 1)
        if not b.strip().isdigit() and b.strip():
            # 冒号前是店铺展示名(拼多多), 冒号后是昵称
            return b.strip(), a.strip()
    return sv.strip(), ""


def sync_staff_names_from_tanyu(triggered_by=None):
    """触发客服昵称同步(异步后台线程); 互斥检查: 核算/刷新/今日进行中不可同时跑"""
    try:
        _assert_no_running_task()
    except RuntimeError as e:
        raise RuntimeError(f"其他任务进行中, 请稍后再试: {e}")
    with _lock:
        if _staff_sync_state.get("running"):
            raise RuntimeError("客服昵称同步任务进行中")
        _staff_sync_state["running"] = True
        _staff_sync_state["error"] = None
        _staff_sync_state["triggered_by"] = triggered_by
    # 用"昨天~昨天"(客服表按天, 一天足矣; 该接口反映当前客服成员)
    end_day = (datetime.date.today() - datetime.timedelta(days=1)).isoformat()
    threading.Thread(target=_sync_staff_names_worker, args=(end_day, end_day), daemon=True).start()
    return {"ok": True, "message": "客服昵称同步已启动"}


@app.get("/api/staff/names")
def staff_names():
    """客服昵称映射: {fetched_at, map, count}(供客服池渲染昵称/店铺名)"""
    data = _load_staff_names()
    if not data:
        return {"fetched_at": None, "map": {}, "count": 0}
    name_map = data.get("map") or {}
    return {"fetched_at": data.get("fetched_at"), "map": name_map, "count": len(name_map)}


@app.get("/api/staff/shops")
def staff_shops(platform: int | None = None):
    """跨集团全量店铺(客服池店铺筛选下拉的数据源, 不依赖当前激活集团)

    与 /api/shops 不同: /api/shops 读 shops.json(当前激活集团), 客服池查询走
    SQLite 跨集团表(trace_store.get_shops), 下拉必须与该作用域一致。
    """
    try:
        shops = trace_store.get_shops()
    except Exception:
        shops = []
    for s in shops:
        s["platformName"] = PLATFORM_NAMES.get(s.get("platform"), str(s.get("platform")))
    if platform is not None:
        shops = [s for s in shops if s.get("platform") == platform]
    shops.sort(key=lambda s: (s.get("platformName") or "", s.get("shopName") or ""))
    return {"shops": shops}


@app.post("/api/staff/names/sync")
def staff_names_sync(request: Request):
    """手动触发客服昵称同步(异步); 多用户队列: 忙时入队排队"""
    _task_label = "同步客服昵称"
    _task_params = {}
    _queued = _queue_if_busy(request, "staff_sync", _task_label, _task_params)
    if _queued:
        return _queued
    with _lock:
        if _any_task_running():
            if request.headers.get("x-scheduler-task") == "1":
                raise HTTPException(409, "任务进行中, 调度器稍后重试")
            _task = _enqueue_task("staff_sync", _task_label, _operator_label(request), _task_params, _operator_role(request))
            return {"status": "queued", "taskId": _task["id"],
                    "queuePosition": _queue_position(_task["id"]), "task": _task_public(_task)}
    try:
        r = sync_staff_names_from_tanyu(_operator_label(request))
    except RuntimeError as e:
        raise HTTPException(409, str(e))
    # 直接启动(非调度器)也登记任务列表: 前端显示"谁发起的 + 进度条"
    if request.headers.get("x-scheduler-task") != "1":
        _register_running_task(_staff_sync_state, "staff_sync", _task_label,
                               _operator_label(request), _task_params, _operator_role(request))
    return r


@app.get("/api/staff/names/status")
def staff_names_status(request: Request):
    """客服昵称同步任务状态"""
    with _lock:
        st = dict(_staff_sync_state)
    if st.get("triggered_by"):
        st["triggered_by"] = _mask_admin_label(request, st["triggered_by"])
    return st


# ---------- 风控保护 ----------
import random

# 请求节奏(秒): 用随机区间而非固定值, 接近人工操作习惯
RISK_PAGE_INTERVAL = (0.4, 1.2)    # 消息分页之间随机等待
RISK_SHOP_INTERVAL = (0.8, 2.5)    # 店铺之间随机等待
RISK_MAX_RPM = 30                  # 全局限速: 每分钟最多请求数
RISK_CACHE_HOURS = 24              # 核算缓存有效期(同时间段重复核算零请求)
RISK_KEYWORDS = ["登录", "过期", "验证", "风控", "频繁", "操作过频",
                 "forbidden", "unauthorized", "rate limit", "restrict"]

# 风控全局状态: 一旦触发, 所有抓取任务立即停止, 需重新登录后手动清除
_risk_state = {"triggered": False, "reason": "", "at": None, "last_code": None}
_request_times = []  # 滑动窗口请求时间戳


class RiskTriggered(RuntimeError):
    """风控/登录失效信号, 触发后任务立即停止; 继承 RuntimeError 使各端点
    except RuntimeError 能统一兜底(返回 502/503), 避免漏网 500"""


class TraceTruncatedError(RuntimeError):
    """区间内消息数超过 tanyu 接口单区间硬上限且细分后仍触顶:
    返回的数据不完整, 调用方不得将其落缓存/落库(避免残缺天被永久固化)"""


def _set_risk(reason, code=None):
    need_push = False
    with _lock:
        if not _risk_state["triggered"]:
            _risk_state["triggered"] = True
            _risk_state["reason"] = reason
            _risk_state["at"] = time.time()
            _risk_state["last_code"] = code
            need_push = True
    log_line("risk", f"⚠️ 检测到风控/登录失效信号: {reason} — 已停止所有抓取任务, 请重新登录")
    if need_push:
        # 首次触发(非重复): 写系统通知(页面铃铛/横幅), 不外推钉钉
        _notify_risk_alert(reason, code)


# ---------- 风控/登录失效系统通知 ----------
# 场景: 抓取过于频繁触发探域风控 / 微信账号被踢下线 / 密码被改 → 登录态失效,
# 所有请求开始报 401/403/429 或会话错误码 → _set_risk 全局止损 → 写入系统通知
# (页面右上角铃铛 + 风控横幅), 不外推钉钉。
NOTIFY_FILE = DATA_DIR / "notifications.json"
NOTIFY_MAX = 100            # 通知保留条数(超出滚动删除最旧)
_notify_lock = threading.Lock()


def _notify_load():
    try:
        if NOTIFY_FILE.exists():
            d = json.loads(NOTIFY_FILE.read_text(encoding="utf-8"))
            if isinstance(d, dict) and isinstance(d.get("items"), list):
                return d["items"]
    except Exception:
        pass
    return []


def _notify_save(items):
    _atomic_write_text(NOTIFY_FILE, json.dumps({"items": items[-NOTIFY_MAX:]},
                                               ensure_ascii=False, separators=(",", ":")))


def _notify_add(level, title, body):
    """写入一条系统通知; 永不抛异常"""
    try:
        with _notify_lock:
            items = _notify_load()
            items.append({
                "id": f"n{int(time.time() * 1000)}",
                "level": level,          # danger / success / info
                "title": title,
                "body": body,
                "at": time.time(),
                "read": False,
            })
            _notify_save(items)
    except Exception as e:
        log_line("risk", f"写通知失败: {e}")


def _notify_risk_alert(reason, code=None):
    """风控/登录失效(微信账号被登出)时写系统通知"""
    try:
        t = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cur = None
        try:
            cfg = load_config()
            gid = cfg.get("cookies", {}).get("tanyu-group-id")
            g = next((x for x in cfg.get("groups", []) if x.get("groupId") == gid), None)
            cur = g.get("groupName") if g else None
        except Exception:
            pass
        body = (
            f"时间: {t}\n"
            f"原因: {reason}" + (f" (code={code})" if code else "") + "\n"
            f"当前集团: {cur or '未知'}\n"
            f"影响: 所有抓取/核算任务已停止, 新数据不再入库\n"
            f"处理: 系统设置 → 微信扫码登录(二维码), 扫码后自动续期 Cookie 并恢复抓取。"
        )
        _notify_add("danger", "探域登录失效/风控告警", body)
        log_line("risk", "已写入系统通知: 登录失效/风控告警")
    except Exception as e:
        log_line("risk", f"风控通知写入异常: {e}")


def _notify_risk_recovered():
    """重新登录成功(风控解除)时写系统通知"""
    try:
        t = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        _notify_add("success", "探域登录已恢复",
                    f"时间: {t}\nCookie 已更新, 抓取/核算任务可正常运行。")
        log_line("risk", "已写入系统通知: 登录已恢复")
    except Exception as e:
        log_line("risk", f"恢复通知写入异常: {e}")


def _assert_no_risk():
    if _risk_state["triggered"]:
        raise RiskTriggered(_risk_state["reason"])


def _reset_risk():
    was_triggered = False
    with _lock:
        was_triggered = _risk_state["triggered"]
        _risk_state["triggered"] = False
        _risk_state["reason"] = ""
        _risk_state["at"] = None
        _risk_state["last_code"] = None
    _request_times.clear()
    if was_triggered:
        # 曾处于风控停止状态 → 本次登录成功解除, 写系统通知
        _notify_risk_recovered()


def _sleep_random(lo, hi):
    """随机间隔, 打散请求节奏"""
    time.sleep(random.uniform(lo, hi))


def _rate_limit():
    """滑动窗口限速: 超限则排队等待, 避免短时间大量请求(独立锁, 防多线程竞态)"""
    global _request_times
    now = time.time()
    with _rate_lock:
        _request_times = [t for t in _request_times if now - t < 60]
        if len(_request_times) >= RISK_MAX_RPM:
            wait = 60 - (now - _request_times[0]) + random.uniform(0.3, 1.2)
            print(f"[rate] 已达限速(>={RISK_MAX_RPM}次/分), 等待 {wait:.1f}s")
            time.sleep(wait)
            now = time.time()
            _request_times = [t for t in _request_times if now - t < 60]
        _request_times.append(time.time())


def _rate_limit_available():
    """限速槽位是否可用(不占用): 供交互式请求判断是否需要立即返回(独立锁)"""
    now = time.time()
    with _rate_lock:
        return len([t for t in _request_times if now - t < 60]) < RISK_MAX_RPM


def _check_risk(resp, data=None):
    """检测响应中的风控/登录失效信号, 命中则全局止损"""
    if resp.status_code in (401, 403, 429):
        _set_risk(f"HTTP {resp.status_code}", resp.status_code)
        return True
    if not data:
        return False
    code, msg = data.get("code"), str(data.get("msg") or "")
    # 授权/会话失效类错误码
    if code in (3014, 4001, 4010, 5001):
        _set_risk(f"code={code} {msg[:60]}", code)
        return True
    for kw in RISK_KEYWORDS:
        if kw.lower() in msg.lower():
            _set_risk(f"{msg[:60]}", code)
            return True
    return False


def _assert_no_running_task():
    """任务互斥: 核算 / 数据刷新 / 今日抓取 / 手动补抓 不能同时跑"""
    if _refresh_state.get("running"):
        raise RuntimeError("数据刷新任务进行中, 请稍后再试")
    if _trace_state.get("running"):
        raise RuntimeError("核算任务进行中, 请稍后再试")
    if _today_state.get("running"):
        raise RuntimeError("今日数据抓取任务进行中, 请稍后再试")
    if _prefetch_state.get("running"):
        raise RuntimeError("手动补抓任务进行中, 请稍后再试")


# ---------- 核算(消息轨迹) ----------
TRACE_API = "https://agent.tanyuai.com/api/im/agent-trace/paginateV2"

# 发送状态: 1=自动发送 2=侧边栏点击发送 3=编辑后发送 (None=未发送)
SEND_TYPES = {1: "自动发送", 2: "侧边栏点击发送", 3: "编辑后发送"}
ADOPTED_SEND_TYPES = (1, 2, 3)

# 请求间隔(秒): 随机区间抖动, 打散节奏接近人工, 降低触发平台风控的概率
TRACE_PAGE_INTERVAL = (0.2, 0.6)   # 分页之间(大分页后页数极少, 间隔可收紧)
TRACE_SHOP_INTERVAL = (0.8, 2.5)   # 店铺之间
TRACE_PAGE_SIZE = 2000             # 单页消息数: 实测 paginateV2 支持到 5000+,
                                   # 大分页把页数从 N/100 降到 N/2000(20倍), 核算快一个数量级
TRACE_QUERY_CAP = 10000            # 接口单次核算区间最多返回 10000 条(实测硬上限),
                                   # 超出部分翻页也拿不到(results 为 None)


def sleep_trace_page():
    """分页间随机等待"""
    _sleep_random(*TRACE_PAGE_INTERVAL)


def sleep_trace_shop():
    """店铺间随机等待"""
    _sleep_random(*TRACE_SHOP_INTERVAL)


def fetch_trace_page(shop_id, begin, end, page_index, page_size=TRACE_PAGE_SIZE):
    """拉取一页消息轨迹; 带限速+风控检测. 返回 (total, results)

    total 为区间消息总数(用于判断是否超 1 万上限), results 为当前页列表
    """
    payload = {
        "thirdShopId": shop_id,
        "pageIndex": page_index,
        "pageSize": page_size,
        "beginTime": begin,
        "endTime": end,
    }
    _assert_no_risk()
    # 夜间抓取进程正在逐集团切激活集团 cookie → 轨迹直连请求不发(否则带错集团
    # cookie 出网会跨集团串数据, 正是夜间标志要防的; 核算/今日/单店都经此路径)。
    # pid==os.getpid() 自排除, 夜间进程自身抓取不受影响。
    if _nightly_fetch_active():
        raise BusyQueueError("夜间抓取进行中, 请稍后重试")
    _rate_limit()
    resp = requests.post(TRACE_API, json=payload, headers=get_headers(), timeout=20)
    data = resp.json()
    if _check_risk(resp, data):
        # 同 post_api: 抛 RiskTriggered 让调用方首次命中即停, 不被通用 except 吞成假零
        raise RiskTriggered(f"风控/登录失效: {data.get('msg', resp.status_code)}")
    if data.get("success") is False:
        raise RuntimeError(data.get("msg", "接口错误"))
    d = data.get("data") or {}
    return d.get("total") or 0, d.get("results") or []


# 客服昵称同步(人工客服账号池: 真实昵称 + 子账号)
# 数据源: customer-service/table 经营罗盘(实测确认)
#   拼多多(1): serviceAccount='cs_340410493:154573412' serviceName='易方配件专营店:哆啦A梦'(店铺名:昵称)
#   抖音(5)/京东(7): serviceAccount='48653908:小罗' serviceName='小罗'(仅昵称)
# 每店含 '__SUMMARY__' 汇总行(跳过); section 用 'service'(operations 全 0 无数据)
CUSTOMER_SERVICE_TABLE_API = API_BASE + "/customer-service/table"
STAFF_NAMES_FILE = DATA_DIR / "staff_names.json"
STAFF_NAMES_TTL = 6 * 3600  # 同步结果有效期(小时级刷新即可, 客服昵称低频变化)

# 客服昵称同步状态(异步任务, 独立于核算/刷新/今日)
_staff_sync_state = {
    "running": False,
    "progress": {"done": 0, "total": 0, "current": ""},
    "last_run": None,
    "error": None,
    "triggered_by": None,  # 谁发起的客服同步(操作者用户名/昵称), 供状态展示
}


def _load_staff_names():
    """读本地客服昵称映射(不读 config cookie, 无敏感信息)"""
    if not STAFF_NAMES_FILE.exists():
        return None
    try:
        return json.loads(STAFF_NAMES_FILE.read_text(encoding="utf-8"))
    except Exception:
        return None


def _split_staff(acct):
    """拆解 sellerAccount 为账号与真实名称: '48653908:柯柯' -> 账号=48653908 名称=柯柯

    抖音(5)/京东(7) 冒号后是真实客服名; 拼多多(1) 'cs_340410493:154573412'
    冒号后是数字ID(接口不返回真实名), 此时无 name、账号显示完整串。
    拼多多账号形如 cs_<卖家ID>:<客服数字ID>, accountShort=cs_340410493, accountId=154573412。
    客服维度主键始终是完整 sellerAccount(同一个人名可对应多个账号)。
    """
    if not acct:
        return {"account": acct or "", "name": None, "accountShort": acct or "", "accountId": ""}
    if ":" in acct:
        a, b = acct.split(":", 1)
        if b.strip().isdigit():
            return {"account": acct, "name": None, "accountShort": a, "accountId": b.strip()}
        return {"account": acct, "name": b.strip(), "accountShort": a, "accountId": ""}
    return {"account": acct, "name": None, "accountShort": acct, "accountId": ""}


def _staff_list_from_agg(by_staff):
    """把 {account: {total, adopted}} 聚合成 byStaff 列表(含拆出的真实名称), 按 total 降序

    用本地客服昵称映射(staff_names.json)给拼多多客服补真实昵称/店铺名:
    拼多多 sellerAccount 是数字ID('cs_340410493:154573412'), 接口不返回名字,
    昵称来自 customer-service/table 的 serviceName('店铺名:客服昵称')。字段:
      nick     - 真实昵称(拼多多从映射取, 无映射则 None)
      shopName - 该客服所在店铺的展示名(拼多多从 serviceName 前缀取)
    抖音/京东 sellerAccount 已含昵称(_split_staff 的 name), 不受影响。
    """
    staff_names = _load_staff_names()
    name_map = (staff_names or {}).get("map") or {}
    out = []
    for acct, v in by_staff.items():
        meta = _split_staff(acct)
        rec = {
            "account": meta["account"],
            "name": meta["name"],
            "accountShort": meta["accountShort"],
            "total": v["total"],
            "adopted": v["adopted"],
            "rate": round(v["adopted"] / v["total"] * 100, 2) if v["total"] else 0,
        }
        # 拼接多多真实昵称/店铺名(映射键是完整 serviceAccount)
        n = name_map.get(acct)
        if n:
            rec["nick"] = n.get("nick") or rec["name"]
            if n.get("shopName"):
                rec["shopName"] = n["shopName"]
        out.append(rec)
    return sorted(out, key=lambda x: -x["total"])


def _staff_list_per_shop(by_shop, platform=None, shop_filter=None):
    """按 (店铺,客服) 组合列出客服(不去重/不跨店合并), 每项带店铺归属

    客服账号池"按店铺区分客服"的数据源: 同一客服账号在不同店铺各自成行。
    以 staff_names(在编客服)为权威, 合并消息统计:
      - 该店在编客服(含窗口内无消息)全部列出, total=消息统计(无则 0)
      - 消息中出现但不在 staff_names 的客服(新客服/已离职)也列出
    归属: 用 shops 表 seller_id 精确匹配 account 前缀(PDD 去 cs_), 而非
    serviceName 前缀(PDD 大量 serviceName 只有昵称、无店铺名前缀)。
    """
    shop_map = {s["thirdShopId"]: s for s in trace_store.get_shops()}
    # (seller_id, platform) → (shop_id, shop_name): 权威店铺归属表
    seller_shop = {}
    for sid, s in shop_map.items():
        if s.get("sellerId"):
            seller_shop.setdefault((str(s["sellerId"]), s.get("platform")),
                                   (sid, s.get("shopName") or sid))
    staff_names = _load_staff_names()
    name_map = (staff_names or {}).get("map") or {}

    def _acct_shop(acct, plat):
        pre = acct.split(":", 1)[0]
        if pre.startswith("cs_"):
            pre = pre[3:]
        return seller_shop.get((pre, plat))

    # 消息侧统计: shop_id -> {acct: {total, adopted}}
    msg_agg = {sid: dict(shop_agg) for sid, shop_agg in by_shop.items()}
    # 在编客服(权威): staff_names 按 seller_id 归属到店, 同时按平台/子集过滤
    roster = {}
    for acct, rec in name_map.items():
        plat = rec.get("platform")
        if platform is not None and plat != platform:
            continue
        hit = _acct_shop(acct, plat)
        if not hit:
            continue
        sid = hit[0]
        if shop_filter and sid not in shop_filter:
            continue
        roster.setdefault(sid, {})[acct] = rec
    # 导入平台(天猫1/2)在编客服: 用 imported_roster(Excel 登记的客服清单, 不被
    # tanyu 同步覆写)。is_excluded=1 的剔除账号不进入在编池(仍会因消息出现而列出)。
    import_plat_set = set(IMPORT_PLATFORMS)
    if platform in import_plat_set:
        for r in trace_store.get_import_roster(platform):
            if r["isExcluded"]:
                continue
            sid = r["shopId"]
            if shop_filter and sid not in shop_filter:
                continue
            rec = {"platform": platform, "nick": r.get("nick"), "serviceName": r.get("nick")}
            roster.setdefault(sid, {})[r["account"]] = rec
    combos = []
    for sid in set(msg_agg) | set(roster):
        shop = shop_map.get(sid, {})
        sp = shop.get("platform")
        if platform is not None and sp != platform:
            continue
        if shop_filter and sid not in shop_filter:
            continue
        for acct in set(msg_agg.get(sid, {})) | set(roster.get(sid, {})):
            meta = _split_staff(acct)
            v = msg_agg.get(sid, {}).get(acct, {"total": 0, "adopted": 0})
            rec = roster.get(sid, {}).get(acct) or {}
            combo = {
                "shopId": sid,
                "shopName": shop.get("shopName") or rec.get("shopName") or sid,
                "shopPlatform": sp,
                "platformName": PLATFORM_NAMES.get(sp, str(sp or "")),
                "account": acct,
                "name": meta["name"],
                "accountShort": meta["accountShort"],
                "accountId": meta["accountId"],
                "total": v.get("total", 0),
                "adopted": v.get("adopted", 0),
                "rate": round(v.get("adopted", 0) / v.get("total", 0) * 100, 2) if v.get("total") else 0,
            }
            if rec:
                combo["nick"] = rec.get("nick") or combo["name"]
                combo["serviceName"] = rec.get("serviceName")
            combos.append(combo)
    combos.sort(key=lambda c: (c.get("shopName") or "", -(c.get("total") or 0)))
    return combos


def stat_trace(shop_id, begin, end, on_progress=None):
    """遍历指定店铺时间段内的全部消息轨迹, 统计发送状态分布

    返回: {total, counts, adopted, rate, byStaff, staffList, byType}
      total    - 消息总条数
      counts   - sendType 分布 {1: n, 2: n, 3: n, None: n}
      adopted  - 采纳条数(自动发送+侧边栏点击发送+编辑后发送)
      rate     - 核算采纳率
      byStaff  - 按人工客服(sellerAccount)维度统计
      staffList- 出现过的客服账号列表
    """
    results = _fetch_trace_range(shop_id, begin, end)
    total = len(results)

    counts = {1: 0, 2: 0, 3: 0, None: 0}
    by_staff = {}  # sellerAccount -> {"total": n, "adopted": n}
    by_type = {}
    daily = {}  # 按天聚合: date -> {"total": n, "adopted": n}
    for r in results:
        st = r.get("sendType")
        counts[st] = counts.get(st, 0) + 1
        staff = r.get("sellerAccount") or "未知"
        entry = by_staff.setdefault(staff, {"total": 0, "adopted": 0})
        entry["total"] += 1
        if st in ADOPTED_SEND_TYPES:
            entry["adopted"] += 1
        t = r.get("type") or "OTHER"
        by_type[t] = by_type.get(t, 0) + 1
        # 按天聚合(time 为毫秒时间戳)
        ts = r.get("time")
        if isinstance(ts, (int, float)):
            day = datetime.datetime.fromtimestamp(ts / 1000).strftime("%Y-%m-%d")
            de = daily.setdefault(day, {"total": 0, "adopted": 0})
            de["total"] += 1
            if st in ADOPTED_SEND_TYPES:
                de["adopted"] += 1

    adopted = sum(counts.get(s, 0) for s in ADOPTED_SEND_TYPES)
    rate = (adopted / total * 100) if total else 0
    daily_list = [
        {"date": d, "total": v["total"], "adopted": v["adopted"],
         "rate": round(v["adopted"] / v["total"] * 100, 2) if v["total"] else 0}
        for d, v in sorted(daily.items())
    ]
    staff_list = _staff_list_from_agg(by_staff)
    # 原始消息列表(最近的最前, 便于核对)
    raw_messages = [
        {
            "time": r.get("createTime") or r.get("createAt") or r.get("time") or "",
            "buyer": (r.get("buyerNick") or r.get("customerName") or r.get("userName") or ""),
            "sendType": r.get("sendType"),
            "content": (r.get("content") or r.get("replyContent") or r.get("question") or "")[:200],
            "staff": r.get("sellerAccount") or r.get("staffName") or "",
            "type": r.get("type") or "OTHER",
            "traceId": r.get("traceId") or r.get("id") or "",
        }
        for r in reversed(results)
    ]
    return {
        "total": total,
        "counts": counts,
        "adopted": adopted,
        "rate": round(rate, 2),
        "byStaff": staff_list,
        "byType": by_type,
        "daily": daily_list,
        "messages": raw_messages,
    }


# 聚合/原始消息仅用到这些字段, 抓取落盘时裁剪, 缓存体积可缩小 10 倍+
TRACE_KEEP_KEYS = (
    "time", "createTime", "createAt",
    "sendType", "sellerAccount", "staffName",
    "type", "buyerNick", "customerName", "userName",
    "traceId", "id",
)


def _trim_trace_msg(r):
    """把单条原始消息裁剪成下游聚合需要的字段"""
    m = {k: r.get(k) for k in TRACE_KEEP_KEYS if k in r}
    m["content"] = (r.get("content") or r.get("replyContent") or r.get("question") or "")[:200]
    return m


def _fetch_trace_range(shop_id, begin, end):
    """抓取指定时间区间内的全部消息轨迹

    接口单次核算区间最多返回 TRACE_QUERY_CAP 条(实测硬上限), 超出部分翻页也拿不到。
    若 total 触顶, 说明该区间消息数超上限, 自动按 4 小时细分为多个子区间,
    每个子区间都低于上限, 从而完整拿到全部消息(如 1.4 万条的一天需 6 个 4 小时片)。
    """
    total, first_batch = fetch_trace_page(shop_id, begin, end, 1)
    if total >= TRACE_QUERY_CAP:
        # 触顶: 该区间超过 1 万条, 递归细分以保证完整
        sub_results = []
        try:
            b = datetime.datetime.fromisoformat(begin)
            e = datetime.datetime.fromisoformat(end)
        except Exception:
            b = e = None
        if b and e and (e - b).total_seconds() > 4 * 3600:
            for start in _time_slices(b, e, 4):
                sleep_trace_page()
                seg_end = min(start + datetime.timedelta(hours=4), e)  # 最后一段不越出原始 end
                sub_results.extend(
                    _fetch_trace_range(shop_id, start.strftime("%Y-%m-%d %H:%M:%S"),
                                       (seg_end - datetime.timedelta(seconds=1)).strftime("%Y-%m-%d %H:%M:%S"))
                )
            return sub_results
        # 已是 4 小时以内仍触顶: 数据不完整, 抛出信号让调用方不落缓存/库
        raise TraceTruncatedError(
            f"{shop_id} {begin[:10]}~{end[:10]} 消息数 ≥ {TRACE_QUERY_CAP} 条(接口单区间硬上限), "
            f"细分后仍截断, 该段数据未入库")
    # 未触顶: 正常翻页补齐
    results = first_batch
    page = 2
    while len(results) < total and len(results) < TRACE_QUERY_CAP:
        sleep_trace_page()
        try:
            _, batch = fetch_trace_page(shop_id, begin, end, page)
        except RiskTriggered:
            raise
        except BusyQueueError:
            raise  # 夜间抓取占用激活集团 cookie: 不得吞掉(否则门控被绕过返回空结果)
        except Exception as e:
            print(f"[trace] {shop_id} {begin[:10]} page={page} 失败: {e}")
            break
        if not batch:
            break
        results.extend(batch)
        page += 1
    return results


def _time_slices(start, end, hours):
    """把 [start, end) 按 hours 小时切成若干子区间(返回每个子区间的起点)"""
    slices = []
    cur = start
    while cur < end:
        slices.append(cur)
        cur += datetime.timedelta(hours=hours)
    return slices


def _fetch_trace_day(shop_id, day_str):
    """抓取某一天的全部消息轨迹(按天遍历, 缓存可精确到天)

    返回该天所有消息的 results 列表; 失败返回 None(调用方决定是否跳过)
    """
    begin = f"{day_str} 00:00:00"
    end = f"{day_str} 23:59:59"
    return _fetch_trace_range(shop_id, begin, end)


def days_all_cached(shop_id, start, end):
    """判断区间内所有天是否已缓存且未过期(纯聚合, 可跳过限速停顿)"""
    cache = _load_trace_days_cache(shop_id)
    if not cache or not cache.get("days"):
        return False
    _sms, _ems, sday, eday, _ = _split_bounds(start, end)
    start_d = datetime.date.fromisoformat(sday)
    end_d = datetime.date.fromisoformat(eday)
    need = {(start_d + datetime.timedelta(days=i)).isoformat()
            for i in range((end_d - start_d).days + 1)}
    return all(_cached_days_usable(cache, ds) for ds in need)


def _shop_platform(shop_id):
    """查店铺平台: 先看当前 shops 列表, 找不到回退 shop_seller 缓存"""
    try:
        for s in load_all_shops():
            if s.get("thirdShopId") == shop_id:
                return s.get("platform", 0)
    except Exception:
        pass
    return 0


def _upsert_shop_day_db(shop_id, day_str, msgs):
    """抓取成功后把该店当天消息同步进 SQLite(平台从店铺列表取)"""
    platform = _shop_platform(shop_id)
    trace_store.upsert_shop_day(shop_id, platform, day_str, msgs)


def stat_trace_daily(shop_id, start, end, force=False, force_days=None, trim_ms=None,
                     history_mode=False):
    """按天遍历消息轨迹, 逐日缓存, 支持增量更新

    - 已缓存的天直接复用(零请求)
    - 未缓存的天按天抓取
    - force=True 时忽略缓存重新抓取全部天(用于"重新抓取"按钮)
    - force_days: 这些天即使缓存有效也强制重抓(如"昨天"防采纳口径漂移:
      tanyu 会对已抓消息回溯更新 sendType, 不重抓则昨天采纳数停在预抓时刻)
    - history_mode: 夜间预抓专设。历史天缓存"已抓就复用"(不按 7 天 TTL 判过期),
      使每晚增量只抓昨天(未缓存/缺格天仍会补抓), 避免每晚周期性重抓整窗旧天
      (7 天 TTL 会把超过 7 天未重抓的历史天判过期→整窗重抓, 违背"只抓昨天"契约)。
      普通用户/核算请求不传此参数, 保持既有 TTL 保鲜语义。
    - trim_ms=(start_ms, end_ms): 秒级核算区间, 聚合前按消息 msg_time(毫秒)过滤
      (trace_daily 只有整天聚合, 秒级边界必须落到原始消息时间戳)
    - 返回聚合 stat(messages/daily 齐全, 供折线图/核算/原始消息)
    """
    force_days = force_days or set()
    # 按天拆分区间(接受纯日期或带时间的 start/end, 一律取 day 边界)
    _sms, _ems, start_day, end_day, _ht = _split_bounds(start, end)
    start_d = datetime.date.fromisoformat(start_day)
    end_d = datetime.date.fromisoformat(end_day)
    days = [start_d + datetime.timedelta(days=i) for i in range((end_d - start_d).days + 1)]
    day_strs = [d.isoformat() for d in days]
    trim_lo, trim_hi = (trim_ms if trim_ms else (None, None))

    cache = {} if force else (_load_trace_days_cache(shop_id) or {})
    cached_days = cache.get("days") or {}  # {day_str: [messages...]}
    day_fetched_at = cache.get("day_fetched_at") or {}  # {day_str: fetched_ts}
    total_results = []

    for ds in day_strs:
        if not force and ds not in force_days:
            if history_mode:
                # 夜间预抓: 历史天"已抓就复用"(不判 TTL 过期); 只补未抓的缺格天
                if ds in cached_days:
                    total_results.extend(cached_days[ds])
                    continue
            elif _cached_days_usable(cache, ds):
                total_results.extend(cached_days[ds])
                continue
        try:
            # 网络瞬时错误(SSL 断连/限流)退避重试 3 次, 避免一次失败就缺一天
            res = [_trim_trace_msg(r) for r in _with_network_retry(
                lambda: _fetch_trace_day(shop_id, ds), f"{shop_id} {ds}", tag="trace")]
        except RiskTriggered:
            raise
        except BusyQueueError:
            raise  # 夜间抓取占用激活集团 cookie: 不得吞掉(否则门控被绕过返回空结果)
        except TraceTruncatedError as e:
            # 该天数据超接口上限且细分后仍截断: 不落缓存/库, 记日志供人工补抓
            log_line("trace", f"⚠️ {shop_id} {ds} 数据超上限未入库: {e}")
            continue
        except Exception as e:
            log_line("trace", f"{shop_id} {ds} 抓取失败: {e}")
            continue
        cached_days[ds] = res
        day_fetched_at[ds] = time.time()
        total_results.extend(res)
        # 每天抓完立即落盘, 中途失败也不丢已抓数据
        save_cache("trace_days", shop_id, {
            "fetched_at": time.time(),
            "day_fetched_at": day_fetched_at,
            "days": cached_days,
        })
        # 同步写入 SQLite(35 天滚动窗口, 全平台), 供核算/展示纯本地聚合
        try:
            _upsert_shop_day_db(shop_id, ds, res)
        except Exception as e:
            log_line("db", f"{shop_id} {ds} 写入失败: {e}")
        _trace_days_cache[shop_id] = (time.time(), {
            "fetched_at": time.time(),
            "day_fetched_at": day_fetched_at,
            "days": cached_days,
        })
        log_line("trace", f"{shop_id} {ds} 抓取完成 {len(res)} 条")

    # 秒级区间: 聚合前按消息毫秒时间戳过滤(整天的首尾消息在窗口外被裁掉)
    if trim_lo is not None:
        total_results = [
            r for r in total_results
            if trim_lo <= (r.get("time") or r.get("createTime") or r.get("createAt") or 0) <= trim_hi
        ]

    # 聚合成与 stat_trace 相同的结构
    counts = {1: 0, 2: 0, 3: 0, None: 0}
    by_staff = {}
    by_type = {}
    daily = {}
    total = 0
    for r in total_results:
        st = r.get("sendType")
        counts[st] = counts.get(st, 0) + 1
        staff = r.get("sellerAccount") or "未知"
        entry = by_staff.setdefault(staff, {"total": 0, "adopted": 0})
        entry["total"] += 1
        if st in ADOPTED_SEND_TYPES:
            entry["adopted"] += 1
        t = r.get("type") or "OTHER"
        by_type[t] = by_type.get(t, 0) + 1
        ts = r.get("time")
        if isinstance(ts, (int, float)):
            day = datetime.datetime.fromtimestamp(ts / 1000).strftime("%Y-%m-%d")
            de = daily.setdefault(day, {"total": 0, "adopted": 0, "replies": 0})
            de["total"] += 1
            if st in ADOPTED_SEND_TYPES:
                de["adopted"] += 1
            if t == "CONSULT_REPLY":
                de["replies"] += 1
        total += 1

    adopted = sum(counts.get(s, 0) for s in ADOPTED_SEND_TYPES)
    rate = (adopted / total * 100) if total else 0
    daily_list = [
        {"date": d, "total": v["total"], "adopted": v["adopted"], "replies": v["replies"],
         "rate": round(v["adopted"] / v["total"] * 100, 2) if v["total"] else 0}
        for d, v in sorted(daily.items())
    ]
    staff_list = _staff_list_from_agg(by_staff)
    raw_messages = [
        {
            "time": r.get("createTime") or r.get("createAt") or r.get("time") or "",
            "buyer": (r.get("buyerNick") or r.get("customerName") or r.get("userName") or ""),
            "sendType": r.get("sendType"),
            "content": (r.get("content") or r.get("replyContent") or r.get("question") or "")[:200],
            "staff": r.get("sellerAccount") or r.get("staffName") or "",
            "type": r.get("type") or "OTHER",
            "traceId": r.get("traceId") or r.get("id") or "",
        }
        for r in reversed(total_results)
    ]
    return {
        "total": total,
        "counts": counts,
        "adopted": adopted,
        "rate": round(rate, 2),
        "byStaff": staff_list,
        "byType": by_type,
        "daily": daily_list,
        "messages": raw_messages,
    }


# ---------- 接口 ----------
class CookieUpdate(BaseModel):
    cookies: dict


class GroupSwitch(BaseModel):
    group_id: str


@app.get("/api/health")
def health():
    return {"ok": True}


@app.get("/api/shops")
def shops_list(platform: int | None = None):
    """店铺列表: 跨集团全量(平台切换后任何平台都有数据)"""
    return {"shops": load_all_shops(platform)}


@app.get("/api/platforms")
def platforms_info():
    """平台配置信息: 哪些平台抓取, 哪些保留接口, 哪些靠 Excel 文档导入(10/11 天猫1/2)"""
    return {
        "fetch": [{"id": p, "name": PLATFORM_NAMES.get(p, str(p))} for p in FETCH_PLATFORMS],
        "import": [{"id": p, "name": PLATFORM_NAMES.get(p, str(p))} for p in IMPORT_PLATFORMS],
        "keep": [{"id": p, "name": PLATFORM_NAMES.get(p, str(p))} for p in KEEP_PLATFORMS],
        "all": [{"id": p, "name": PLATFORM_NAMES.get(p, str(p))} for p in sorted(PLATFORM_NAMES)],
    }


# ---------- 导入平台(天猫1/2) Excel 文档上传 ----------
# 平台 10/11 无 tanyu 抓取能力, 数据靠 RPA 人工登记的 Excel 上传入库。
# 解析目标 sheet:
#   明细数据1/明细数据2        A-I: 时间段/店铺/店铺消息总量/话术未生成/未发送/话术总量/采纳数/采纳率/生成率
#   探域店铺数据抓取           A-I: 同上(时间段为 "2026-06-01 00:00:00" datetime, 取日期部分)
#   客服数据抓取               A-G: 时间段/店铺/客服/客服消息总量/未发送/采纳数/采纳率
#   后台客服数据抓取           A-L: 时间段(5.1-5.7)/店铺简称/客服/接待量/询单量/下单量/转化/.../满意率
#   要抓的客服账号             A-B: 店铺/客服账号
#   剔除账号登记               A-B: 店铺/账号
# 所有 sheet 按表头嗅探识别, 列位置不写死(容错), 解析失败返回 400 并定位行列。
# 导入策略: 天级聚合写 trace_daily(by_staff_json), 幂等覆盖; 客服 KPI 写 imported_staff_kpi;
#           在编/剔除写 imported_roster。店铺由 IMPORT_SHOPS 常量按店名匹配(未知店名跳行记 warning)。
# 绝不向 tanyu 发任何请求, 绝不清 shops.json(店铺只 upsert 进 SQLite shops 表)。
_IMPORT_MAX_BYTES = 20 * 1024 * 1024  # 20MB 上限

# 店铺简称 → 全名(后台客服数据抓取用简称: 华荣/魔声/博睿兴/星诚/聚源/青橙/果时代/星科/飞利浦)
_IMPORT_SHOP_ALIASES = {
    "华荣": "联想华荣专卖店", "魔声": "MONSTER魔声安诺专卖店", "博睿兴": "联想博睿兴专卖店",
    "星诚": "星诚影音专营店", "聚源": "联想聚源专卖店", "青橙": "魔鬼猫青橙专卖店",
    "果时代": "果时代数码旗舰店", "星科": "星科数码专营", "飞利浦": "飞利浦昕屿专卖店",
}


def _import_parse_date(v, col):
    """把 Excel 时间段单元格解析成 'YYYY-MM-DD'; 非法抛 ValueError(定位列名)"""
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    s = str(v).strip()
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    # 取年份四位 + 两段数字作为 月/日
    raise ValueError(f"无法解析日期列 {col}={v!r}")


def _import_num(v):
    """Excel 数值/文本 → int; 空/'-'/None → 0; 非法抛 ValueError"""
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip().replace(",", "")
    if s in ("", "-", "--", "#N/A", "None"):
        return 0
    if s.endswith("%"):
        s = s[:-1]
    try:
        return int(float(s))
    except ValueError:
        raise ValueError(f"无法解析数值 {v!r}")


def _parse_shop_day_sheet(ws, warnings, by_day, gen_rate_by_day):
    """明细/店铺数据抓取(时间段×店铺, A-I) → by_day[(shop_id, day)] += 聚合 + gen_rate"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0
    header = [("" if c is None else str(c)).strip().replace("\n", "") for c in rows[0]]
    # 列定位: 按表头关键字(允许"时间段/店铺/店铺消息总量/采纳数/生成率"变体)
    def find_col(*keys):
        for i, h in enumerate(header):
            for k in keys:
                if k in h:
                    return i
        return None
    i_day = find_col("时间段") or 0
    i_shop = find_col("店铺")
    i_total = find_col("店铺消息总量")
    i_adopted = find_col("采纳数")
    i_gen = find_col("生成率")
    if i_shop is None or i_total is None:
        raise ValueError(f"sheet[{ws.title}] 表头缺『店铺/店铺消息总量』列, 实际: {header}")
    n = 0
    for r_i, row in enumerate(rows[1:], start=2):
        vals = list(row)
        if not vals or all(c is None or str(c).strip() == "" for c in vals[:6]):
            continue
        try:
            day = _import_parse_date(vals[i_day], "时间段")
            shop_name = str(vals[i_shop]).strip() if vals[i_shop] is not None else ""
        except (ValueError, IndexError):
            warnings.append(f"{ws.title} 第{r_i}行: 日期/店铺字段缺失或非法, 跳过")
            continue
        shop_id = _import_shop_id(shop_name)
        if not shop_id:
            warnings.append(f"{ws.title} 第{r_i}行: 未登记的店铺『{shop_name}』跳过(仅支持 9 家天猫1店铺)")
            continue
        try:
            total = _import_num(vals[i_total] if i_total < len(vals) else None)
            adopted = _import_num(vals[i_adopted] if i_adopted is not None and i_adopted < len(vals) else None)
            gen = None
            if i_gen is not None and i_gen < len(vals):
                gv = vals[i_gen]
                if isinstance(gv, (int, float)):
                    gen = float(gv)
                elif gv is not None and str(gv).strip() not in ("", "-", "#DIV/0!"):
                    gen = float(str(gv).strip().rstrip("%")) / 100 if str(gv).strip().endswith("%") else float(str(gv).strip())
        except ValueError as e:
            warnings.append(f"{ws.title} 第{r_i}行: {e}, 跳过")
            continue
        key = (shop_id, day)
        prev = by_day.get(key)
        if prev:
            # 探域店铺数据抓取与明细可能重叠(同一(店,天)多行); 同源值应一致, 取 max 防 RPA 重复
            by_day[key] = {
                "total": max(prev["total"], total),
                "adopted": max(prev["adopted"], adopted),
            }
            if gen is not None:
                gen_rate_by_day[key] = max(gen_rate_by_day.get(key, 0.0), gen)
        else:
            by_day[key] = {"total": total, "adopted": adopted}
            if gen is not None:
                gen_rate_by_day[key] = gen
        n += 1
    return n


def _parse_staff_sheet(ws, warnings, staff_by_day):
    """客服数据抓取(时间段×店铺×客服, A-G) → staff_by_day[(shop_id, day)][account] = {total, adopted}"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0
    header = [("" if c is None else str(c)).strip().replace("\n", "") for c in rows[0]]

    def find_col(*keys):
        for i, h in enumerate(header):
            for k in keys:
                if k in h:
                    return i
        return None
    i_day = find_col("时间段") or 0
    i_shop = find_col("店铺")
    i_staff = find_col("客服")
    i_total = find_col("客服消息总量")
    i_adopted = find_col("采纳数")
    if i_shop is None or i_staff is None:
        raise ValueError(f"sheet[{ws.title}] 表头缺『店铺/客服』列, 实际: {header}")
    n = 0
    for r_i, row in enumerate(rows[1:], start=2):
        vals = list(row)
        if not vals or all(c is None or str(c).strip() == "" for c in vals[:5]):
            continue
        try:
            day = _import_parse_date(vals[i_day], "时间段")
            shop_name = str(vals[i_shop]).strip() if vals[i_shop] is not None else ""
            account = str(vals[i_staff]).strip() if vals[i_staff] is not None else ""
        except (ValueError, IndexError):
            warnings.append(f"{ws.title} 第{r_i}行: 日期/店铺/客服字段缺失或非法, 跳过")
            continue
        if not account or account == shop_name:
            continue  # 店铺占位行(非客服), 跳过
        shop_id = _import_shop_id(shop_name)
        if not shop_id:
            warnings.append(f"{ws.title} 第{r_i}行: 未登记的店铺『{shop_name}』跳过")
            continue
        try:
            total = _import_num(vals[i_total] if i_total < len(vals) else None)
            adopted = _import_num(vals[i_adopted] if i_adopted is not None and i_adopted < len(vals) else None)
        except ValueError as e:
            warnings.append(f"{ws.title} 第{r_i}行: {e}, 跳过")
            continue
        staff_by_day.setdefault((shop_id, day), {}).setdefault(account, {"total": 0, "adopted": 0})
        # RPA 可能对同一(天,店,客服)重复登记, 值应一致, 取 max 幂等
        cur = staff_by_day[(shop_id, day)][account]
        staff_by_day[(shop_id, day)][account] = {
            "total": max(cur["total"], total), "adopted": max(cur["adopted"], adopted),
        }
        n += 1
    return n


def _parse_kpi_sheet(ws, warnings, kpi_rows):
    """后台客服数据抓取(周×店铺简称×客服, A-L) → imported_staff_kpi rows"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0
    header = [("" if c is None else str(c)).strip().replace("\n", "") for c in rows[0]]
    # 列: 时间段/店铺/客服/接待量/询单量/下单量/转化/(评价参与量/很满意/满意/总满意数)/满意率
    def find_col(*keys):
        for i, h in enumerate(header):
            for k in keys:
                if k in h:
                    return i
        return None
    i_week = find_col("时间段") or 0
    i_shop = find_col("店铺")
    i_staff = find_col("客服")
    i_receive = find_col("接待量")
    i_inquiry = find_col("询单量")
    i_order = find_col("下单量")
    i_conv = find_col("转化")
    i_sat = find_col("满意率")
    if i_shop is None or i_staff is None:
        raise ValueError(f"sheet[{ws.title}] 表头缺『店铺/客服』列, 实际: {header}")

    def week_start(week_str):
        """'5.1-5.7' → 周一日期 'YYYY-05-01'; '2026-05-01 00:00:00' → 取日期"""
        s = str(week_str).strip()
        try:
            return _import_parse_date(s, "时间段")
        except ValueError:
            pass
        if "-" in s and len(s) < 12:
            a = s.split("-")[0].strip()
            # 5.1-5.7 → 2026-05-01(用当年年份, 不写死; RPA 周表通常不含年份)
            parts = a.split(".")
            if len(parts) == 2:
                try:
                    return f"{datetime.date.today().year}-{int(parts[0]):02d}-{int(parts[1]):02d}"
                except ValueError:
                    pass
        raise ValueError(f"无法解析周字段 {week_str!r}")

    n = 0
    for r_i, row in enumerate(rows[1:], start=2):
        vals = list(row)
        if not vals or all(c is None or str(c).strip() == "" for c in vals[:4]):
            continue
        try:
            w = week_start(vals[i_week])
            shop_alias = str(vals[i_shop]).strip() if vals[i_shop] is not None else ""
            account = str(vals[i_staff]).strip() if vals[i_staff] is not None else ""
            shop_name = _IMPORT_SHOP_ALIASES.get(shop_alias, shop_alias)
        except (ValueError, IndexError):
            warnings.append(f"{ws.title} 第{r_i}行: 周/店铺/客服字段非法, 跳过")
            continue
        shop_id = _import_shop_id(shop_name)
        if not shop_id:
            warnings.append(f"{ws.title} 第{r_i}行: 未登记的店铺『{shop_name}』跳过")
            continue
        if not account:
            continue
        try:
            kpi_rows.append({
                "third_shop_id": shop_id, "week_start": w, "account": account,
                "receive_cnt": _import_num(vals[i_receive] if i_receive is not None and i_receive < len(vals) else None),
                "inquiry_cnt": _import_num(vals[i_inquiry] if i_inquiry is not None and i_inquiry < len(vals) else None),
                "order_cnt": _import_num(vals[i_order] if i_order is not None and i_order < len(vals) else None),
                "conversion": float(vals[i_conv]) if i_conv is not None and i_conv < len(vals) and isinstance(vals[i_conv], (int, float)) else None,
                "satisfaction": float(vals[i_sat]) if i_sat is not None and i_sat < len(vals) and isinstance(vals[i_sat], (int, float)) else None,
            })
        except (ValueError, IndexError, TypeError) as e:
            warnings.append(f"{ws.title} 第{r_i}行: {e}, 跳过")
            continue
        n += 1
    return n


def _parse_roster_sheet(ws, warnings, roster_rows, is_excluded):
    """在编/剔除客服(店铺×账号) → imported_roster rows"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0
    header = [("" if c is None else str(c)).strip() for c in rows[0]]
    i_shop = 0 if "店铺" in header[0] else None
    i_acc = None
    for i, h in enumerate(header):
        if "账号" in h or "客服" in h:
            i_acc = i
            break
    if i_shop is None or i_acc is None:
        raise ValueError(f"sheet[{ws.title}] 表头缺『店铺/账号』列, 实际: {header}")
    n = 0
    for r_i, row in enumerate(rows[1:], start=2):
        vals = list(row)
        if not vals or all(c is None or str(c).strip() == "" for c in vals[:2]):
            continue
        shop_name = str(vals[i_shop]).strip() if vals[i_shop] is not None else ""
        account = str(vals[i_acc]).strip() if i_acc < len(vals) and vals[i_acc] is not None else ""
        if not shop_name or not account:
            continue
        shop_id = _import_shop_id(shop_name)
        if not shop_id:
            warnings.append(f"{ws.title} 第{r_i}行: 未登记的店铺『{shop_name}』跳过")
            continue
        roster_rows.append({
            "third_shop_id": shop_id, "account": account,
            "nick": account.split(":")[-1] if ":" in account else account,
            "is_excluded": is_excluded,
        })
        n += 1
    return n


@app.post("/api/import/trace")
def import_trace(file: UploadFile = File(...), platform: int = Form(10),
                 _: dict = Depends(auth.require_admin)):
    """上传 Excel 文档导入天猫1/2 平台聚合数据。仅管理员可导入。

    同步端点(FastAPI 自动放线程池): openpyxl 解析 + 全量落库是重活, async def
    会占事件循环数秒~数十秒, 期间健康检查/其它 async 处理全部停滞。
    """
    if platform not in IMPORT_PLATFORMS:
        raise HTTPException(400, f"platform={platform} 非导入平台(仅 {IMPORT_PLATFORMS})")
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(400, "仅支持 .xlsx 文件")
    content = file.file.read()
    if len(content) > _IMPORT_MAX_BYTES:
        raise HTTPException(400, f"文件超过 20MB 上限({len(content)} 字节)")
    try:
        # 不用 read_only: 该 RPA 生成的工作簿 read_only 模式维度探测异常
        # (iter_rows 只读到第 1 列/第 1 行), 普通模式配 20MB 上限即可控内存
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"无法解析 Excel: {e}")

    warnings = []
    by_day = {}          # (shop_id, day) -> {total, adopted}
    gen_rate_by_day = {}  # (shop_id, day) -> float
    staff_by_day = {}     # (shop_id, day) -> {account: {total, adopted}}
    kpi_rows = []
    roster_rows = []

    try:
        for ws in wb.worksheets:
            title = ws.title.strip()
            if title in ("明细数据 1", "明细数据 2"):
                _parse_shop_day_sheet(ws, warnings, by_day, gen_rate_by_day)
            elif title == "探域店铺数据抓取":
                _parse_shop_day_sheet(ws, warnings, by_day, gen_rate_by_day)
            elif title == "客服数据抓取":
                _parse_staff_sheet(ws, warnings, staff_by_day)
            elif title == "后台客服数据抓取":
                _parse_kpi_sheet(ws, warnings, kpi_rows)
            elif title in ("要抓的客服账号", "剔除账号登记"):
                _parse_roster_sheet(ws, warnings, roster_rows, is_excluded=(title == "剔除账号登记"))
            else:
                # 未知 sheet(如 Sheet6 是纯粘贴文本): 跳过不报错, 不致命
                continue
    except ValueError as e:
        raise HTTPException(400, f"Excel 解析失败: {e}")

    if not by_day and not staff_by_day:
        raise HTTPException(400, f"Excel 中未找到『明细数据1/2 或 探域店铺数据抓取』sheet 的有效行({len(warnings)} 条告警: {warnings[:5]})")

    # 合并客服级数据进天级聚合: by_staff_json 按 total 降序, 与 rebuild_daily 同构
    for (shop_id, day), staff_map in staff_by_day.items():
        agg = by_day.setdefault((shop_id, day), {"total": 0, "adopted": 0})
        staff_total = sum(s["total"] for s in staff_map.values())
        staff_adopted = sum(s["adopted"] for s in staff_map.values())
        agg["total"] = max(agg["total"], staff_total)
        agg["adopted"] = max(agg["adopted"], staff_adopted)
        gen_rate_by_day.setdefault((shop_id, day), None)

    # 落库
    trace_store.upsert_shops(IMPORT_SHOPS)  # 店铺先确保在 shops 表
    trace_store.clear_import_data(platform)  # 全量重传即替换, 防旧日行残留
    for (shop_id, day), agg in by_day.items():
        trace_store.upsert_import_shop_day(
            shop_id, day, agg["total"], agg["adopted"],
            by_staff_map=staff_by_day.get((shop_id, day)),
            generation_rate=gen_rate_by_day.get((shop_id, day)),
        )
    trace_store.upsert_import_kpi(kpi_rows)
    trace_store.upsert_import_roster(roster_rows)

    _invalidate_audit_caches()  # 清核算磁盘/内存缓存, 强制下次从新数据聚合
    log_line("import", f"导入平台{platform}成功: {len(set(k[0] for k in by_day))}店 {len(by_day)}天 {len(roster_rows)}客服 {len(kpi_rows)}条KPI")
    return {
        "ok": True, "platform": platform,
        "shops": len(set(k[0] for k in by_day)),
        "days": len(by_day),
        "staff": len(set(r["account"] for r in roster_rows)),
        "kpis": len(kpi_rows),
        "warnings": warnings[:50],
        "warningCount": len(warnings),
    }


@app.get("/api/import/template")
def import_template(platform: int = 10, _: dict = Depends(auth.require_admin)):
    """下载天猫1/2 Excel 导入模板(表头与解析规则一致, 附该平台已登记店铺名供填写)

    只生成结构: sheet 名/表头与 import_trace 解析器严格对应; 含 1 行示例。
    """
    if platform not in IMPORT_PLATFORMS:
        raise HTTPException(400, f"platform={platform} 非导入平台(仅 {IMPORT_PLATFORMS})")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "明细数据 1"
    ws.append(["时间段", "店铺", "店铺消息总量", "采纳数", "生成率"])
    ws.append([f"{datetime.date.today() - datetime.timedelta(days=1)} 00:00:00", "星科数码专营", 1200, 800, 0.65])
    ws2 = wb.create_sheet("客服数据抓取")
    ws2.append(["时间段", "店铺", "客服", "客服消息总量", "采纳数"])
    ws2.append([f"{datetime.date.today() - datetime.timedelta(days=1)} 00:00:00", "星科数码专营", "客服A", 300, 200])
    ws3 = wb.create_sheet("后台客服数据抓取")
    ws3.append(["时间段", "店铺", "客服", "接待量", "询单量", "下单量", "转化", "满意率"])
    ws3.append(["5.1-5.7", "星科数码专营", "客服A", 100, 60, 30, 0.5, 0.95])
    ws4 = wb.create_sheet("要抓的客服账号")
    ws4.append(["店铺", "账号"])
    ws4.append(["星科数码专营", "客服A"])
    ws5 = wb.create_sheet("剔除账号登记")
    ws5.append(["店铺", "账号"])
    # 该平台已登记店铺名(附录 sheet, 供复制填写; 解析器会跳过未知 sheet)
    shops = [s for s in IMPORT_SHOPS if s["platform"] == platform]
    if shops:
        ws6 = wb.create_sheet("已登记店铺(参考)")
        ws6.append(["店铺名(必须完全一致)"])
        for s in shops:
            ws6.append([s["shopName"]])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    # 文件名保持 ASCII(HTTP 头 latin-1 限制); 前端下载时用中文名
    name = "tianmao_import_template.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{name}"',
                 **_NO_CACHE_HEADERS},
    )


@app.get("/api/groups")
def groups_list():
    """集团列表 + 当前集团"""
    return get_group_list()


@app.post("/api/groups/switch")
def groups_switch(body: GroupSwitch):
    """切换到指定集团并同步店铺"""
    try:
        return switch_group(body.group_id)
    except ValueError as e:
        raise HTTPException(400, str(e))
    except RuntimeError as e:
        raise HTTPException(502, str(e))


@app.get("/api/groups/sync-status")
def groups_sync_status():
    """切换集团时店铺同步阶段(供前端加载读条)

    phase: idle 无切换 / syncing 请求中 / retry 网络波动重试中(attempt 第几次) /
           fallback 已降级本地缓存 / done 完成。detail 为可读文案。
    """
    return dict(_group_sync_state)


@app.post("/api/shops/sync")
def shops_sync():
    """从探域接口同步当前集团的店铺列表"""
    try:
        return sync_shops_from_tanyu()
    except RuntimeError as e:
        raise HTTPException(502, str(e))


@app.get("/api/shops/cached-groups")
def shops_cached_groups():
    """返回 shops.json 里出现过哪些集团(按 groupId), 供比对"""
    return {"ok": True}


# ---------- 历史周(本地 SQLite 聚合) ----------
# tanyu /summary 忽略日期永远只给当前周(已实测), 历史周只能从本地 trace_daily
# 聚合(消息量/采纳数/采纳率)。纯本地零上游请求, 无需风控/限速。
# 历史周本地指标(前端据此渲染), 无 tanyu 专属指标(unavailable)。

def _week_anchor_range(anchor):
    """周一锚点 -> (start, end) 周区间"""
    ws = datetime.date.fromisoformat(anchor)
    return ws.isoformat(), (ws + datetime.timedelta(days=6)).isoformat()


def _history_week_summary(platform, week_anchor, shop_filter=None):
    """本地聚合指定历史周的平台/店铺子集汇总。

    返回 overview 同构 dict(source="history", items 含消息量/采纳数/采纳率,
    coverage 含缺天统计) 或 None(该周无任何本地数据)。
    shop_filter: 可选店铺子集(集合), 只聚合勾选店铺(账号池模式)。
    """
    ws, we = _week_anchor_range(week_anchor)
    today_str = datetime.date.today().isoformat()
    if ws <= today_str <= we:
        return None  # 当前周走 tanyu summary, 不落历史分支
    agg = trace_store.overview_aggregate(ws, we, platform, shop_filter)
    shop_list = agg["shop_list"]
    staff_agg = agg["staff_agg"]
    if not shop_list:
        return None
    total_msgs = sum(v["total"] for _, v in shop_list)
    total_adopted = sum(v["adopted"] for _, v in shop_list)
    # 生成率: 仅导入平台(10/11)有值(Excel 生成率列); 抓取平台恒 None, 不设卡。
    # 恒写 history_gen_rate(值为 None 时前端显示 '—' 无数据), 与 _import_overview_summary
    # 行为一致; 若只在 gen 非 None 时写键, 导入平台某周缺生成率列时整卡消失。
    gen = trace_store.avg_generation_rate(ws, we, platform, shop_filter)
    items_gen = {
        "history_msg_total": {"current": total_msgs, "previous": 0, "comparePercent": None, "label": "消息量"},
        "history_adopted_total": {"current": total_adopted, "previous": 0, "comparePercent": None, "label": "采纳数"},
        "history_accept_rate": {"current": (round(total_adopted * 100.0 / total_msgs, 2) if total_msgs else 0), "previous": 0, "comparePercent": None, "label": "采纳率"},
        "history_gen_rate": {"current": round(gen * 100, 2) if gen is not None else None,
                             "previous": 0, "comparePercent": None, "label": "生成率"},
    }
    items = {k: v for k, v in items_gen.items() if k != "history_gen_rate" or platform in IMPORT_PLATFORMS}
    # 缺天统计(仅该平台): 复用 week_coverage
    wc = trace_store.week_coverage(platform=platform)
    cov = next((w for w in wc if w["week_start"] == week_anchor), None)
    shop_total = dict(shop_list)
    coverage = {
        "days": cov["days"] if cov else 7,
        "shops": len(shop_list),
        "missing": {sid: miss for sid, miss in (cov["missing"] if cov else {}).items() if sid in shop_total},
    }
    return {
        "statType": "natural_week",
        "startDate": ws,
        "endDate": we,
        "platform": platform,
        "platformName": PLATFORM_NAMES.get(platform),
        "source": "history",
        "items": items,
        "coverage": coverage,
    }


@app.get("/api/history/weeks")
def history_weeks(platform: int | None = None):
    """可回溯的历史周列表(基于本地 trace_daily 覆盖), 供历史周下拉。

    纯本地零上游请求; 返回按周降序的 [{start: 周一锚点, end, label, days}]。
    平台可选过滤。仅返回有数据的周(不含当前周——当前周走 tanyu summary)。
    """
    today = datetime.date.today()
    cur_ws = today - datetime.timedelta(days=today.weekday())
    weeks = trace_store.week_coverage(platform=platform)
    out = []
    for w in weeks:
        if w["week_start"] == cur_ws.isoformat():
            continue  # 当前周不在下拉(走 tanyu summary)
        if not w["shop_days"]:
            continue
        out.append({
            "start": w["week_start"],
            "end": w["week_end"],
            "days": w["days"],
            "shops": len(w["shop_days"]),
        })
    return {"weeks": out}


def _period_max_age(stat_type):
    """summary 缓存有效期上限: 周期稳定数据按"本期边界"放宽 TTL。

    natural_week 一周内(M~日)值基本不变, 昨天(同周)写的缓存今天仍有效,
    不因 6h TTL 过期导致推送/概览生成率回落 None; natural_month 同理。
    natural_day 保持 6h(每天换值, 且当天数据还随时间更新)。
    返回秒数, 至少 6h(防边界刚切新周期时缓存刚过期导致短时全 None)。
    """
    now = datetime.datetime.now()
    if stat_type == "natural_week":
        boundary = (now - datetime.timedelta(days=now.weekday())).replace(
            hour=0, minute=0, second=0, microsecond=0)
    elif stat_type == "natural_month":
        boundary = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    else:
        return 6 * 3600
    return max(6 * 3600, int((now - boundary).total_seconds()) + 3600)


def _aggregate_shop_gen_rate(platform, stat_type, anchor=None):
    """抓取平台(抖音/京东等)平台级生成率: 从店铺 summary 缓存加权聚合

    tanyu 平台级 summary 对部分平台(抖音=5/京东=7)返回空 items, 但店铺级
    summary 在线返回 ai_consult_response_rate(生成率, 0-100 同本地标度)。
    这里遍历该平台店铺, 读各自 summary 缓存(data/summary/{shop_id}__{stat_type}
    [__{anchor}].json), 取 ai_consult_response_rate 按 service_consult_cnt(咨询量)
    加权聚合出平台级 current/previous, 供 _import_overview_summary 的生成率卡
    兜底(在线平台级 tanyu > 本聚合 > None)。

    anchor: 历史周周一锚点(natural_week 专属); 传 None 时读当前周缓存
    (无锚点键, 与 shop_detail 的键构造一致)。缓存命中受 6h TTL 约束。
    返回 (current_float|None, previous_float|None), 标度与 trace_store.
    avg_generation_rate 一致(分数 0-1, 由调用方 *100 转百分比), 无任何带
    生成率的店铺缓存时返回 (None, None)。
    """
    if platform not in FETCH_PLATFORMS:
        return None, None
    # 仅当前期有缓存且未被 TTL 淘汰时参与聚合; 带锚点(历史周)缓存 data 为空,
    # 读不到生成率自然排除, 不会误计。
    max_age = _period_max_age(stat_type)
    rows = []
    for s in load_all_shops(platform):
        sid = s.get("thirdShopId")
        if not sid:
            continue
        key = f"{sid}__{stat_type}" + (f"__{anchor}" if anchor else "")
        cache = load_cache("summary", key, max_age=max_age)
        if not cache:
            continue
        data = cache.get("data") or {}
        gen = (data.get("ai_consult_response_rate") or {}).get("current")
        consult = (data.get("service_consult_cnt") or {}).get("current")
        if gen is None or consult is None:
            continue
        prev = (data.get("ai_consult_response_rate") or {}).get("previous")
        # tanyu 返回 0-100 标度, 转 0-1 分数与 avg_generation_rate 同标, 由调用方转回
        rows.append((gen / 100.0, consult, (prev / 100.0 if prev is not None else None)))
    if not rows:
        return None, None

    def _wmean(pick):
        # 过滤 pick 值为 None 的行(previous 对部分新店可能缺失), 否则 None*权重崩 TypeError
        pairs = [(pick(row), row[1]) for row in rows]
        pairs = [(v, w) for v, w in pairs if v is not None]
        tot_w = sum(w for _, w in pairs)
        if tot_w <= 0:
            # 全部零权重(咨询量全 0): 退化为等权, 不因除零吞掉值
            vals = [v for v, _ in pairs]
            return (sum(vals) / len(vals)) if vals else None
        return sum(v * w for v, w in pairs) / tot_w

    cur = _wmean(lambda row: row[0])
    prev_vals = [row[2] for row in rows if row[2] is not None]
    prev = _wmean(lambda row: row[2]) if prev_vals else None
    return cur, prev


def _import_overview_summary(platform, stat_type, week_anchor=None, cap_to_data_end=False):
    """导入平台(天猫1/2)平台维度汇总: 纯本地 trace_daily 聚合, 零上游请求

    与 overview 返回同构(source="import"), 口径与 tanyu summary 一致(忽略日期):
      natural_day   = 昨天 vs 前天
      natural_week  = 本周(周一~昨天) vs 上周
      natural_month = 本月(1~昨天) vs 上月
    4 张卡: 消息量/采纳数/采纳率/生成率(generation_rate 来自导入 Excel 生成率列)。
    统计字段: 消息量/采纳数直接聚合, 采纳率=adopted/total, 生成率=trace_daily.generation_rate。
    返回额外带 data_start/data_end = 该平台 trace_daily 实际数据范围(按导入表格),
    供钉钉推送等按"导入表格截止"显示数据截止, 而非自然日口径默认区间。

    cap_to_data_end: 钉钉推送等"按导入表截止显示"的场景传 True。当导入表滞后
    (data_end < 昨天)时, 数值聚合窗口截止到表末天, 避免"显示 06-01~08-06 但
    数值按昨天聚合=0"的自相矛盾(表还没导入昨天, 却显示昨天 0 消息)。
    默认 False 保持既有 natural 口径(看板页面不改变行为)。
    """
    today = datetime.date.today()
    one_day = datetime.timedelta(days=1)
    yesterday = today - one_day

    # 该平台 trace_daily 实际数据范围(按导入表格覆盖到哪天, 而非自然日默认区间)
    data_start = data_end = None
    try:
        row = trace_store._conn().execute(
            """SELECT MIN(d.day) m, MAX(d.day) x FROM trace_daily d
               JOIN shops s ON s.third_shop_id = d.third_shop_id
               WHERE s.platform = ?""", (platform,)
        ).fetchone()
        if row:
            data_start, data_end = row["m"], row["x"]
    except Exception:
        pass

    # 数值聚合截止天: 默认昨天; cap_to_data_end 且表滞后时用表末天(见 docstring)
    win_end = yesterday
    if cap_to_data_end and data_end:
        try:
            de = datetime.date.fromisoformat(data_end)
            if de < yesterday:
                win_end = de
        except Exception:
            pass

    if stat_type == "natural_week":
        cur_ws = win_end - datetime.timedelta(days=win_end.weekday())
        # 周一当天: cur_ws==win_end(周一) > 上一窗口末, 区间反转会让 BETWEEN
        # 匹配零行、本周四卡全空。归一为 min(周一, 窗口末)..窗口末(周一当天=单日)。
        cur_start = min(cur_ws.isoformat(), win_end.isoformat())
        cur_end = win_end.isoformat()
        prev_start = (cur_ws - datetime.timedelta(days=7)).isoformat()
        prev_end = (cur_ws - one_day).isoformat()
    elif stat_type == "natural_month":
        # 每月1号: win_end.replace(day=1)==1号 > 上一窗口末(上月最后一天), 同样反转。
        cur_start = min(win_end.replace(day=1).isoformat(), win_end.isoformat())
        cur_end = win_end.isoformat()
        prev_start = (win_end.replace(day=1) - one_day).replace(day=1).isoformat()
        prev_end = (win_end.replace(day=1) - one_day).isoformat()
    else:  # natural_day
        cur_start = cur_end = win_end.isoformat()
        prev_start = prev_end = (win_end - one_day).isoformat()

    def _agg(s, e):
        agg = trace_store.overview_aggregate(s, e, platform)
        total = sum(v["total"] for _, v in agg["shop_list"])
        adopted = sum(v["adopted"] for _, v in agg["shop_list"])
        # 生成率: 导入平台(10/11)来自 trace_daily.generation_rate(Excel 直接给出);
        # 抓取平台(1/5/7)本地库无该列, 从店铺 summary 缓存按咨询量加权聚合(当前期)。
        # 历史周(anchor 键)缓存 data 为空, 聚合自然返回 None, 生成率卡显示 '—'。
        if platform in FETCH_PLATFORMS:
            gen, gen_prev = _aggregate_shop_gen_rate(platform, stat_type, anchor=week_anchor)
            return total, adopted, gen, gen_prev
        gen = trace_store.avg_generation_rate(s, e, platform)
        return total, adopted, gen, None

    cur_total, cur_adopted, cur_gen, cur_gen_prev = _agg(cur_start, cur_end)
    prev_total, prev_adopted, prev_gen, prev_gen_prev = _agg(prev_start, prev_end)
    # 抓取平台聚合给的 previous 才是上一期生成率(_agg 的 s/e 是 local 口径聚合用,
    # 店铺缓存按 stat_type 语义自带 current/previous), 用它替换掉占位的 prev_gen。
    if cur_gen_prev is not None:
        prev_gen = cur_gen_prev

    def _card(current, previous, label):
        # comparePercent 需 current/previous 都非 None 才计算; 生成率可能为 None
        # (某区间无 Excel 生成率列), None - float 会抛 TypeError 让整卡 500。
        pct = None
        if current is not None and previous:
            pct = round((current - previous) / previous * 100, 2)
        return {
            "current": current,
            "previous": previous,
            "comparePercent": pct,
            "label": label,
        }

    items = {
        "history_msg_total": _card(cur_total, prev_total, "消息量"),
        "history_adopted_total": _card(cur_adopted, prev_adopted, "采纳数"),
        "history_accept_rate": _card(round(cur_adopted * 100.0 / cur_total, 2) if cur_total else 0,
                                     round(prev_adopted * 100.0 / prev_total, 2) if prev_total else 0, "采纳率"),
        "history_gen_rate": _card(round(cur_gen * 100, 2) if cur_gen is not None else None,
                                  round(prev_gen * 100, 2) if prev_gen is not None else None, "生成率"),
    }
    return {
        "statType": stat_type,
        "startDate": _stat_type_range(stat_type)[0],
        "endDate": _stat_type_range(stat_type)[1],
        "data_start": data_start,
        "data_end": data_end,
        "platform": platform,
        "platformName": PLATFORM_NAMES.get(platform),
        "source": "import",
        "items": items,
    }


@app.get("/api/overview")
def overview(platform: int = 1, stat_type: str = "natural_day", start: str | None = None, end: str | None = None,
             week: str | None = None, cap_to_data_end: bool = False):
    """平台维度汇总, 按统计口径(自然日/自然周/自然月)。

    tanyu summary 固定忽略 startDate/endDate: 三种口径分别返回 昨天vs前天 /
    本周vs上周 / 本月vs上月。days 参数已废弃(曾误导为可调范围), 仅保留
    start/end 透传给 tanyu(无实际作用, 但保接口兼容)。

    week(可选, 仅 natural_week): 周一锚点 YYYY-MM-DD。传入历史周时改用本地
    SQLite 聚合(source="history", 消息量/采纳数/采纳率 + coverage), 因为 tanyu
    summary 忽略日期永远只给当前周(实测); 不传或传当前周走 tanyu summary。

    cap_to_data_end(钉钉推送专用): 导入平台表滞后时数值聚合截止到表末天,
    与 data_end 显示区间对齐(否则"显示 06-01~08-06 但数值按昨天聚合=0")。
    默认 False 保持既有 natural 口径(看板页面不受影响)。
    """
    if platform not in PLATFORM_NAMES:
        raise HTTPException(400, f"不支持的平台: {platform}")
    if not _valid_stat_type(stat_type):
        raise HTTPException(400, f"不支持的统计口径: {stat_type}")
    # 未显式传区间时按口径给默认值(周=本周一~周日, 月=本月1日~今天, 日=昨天)
    _start, _end = _stat_type_range(stat_type)
    start = _valid_date_iso(start) or _start
    end = _valid_date_iso(end) or _end
    # 历史周分支必须先于导入平台分支判定: 导入平台(天猫1/2)无限期存储,
    # 选历史周时必须能回溯到本地库的旧周数据, 不能吞掉 week 参数直接给当前周聚合。
    # 当前周(week 空/当前锚点)时 _history_week_summary 返回 None, 自然落到下方对应分支。
    if stat_type == "natural_week" and week and _valid_date_iso(week):
        hist = _history_week_summary(platform, week)
        if hist is not None:
            return hist
    # 导入平台(天猫1/2)分支: 纯本地 trace_daily 聚合, 绝不请求 tanyu summary
    if platform in IMPORT_PLATFORMS:
        return _import_overview_summary(platform, stat_type, cap_to_data_end=cap_to_data_end)
    payload = {
        "statType": stat_type,
        "startDate": start,
        "endDate": end,
        "platform": platform,
        "dimension": "platform",
    }
    try:
        summary = fetch_summary_interactive(payload)
    except RiskTriggered:
        raise HTTPException(503, "风控触发, 请先登录更新Cookie")
    except BusyQueueError as e:
        raise HTTPException(503, str(e))
    except (RuntimeError, requests.exceptions.RequestException, ValueError) as e:
        # tanyu 不可达(SSL 断连/超时/限流/响应非 JSON): 回落本地 DB 聚合, 不返回空。
        # SSLEOFError 等网络错误是 requests.exceptions.RequestException(OSError 子类),
        # 并非 RuntimeError —— 只 catch RuntimeError 会让 SSL 断连 500/空面板逃逸。
        # 抓取平台数据每晚预抓已入库, 本地聚合口径与 tanyu summary 一致,
        # 只是缺失 tanyu 专属指标(service_3m_response_rate 等)。
        # 风控例外: RiskTriggered 已单独处理, 不会走到这里。
        print(f"[overview] tanyu summary 失败({type(e).__name__}: {e}), 回落本地 DB 聚合 platform={platform}")
        local = _import_overview_summary(platform, stat_type)
        local["source"] = "local-fallback"
        return local
    if not summary:
        # tanyu summary 成功但为空(部分平台当前无 summary 数据/未启用): 同样回落本地 DB,
        # 否则前端三张 tanyu 指标卡全空显示"无数据"。抓取平台数据每晚预抓已入库。
        print(f"[overview] tanyu summary 为空, 回落本地 DB 聚合 platform={platform}")
        local = _import_overview_summary(platform, stat_type)
        local["source"] = "local-fallback"
        return local
    return {
        "startDate": start,
        "endDate": end,
        "statType": stat_type,
        "platform": platform,
        "platformName": PLATFORM_NAMES.get(platform),
        "items": summary,
    }


@app.get("/api/shop/{shop_id}")
def shop_detail(shop_id: str, stat_type: str = "natural_day", start: str | None = None, end: str | None = None,
                tables: int = 1, week: str | None = None):
    """单店汇总 + 明细(带本地缓存); 按统计口径(自然日/自然周/自然月)。

    tanyu summary 固定忽略 startDate/endDate(见 SUMMARY_STAT_TYPES 注释);
    明细表(section/table)对日期敏感, natural_week/natural_month 时按传入的
    周/月区间定位。days 参数已废弃, 保留 start/end 透传。
    tables=0: 只返回 summary(店铺列表用, 跳过明细表请求, 省上游配额)。

    week(可选, 仅 natural_week): 周一锚点。历史周时走本地 SQLite 聚合
    (source="history"), 返回该店该周消息量/采纳数/采纳率 + missing_days;
    tanyu 专属指标标记 unavailable。当前周不传/传当前周走 tanyu summary。
    """
    shops = {s["thirdShopId"]: s for s in load_all_shops()}
    shop = shops.get(shop_id)
    if not shop:
        raise HTTPException(404, "店铺不存在")
    if not _valid_stat_type(stat_type):
        raise HTTPException(400, f"不支持的统计口径: {stat_type}")

    # 未显式传区间时按口径给默认值(与 /api/overview 一致)
    _start, _end = _stat_type_range(stat_type)
    start = _valid_date_iso(start) or _start
    end = _valid_date_iso(end) or _end
    # 出网 endDate clamp 基准: 自然周默认区间含未来周日, 上游 payload 发送时截到今天
    today_str = datetime.date.today().isoformat()

    # 历史周分支: 本地 SQLite 聚合(零上游请求)
    if stat_type == "natural_week" and week and _valid_date_iso(week):
        hist = _history_week_shop(shop, week)
        if hist is not None:
            return hist
    # 导入平台(天猫1/2)店: 纯本地 trace_daily 聚合, 无原始消息/无 tanyu 明细表,
    # 绝不请求 tanyu summary/section。source="import", tables 空。
    if _is_import_shop(shop):
        stat = _import_shop_stat(shop_id, start, end)
        items = {}
        if stat:
            items = {
                "history_msg_total": {"current": stat["total"], "previous": 0, "comparePercent": None, "label": "消息量"},
                "history_adopted_total": {"current": stat["adopted"], "previous": 0, "comparePercent": None, "label": "采纳数"},
                "history_accept_rate": {"current": stat["rate"], "previous": 0, "comparePercent": None, "label": "采纳率"},
                "history_gen_rate": {"current": stat["generation_rate"], "previous": 0, "comparePercent": None, "label": "生成率"},
            }
        return {
            "shop": shop,
            "startDate": start,
            "endDate": end,
            "statType": stat_type,
            "source": "import",
            "items": items,
            "fetchedAt": time.time(),
            "tables": {},
        }
    # 平台/店铺自然日数据一天内基本不变, 缓存 6 小时; 页面加载/切平台走交互快通道, 不排队
    # 缓存键按口径拆维度: 三口径各用独立后缀键({id}__natural_day/week/month), 防批量刷新串写
    # natural_week 键额外绑周一锚点(start): 周界 0 点换新键必然 miss, 杜绝"上周数值服务到
    # 周一早高峰"的 6h 交叉陈旧窗口(旧键 6h TTL 后自然失效, 孤儿化无危害)。
    cache_key = f"{shop_id}__{stat_type}" + (f"__{start}" if stat_type == "natural_week" else "")
    cache = load_cache("summary", cache_key, max_age=21600)
    if cache and cache.get("data"):
        items = cache["data"]
        fetched_at = cache["fetched_at"]
    else:
        payload = {
            "statType": stat_type,
            "startDate": start,
            "endDate": end,
            "platform": shop.get("platform", 1),
            "dimension": "shop",
            "targetId": shop_id,
        }
        try:
            items = fetch_summary_interactive(payload)
        except RiskTriggered:
            raise HTTPException(503, "风控触发, 请先登录更新Cookie")
        except BusyQueueError as e:
            raise HTTPException(503, str(e))
        except RuntimeError as e:
            raise HTTPException(502, str(e))
        save_cache("summary", cache_key, {"fetched_at": time.time(), "data": items})
        fetched_at = time.time()

    tables_data = {}
    if tables:
        for section in ["operations", "service", "ai"]:
            # 明细表键带日期区间(与 refresh_one 批量刷新写键一致): 日期敏感的
            # /section/table 数据严格绑定其起止日期, 周/月明细不会与自然日串写。
            # 键保持完整自然周(start~周日): 周内恒定不逐日换键; clamp 只作用于出网 payload。
            table_key = f"{shop_id}__{stat_type}__{section}__{start}__{end}"
            c = load_cache("table", table_key, max_age=21600)
            if c and c.get("data"):
                tables_data[section] = c["data"]
            else:
                # 出网 endDate clamp 到今天: 自然周默认区间含未来周日(如周五的 08-09),
                # 未来天物理无数据, 且 tanyu /section/table 对未来 endDate 行为未验证,
                # 故发送 min(end, today)(ISO 字典序==时间序)。历史周 we<today 自动不变。
                eff_end = min(end, today_str)
                payload = {
                    "statType": stat_type,
                    "startDate": start,
                    "endDate": eff_end,
                    "platform": shop.get("platform", 1),
                    "dimension": "shop",
                    "targetId": shop_id,
                    "section": section,
                }
                try:
                    tables_data[section] = fetch_section_table_interactive(payload)
                    # 抓取成功即写 6h 缓存, 与读取键一致形成闭环(否则周/月每次必 miss)
                    save_cache("table", table_key,
                               {"fetched_at": time.time(), "data": tables_data[section]})
                except BusyQueueError as e:
                    raise HTTPException(503, str(e))
                except RiskTriggered:
                    raise HTTPException(503, "风控触发, 请先登录更新Cookie")
                except RuntimeError as e:
                    # 单个 section 上游失败: 记录日志, 返回空表(该区间可能确实无数据), 不打断其余 section
                    log_line("shop_detail", f"shop={shop_id} section={section} 失败: {e}")
                    tables_data[section] = {"dates": [], "rows": []}

    return {
        "shop": shop,
        "startDate": start,
        "endDate": end,
        "statType": stat_type,
        "items": items,
        "fetchedAt": fetched_at,
        "tables": tables_data,
    }


# ---------- 罗盘扩展路由: 趋势 & 客服经营 ----------
TREND_CACHE_TTL = 6 * 3600  # 自然日趋势一天内基本不变, 与 summary 缓存同 TTL


def _trend_days_range(days):
    """近 N 天区间(截至昨天), days 收敛到 1..60"""
    days = max(1, min(int(days or 7), 60))
    return date_range(days=days)


def _metric_keys_param(section, metric_keys, defaults_map):
    """解析 metric_keys 参数: 显式传则过滤, 未传用分节默认键"""
    if metric_keys:
        ks = [k.strip() for k in str(metric_keys).split(",") if k.strip()]
        if ks:
            return ks
    return defaults_map.get(section, ["ops_order_payment_amount"])


@app.get("/api/trend/platform")
def platform_trend(platform: int = 1, days: int = 7, section: str = "operations",
                   metric_keys: str | None = None):
    """平台维度经营趋势(section/trend, 按天序列)。

    数据源: /api/data-service/business/compass/section/trend(dimension=platform)。
    导入平台(天猫1/2)无 tanyu 抓取能力, 短路返回空(source="import")。
    """
    if platform not in PLATFORM_NAMES:
        raise HTTPException(400, f"不支持的平台: {platform}")
    if section not in TREND_SECTION_METRICS:
        raise HTTPException(400, f"不支持的 section: {section}")
    if platform in IMPORT_PLATFORMS:
        return {"platform": platform, "platformName": PLATFORM_NAMES.get(platform),
                "dates": [], "series": [], "source": "import"}
    start, end = _trend_days_range(days)
    keys = _metric_keys_param(section, metric_keys, TREND_SECTION_METRICS)
    cache_key = f"plat{platform}__{section}__{start}__{end}"
    cache = load_cache("trend", cache_key, max_age=TREND_CACHE_TTL)
    if cache and cache.get("dates"):
        return {**cache, "platform": platform,
                "platformName": PLATFORM_NAMES.get(platform), "source": "cache"}
    payload = {
        "statType": "natural_day",
        "startDate": start,
        "endDate": min(end, datetime.date.today().isoformat()),
        "platform": platform,
        "dimension": "platform",
        "section": section,
        "metricKeys": keys,
    }
    try:
        trend = fetch_section_trend_interactive(payload)
    except RiskTriggered:
        raise HTTPException(503, "风控触发, 请先登录更新Cookie")
    except BusyQueueError as e:
        raise HTTPException(503, str(e))
    except RuntimeError as e:
        raise HTTPException(502, str(e))
    save_cache("trend", cache_key, {"dates": trend["dates"], "series": trend["series"],
                                    "startDate": start, "endDate": end})
    return {"platform": platform, "platformName": PLATFORM_NAMES.get(platform),
            "startDate": start, "endDate": end, **trend, "source": "tanyu"}


@app.get("/api/shop/{shop_id}/trend")
def shop_trend(shop_id: str, days: int = 7, section: str = "operations",
               metric_keys: str | None = None):
    """单店经营趋势(section/trend, 按天序列)。导入平台店短路返回空。"""
    shops = {s["thirdShopId"]: s for s in load_all_shops()}
    shop = shops.get(shop_id)
    if not shop:
        raise HTTPException(404, "店铺不存在")
    platform = shop.get("platform", 1)
    if platform in IMPORT_PLATFORMS:
        return {"shop": shop, "dates": [], "series": [], "source": "import"}
    if section not in TREND_SECTION_METRICS:
        raise HTTPException(400, f"不支持的 section: {section}")
    start, end = _trend_days_range(days)
    keys = _metric_keys_param(section, metric_keys, TREND_SECTION_METRICS)
    cache_key = f"{shop_id}__{section}__{start}__{end}"
    cache = load_cache("trend", cache_key, max_age=TREND_CACHE_TTL)
    if cache and cache.get("dates"):
        return {**cache, "shop": shop, "source": "cache"}
    payload = {
        "statType": "natural_day",
        "startDate": start,
        "endDate": min(end, datetime.date.today().isoformat()),
        "platform": platform,
        "dimension": "shop",
        "targetId": shop_id,
        "section": section,
        "metricKeys": keys,
    }
    try:
        trend = fetch_section_trend_interactive(payload)
    except RiskTriggered:
        raise HTTPException(503, "风控触发, 请先登录更新Cookie")
    except BusyQueueError as e:
        raise HTTPException(503, str(e))
    except RuntimeError as e:
        raise HTTPException(502, str(e))
    save_cache("trend", cache_key, {"dates": trend["dates"], "series": trend["series"],
                                    "startDate": start, "endDate": end})
    return {"shop": shop, "startDate": start, "endDate": end, **trend, "source": "tanyu"}


@app.get("/api/shop/{shop_id}/cs-summary")
def shop_cs_summary(shop_id: str, days: int = 7):
    """单店客服经营汇总(customer-service/summary)。

    区间为近 N 天(截至昨天), 与罗盘前端一致; current/previous 为区间 vs 上期。
    """
    shops = {s["thirdShopId"]: s for s in load_all_shops()}
    shop = shops.get(shop_id)
    if not shop:
        raise HTTPException(404, "店铺不存在")
    platform = shop.get("platform", 1)
    if platform in IMPORT_PLATFORMS:
        return {"shop": shop, "items": {}, "labels": CS_SUMMARY_LABELS, "source": "import"}
    start, end = _trend_days_range(days)
    cache_key = f"{shop_id}__cs__{start}__{end}"
    cache = load_cache("trend", cache_key, max_age=TREND_CACHE_TTL)
    if cache and cache.get("items"):
        return {**cache, "shop": shop, "labels": CS_SUMMARY_LABELS, "source": "cache"}
    payload = {
        "startDate": start,
        "endDate": min(end, datetime.date.today().isoformat()),
        "platform": platform,
        "shopId": shop_id,
    }
    try:
        items = fetch_cs_summary_interactive(payload)
    except RiskTriggered:
        raise HTTPException(503, "风控触发, 请先登录更新Cookie")
    except BusyQueueError as e:
        raise HTTPException(503, str(e))
    except RuntimeError as e:
        raise HTTPException(502, str(e))
    for k, it in items.items():
        it.setdefault("label", CS_SUMMARY_LABELS.get(k, k))
    save_cache("trend", cache_key, {"items": items, "startDate": start, "endDate": end})
    return {"shop": shop, "startDate": start, "endDate": end, "items": items,
            "labels": CS_SUMMARY_LABELS, "source": "tanyu"}


@app.get("/api/shop/{shop_id}/cs-detail")
def shop_cs_detail(shop_id: str, days: int = 7, section: str = "service",
                   metric_keys: str | None = None, service_account: str | None = None):
    """单店客服明细表(customer-service/detail/table)。

    service_account 可选: 传则只查该客服账号(完整 sellerAccount, 形如
    '48653908:小罗' / 'cs_340410493:154573412'), 不传返回全部客服行。
    """
    shops = {s["thirdShopId"]: s for s in load_all_shops()}
    shop = shops.get(shop_id)
    if not shop:
        raise HTTPException(404, "店铺不存在")
    platform = shop.get("platform", 1)
    if platform in IMPORT_PLATFORMS:
        return {"shop": shop, "dates": [], "rows": [], "source": "import"}
    if section not in CS_DETAIL_METRICS:
        raise HTTPException(400, f"不支持的 section: {section}")
    start, end = _trend_days_range(days)
    keys = _metric_keys_param(section, metric_keys, CS_DETAIL_METRICS)
    payload = {
        "startDate": start,
        "endDate": min(end, datetime.date.today().isoformat()),
        "platform": platform,
        "shopId": shop_id,
        "section": section,
        "metricKeys": keys,
    }
    if service_account:
        payload["serviceAccount"] = service_account
    try:
        data = fetch_cs_detail_table_interactive(payload)
    except RiskTriggered:
        raise HTTPException(503, "风控触发, 请先登录更新Cookie")
    except BusyQueueError as e:
        raise HTTPException(503, str(e))
    except RuntimeError as e:
        raise HTTPException(502, str(e))
    return {"shop": shop, "startDate": start, "endDate": end, **data, "source": "tanyu"}


@app.get("/api/shop/{shop_id}/cs-trend")
def shop_cs_trend(shop_id: str, days: int = 7, section: str = "service",
                  metric_keys: str | None = None, service_account: str | None = None):
    """单店客服明细趋势(customer-service/detail/trend)。"""
    shops = {s["thirdShopId"]: s for s in load_all_shops()}
    shop = shops.get(shop_id)
    if not shop:
        raise HTTPException(404, "店铺不存在")
    platform = shop.get("platform", 1)
    if platform in IMPORT_PLATFORMS:
        return {"shop": shop, "dates": [], "series": [], "source": "import"}
    if section not in CS_DETAIL_METRICS:
        raise HTTPException(400, f"不支持的 section: {section}")
    start, end = _trend_days_range(days)
    keys = _metric_keys_param(section, metric_keys, CS_DETAIL_METRICS)
    payload = {
        "startDate": start,
        "endDate": min(end, datetime.date.today().isoformat()),
        "platform": platform,
        "shopId": shop_id,
        "section": section,
        "metricKeys": keys,
    }
    if service_account:
        payload["serviceAccount"] = service_account
    try:
        data = fetch_cs_detail_trend_interactive(payload)
    except RiskTriggered:
        raise HTTPException(503, "风控触发, 请先登录更新Cookie")
    except BusyQueueError as e:
        raise HTTPException(503, str(e))
    except RuntimeError as e:
        raise HTTPException(502, str(e))
    return {"shop": shop, "startDate": start, "endDate": end, **data, "source": "tanyu"}


def _history_week_shop(shop, week_anchor):
    """本地聚合单个店铺的历史周数据(纯只读, 零上游请求)。

    返回与 shop_detail 同构的 dict(source="history"), 含该店周消息量/采纳数/
    采纳率 + missing_days(缺天列表); tanyu 专属指标标记 unavailable。
    该周该店无任何数据时返回 None(调用方走 tanyu summary 兜底, 行为不变)。
    """
    ws, we = _week_anchor_range(week_anchor)
    today_str = datetime.date.today().isoformat()
    if ws <= today_str <= we:
        return None  # 当前周走 tanyu summary
    shop_id = shop["thirdShopId"]
    platform = shop.get("platform", 1)
    daily = trace_store.query_daily(shop_id, ws, we)
    if not daily:
        return None
    total = sum(d["total"] for d in daily)
    adopted = sum(d["adopted"] for d in daily)
    have_days = {d["day"] for d in daily}
    # 缺天: 周区间应覆盖天数 vs 实际天数
    missing = []
    for i in range(7):
        ds = (datetime.date.fromisoformat(ws) + datetime.timedelta(days=i)).isoformat()
        if ds not in have_days:
            missing.append(ds)
    items = {
        "history_msg_total": {"current": total, "previous": 0, "comparePercent": None, "label": "消息量"},
        "history_adopted_total": {"current": adopted, "previous": 0, "comparePercent": None, "label": "采纳数"},
        "history_accept_rate": {"current": (round(adopted * 100.0 / total, 2) if total else 0), "previous": 0, "comparePercent": None, "label": "采纳率"},
    }
    # 生成率: 仅导入平台(10/11)有值; 恒写键(值为 None 时前端显示 '—'), 与平台维度一致
    if platform in IMPORT_PLATFORMS:
        gen = trace_store.avg_generation_rate(ws, we, platform=platform,
                                              shop_filter={shop_id})
        items["history_gen_rate"] = {"current": round(gen * 100, 2) if gen is not None else None,
                                     "previous": 0, "comparePercent": None, "label": "生成率"}
    # tanyu 专属指标历史不可用(前端显示"历史不可用")
    for key in ("service_3m_response_rate", "ai_consult_response_accept_rate", "ai_consult_response_rate"):
        items[key] = {"unavailable": True, "label": METRIC_BY_KEY[key]["title"]}
    return {
        "shop": shop,
        "startDate": ws,
        "endDate": we,
        "statType": "natural_week",
        "items": items,
        "fetchedAt": time.time(),
        "source": "history",
        "missing_days": missing,
        "coverage": {"days": 7, "have": len(have_days)},
    }


@app.get("/api/refresh")
def refresh(request: Request):
    """触发后台刷新全部店铺数据(轮询三个集团, 恢复原激活集团)"""
    start, end = date_range(7)
    _task_label = "刷新全部店铺数据"
    _task_params = {"start": start, "end": end}
    # 多用户队列: 有任意任务在跑 → 入队排队, 不 409 拒绝
    _queued = _queue_if_busy(request, "refresh", _task_label, _task_params)
    if _queued:
        return _queued
    with _lock:
        if _any_task_running():
            if request.headers.get("x-scheduler-task") == "1":
                raise HTTPException(409, "任务进行中, 调度器稍后重试")
            _task = _enqueue_task("refresh", _task_label, _operator_label(request), _task_params, _operator_role(request))
            return {"status": "queued", "taskId": _task["id"],
                    "queuePosition": _queue_position(_task["id"]), "task": _task_public(_task)}
        _refresh_state["triggered_by"] = _operator_label(request)
        # 同步置位 running: 不等 worker 首个 switch_group 完成, 杜绝窗口期双启动
        _refresh_state["running"] = True
        _refresh_state["error"] = None
        _refresh_state["progress"] = {"done": 0, "total": 0, "current": "准备中"}
    refresh_all_groups_async(start, end)
    # 直接启动(非调度器)也登记任务列表: 前端显示"谁发起的 + 进度条"
    if request.headers.get("x-scheduler-task") != "1":
        _register_running_task(_refresh_state, "refresh", _task_label,
                               _operator_label(request), _task_params, _operator_role(request))
    return {"ok": True, "message": "刷新任务已启动(三集团轮询)"}


@app.post("/api/trace/prefetch")
def trace_prefetch_start(request: Request, payload: dict = Body(...)):
    """手动补抓数据: 用户指定历史日期区间, 逐集团抓取 trace 写库

    用于夜间自动抓取失败/漏抓时人工补数据。复用 _prefetch_group 逐集团增量抓取
    (history_mode=True 已抓天复用零请求, 与夜间任务同源同节奏)。
    start/end 均可选, 默认昨天~昨天; 不可含今天(当天数据走实时不入库)。
    """
    import datetime as _dt
    yesterday = (_dt.date.today() - _dt.timedelta(days=1)).isoformat()
    start = (payload.get("start") or yesterday).strip()
    end = (payload.get("end") or yesterday).strip()
    # 补抓只接受纯日期 YYYY-MM-DD(历史按天粒度); 不能用 _valid_datetime_iso(接受
    # 秒级 'T' 串)校验, 否则 "2026-08-10T00:00" 过校验后被 date.fromisoformat 解析抛 500
    for v in (start, end):
        if not _valid_date_iso(v):
            raise HTTPException(400, f"日期格式非法(需 YYYY-MM-DD): {v}")
    s, e = _dt.date.fromisoformat(start), _dt.date.fromisoformat(end)
    if s > e:
        raise HTTPException(400, f"开始日期不能晚于结束日期: {start} > {end}")
    today = _dt.date.today()
    if e >= today:
        raise HTTPException(400, f"结束日期不能是今天或未来: {end}(当天数据走实时, 不写入历史)")
    span = (e - s).days + 1
    if span > 35:
        raise HTTPException(400, f"区间超过 35 天上限: {span} 天")
    # 校验集团可用
    cfg = load_config()
    groups = [g for g in (cfg.get("groups") or [])
              if g.get("groupId") and g.get("accountId")]
    if not groups:
        raise HTTPException(400, "config.groups 为空或无可用集团, 请先登录")
    _task_label = f"补抓历史数据 · {start} ~ {end}"
    _task_params = {"start": start, "end": end}
    # 多用户队列: 有任意任务在跑 → 入队排队(前端显示"排队中"), 不 409 拒绝
    _queued = _queue_if_busy(request, "prefetch", _task_label, _task_params)
    if _queued:
        return _queued
    # check+set 原子化(锁内): 并发双开补抓会互切集团 cookie 串数据
    with _lock:
        if _any_task_running():
            if request.headers.get("x-scheduler-task") == "1":
                raise HTTPException(409, "任务进行中, 调度器稍后重试")
            _task = _enqueue_task("prefetch", _task_label, _operator_label(request), _task_params, _operator_role(request))
            return {"status": "queued", "taskId": _task["id"],
                    "queuePosition": _queue_position(_task["id"]), "task": _task_public(_task)}
        _prefetch_state.update({
            "running": True, "canceled": False, "error": None,
            "start_date": start, "end_date": end,
            "progress": {"done": 0, "total": 0, "current": "准备中…"},
            "started_at": time.time(), "last_run": None,
            "triggered_by": _operator_label(request),
        })
        # 直接启动(非调度器)也登记任务列表: 前端显示"谁发起的 + 进度条"
        if request.headers.get("x-scheduler-task") != "1":
            _register_running_task(_prefetch_state, "prefetch", _task_label,
                                   _operator_label(request), _task_params, _operator_role(request))

    def worker():
        g_total = g_failed = 0
        try:
            set_nightly_fetch_flag()  # 手动抓取期间 8080 直连请求阻塞(不串集团 cookie)
            # 按平台优先级排序集团(1 拼多多, 5 抖音, 7 京东), 与夜间任务一致
            platform_order = cfg.get("prefetch_platforms") or [1, 5, 7]
            plat_rank = {p: i for i, p in enumerate(platform_order)}
            ordered = sorted(groups, key=lambda g: plat_rank.get(g.get("platform"), 99))
            try:
                total_shops = len([s for s in load_all_shops()
                                   if s.get("platform") in FETCH_PLATFORMS])
            except Exception:
                total_shops = 0
            _prefetch_state["progress"] = {"done": 0, "total": total_shops, "current": ""}

            def _cb(p):
                _prefetch_state["progress"] = {
                    "done": p["done"], "total": total_shops or p["done"], "current": p["current"],
                }

            # 区间内最近 force_days 天强制重抓(tanyu 回溯更新 sendType, 补抓要跟随
            # 最新口径; 与夜间任务同源: config.prefetch_force_days, 默认7)
            _force_days_n = int(cfg.get("prefetch_force_days", 1) or 0)
            _fset = set()
            if _force_days_n > 0:
                from datetime import timedelta as _td
                for i in range(min(_force_days_n, (e - s).days + 1)):
                    _fset.add((e - _td(days=i)).isoformat())
            for g in ordered:
                if _prefetch_state.get("canceled"):
                    log_line("prefetch", "⏹ 手动补抓被用户取消")
                    break
                if _risk_state.get("triggered"):
                    log_line("prefetch", "⛔ 风控/登录失效, 手动补抓整体停止")
                    _prefetch_state["error"] = _risk_state.get("reason") or "风控/登录失效"
                    break
                try:
                    n, f = _prefetch_group(g["groupId"], start, end,
                                           force_days=_fset or None, progress_cb=_cb,
                                           done_shops=g_total,
                                           cancel_check=lambda: bool(_prefetch_state.get("canceled")))
                    g_total += n
                    g_failed += f
                except RiskTriggered:
                    log_line("prefetch", "⛔ 风控触发, 手动补抓整体停止")
                    _prefetch_state["error"] = _risk_state.get("reason") or "风控/登录失效"
                    break
                except Exception as e:
                    log_line("prefetch", f"⚠️ 集团「{g.get('groupName')}」补抓失败, 跳过: {e}")
                    continue
            if not _prefetch_state.get("canceled") and not _prefetch_state.get("error"):
                log_line("prefetch", f"手动补抓完成: {g_total} 家店铺, {g_failed} 家失败, "
                                     f"区间 {start} ~ {end}, 耗时 {time.time() - _prefetch_state['started_at']:.0f}s")
        except Exception as e:
            log_line("prefetch", f"⚠️ 手动补抓异常: {e}")
            _prefetch_state["error"] = str(e)
        finally:
            _prefetch_state["running"] = False
            _prefetch_state["last_run"] = time.time()
            clear_nightly_fetch_flag()

    threading.Thread(target=worker, daemon=True).start()
    return {"ok": True, "message": f"手动补抓已启动: {start} ~ {end}",
            "start_date": start, "end_date": end}


@app.get("/api/trace/prefetch/status")
def trace_prefetch_status(request: Request):
    """手动补抓进度/状态(供前端轮询; 附带夜间抓取标志)"""
    st = dict(_prefetch_state)
    st["nightly"] = _nightly_flag_info()
    if st.get("triggered_by"):
        st["triggered_by"] = _mask_admin_label(request, st["triggered_by"])
    return st


@app.post("/api/trace/prefetch/cancel")
def trace_prefetch_cancel():
    """取消进行中的手动补抓(worker 每店边界检查标志后停止)"""
    if not _prefetch_state.get("running"):
        return {"ok": False, "message": "当前无手动补抓任务"}
    _prefetch_state["canceled"] = True
    log_line("prefetch", "⏹ 收到取消请求, 正在停止…")
    return {"ok": True, "message": "已请求取消"}


# ---------- 数据校准(强制重抓最近 N 天, 同步 tanyu 人工变更) ----------
@app.post("/api/trace/calibrate")
def trace_calibrate_start(request: Request, payload: dict = Body(...),
                          _: dict = Depends(auth.require_admin)):
    """数据校准: 对最近 N 天(不含今天)全部店铺强制重抓, 忽略缓存。

    背景: tanyu 会对已抓消息回溯更新 sendType(客服补发/编辑后 草稿→已发送),
    夜间预抓虽按 prefetch_force_days 重抓最近 N 天, 但"白天的人工变更"要到次日
    凌晨才同步; 此接口用于随时一键校准, 让看板采纳率与 tanyu 后台对齐。
    与手动补抓共用队列/状态槽(_prefetch_state), 串行执行; 遇风控立即停止。
    """
    import datetime as _dt
    cfg = load_config()
    days = int(payload.get("days") or cfg.get("prefetch_force_days", 1) or 1)
    days = max(1, min(days, 7))
    groups = [g for g in (cfg.get("groups") or [])
              if g.get("groupId") and g.get("accountId")]
    if not groups:
        raise HTTPException(400, "config.groups 为空或无可用集团, 请先登录")
    today = _dt.date.today()
    end = (today - _dt.timedelta(days=1)).isoformat()
    start = (today - _dt.timedelta(days=days)).isoformat()
    _task_label = f"数据校准 · 最近 {days} 天({start} ~ {end})"
    _task_params = {"days": days, "start": start, "end": end}
    _queued = _queue_if_busy(request, "prefetch", _task_label, _task_params)
    if _queued:
        return _queued
    with _lock:
        if _any_task_running():
            _task = _enqueue_task("prefetch", _task_label, _operator_label(request),
                                  _task_params, _operator_role(request))
            return {"status": "queued", "taskId": _task["id"],
                    "queuePosition": _queue_position(_task["id"]), "task": _task_public(_task)}
        _prefetch_state.update({
            "running": True, "canceled": False, "error": None,
            "start_date": start, "end_date": end,
            "progress": {"done": 0, "total": 0, "current": "准备中…"},
            "started_at": time.time(), "last_run": None,
            "triggered_by": _operator_label(request),
        })
        _register_running_task(_prefetch_state, "prefetch", _task_label,
                               _operator_label(request), _task_params, _operator_role(request))

    def worker():
        g_total = g_failed = 0
        try:
            set_nightly_fetch_flag()
            try:
                total_shops = len([s for s in load_all_shops()
                                   if s.get("platform") in FETCH_PLATFORMS])
            except Exception:
                total_shops = 0
            _prefetch_state["progress"] = {"done": 0, "total": total_shops, "current": ""}

            def _cb(p):
                _prefetch_state["progress"] = {
                    "done": p["done"], "total": total_shops or p["done"], "current": p["current"],
                }

            # 校准窗口最近 days 天全部强制重抓(忽略缓存)
            fset = {(_dt.date.fromisoformat(end) - _dt.timedelta(days=i)).isoformat()
                    for i in range(days)}
            platform_order = cfg.get("prefetch_platforms") or [1, 5, 7]
            plat_rank = {p: i for i, p in enumerate(platform_order)}
            ordered = sorted(groups, key=lambda g: plat_rank.get(g.get("platform"), 99))
            for g in ordered:
                if _prefetch_state.get("canceled"):
                    log_line("prefetch", "⏹ 数据校准被用户取消")
                    break
                if _risk_state.get("triggered"):
                    log_line("prefetch", "⛔ 风控/登录失效, 数据校准整体停止")
                    _prefetch_state["error"] = _risk_state.get("reason") or "风控/登录失效"
                    break
                try:
                    n, f = _prefetch_group(g["groupId"], start, end,
                                           force_days=fset, progress_cb=_cb,
                                           done_shops=g_total,
                                           cancel_check=lambda: bool(_prefetch_state.get("canceled")))
                    g_total += n
                    g_failed += f
                except RiskTriggered:
                    log_line("prefetch", "⛔ 风控触发, 数据校准整体停止")
                    _prefetch_state["error"] = _risk_state.get("reason") or "风控/登录失效"
                    break
                except Exception as e:
                    log_line("prefetch", f"⚠️ 集团「{g.get('groupName')}」校准失败, 跳过: {e}")
                    continue
            if not _prefetch_state.get("canceled") and not _prefetch_state.get("error"):
                log_line("prefetch", f"数据校准完成: 最近 {days} 天({start} ~ {end}), "
                                     f"{g_total} 家店铺, {g_failed} 家失败, "
                                     f"耗时 {time.time() - _prefetch_state['started_at']:.0f}s")
        except Exception as e:
            log_line("prefetch", f"⚠️ 数据校准异常: {e}")
            _prefetch_state["error"] = str(e)
        finally:
            _prefetch_state["running"] = False
            _prefetch_state["last_run"] = time.time()
            clear_nightly_fetch_flag()

    threading.Thread(target=worker, daemon=True).start()
    return {"ok": True, "message": f"数据校准已启动: 最近 {days} 天({start} ~ {end})",
            "start_date": start, "end_date": end}


# ---------- 抓取配置 + tanyu 抓取账号(管理员) ----------
_TRACE_COOKIE_KEYS = ["tanyu-account-id", "tanyu-agent-account",
                      "tanyu-group-account", "tanyu-group-id"]


@app.get("/api/config/trace")
def trace_config_get(_: dict = Depends(auth.require_admin)):
    """抓取配置 + tanyu 登录账号信息(凭据脱敏, 不暴露 cookie 值)"""
    cfg = load_config()
    cookies = cfg.get("cookies") or {}
    account_id = str(cookies.get("tanyu-account-id") or "")
    if len(account_id) > 8:
        masked = account_id[:4] + "****" + account_id[-3:]
    elif account_id:
        masked = account_id[:2] + "****"
    else:
        masked = "未配置"
    cur = get_current_group()
    return {
        "prefetch_days": cfg.get("prefetch_days", 7),
        "prefetch_force_days": cfg.get("prefetch_force_days", 1),
        "prefetch_platforms": cfg.get("prefetch_platforms", [1, 5, 7]),
        "prefetch_windows": cfg.get("prefetch_windows", {}),
        "tanyu": {
            "account_id": masked,
            "account_configured": bool(account_id),
            "cookie_configured": {k: bool(cookies.get(k)) for k in _TRACE_COOKIE_KEYS},
            "cookie_expires": dict(cfg.get("cookie_expires") or {}),
        },
        "current_group": {
            "id": cur.get("id") if cur else None,
            "name": cur.get("name") if cur else None,
        },
    }


@app.post("/api/config/trace")
def trace_config_set(payload: dict = Body(...), _: dict = Depends(auth.require_admin)):
    """保存抓取配置(管理员): prefetch_force_days(0~7) / prefetch_days(1~35)"""
    def _apply(cfg):
        v = payload.get("prefetch_force_days")
        if v is not None:
            v = int(v)
            if not (0 <= v <= 7):
                raise HTTPException(400, "prefetch_force_days 需为 0~7(0=不强制重抓)")
            cfg["prefetch_force_days"] = v
        v = payload.get("prefetch_days")
        if v is not None:
            v = int(v)
            if not (1 <= v <= 35):
                raise HTTPException(400, "prefetch_days 需为 1~35")
            cfg["prefetch_days"] = v

    cfg = mutate_config(_apply)
    return {"ok": True,
            "prefetch_force_days": cfg["prefetch_force_days"],
            "prefetch_days": cfg["prefetch_days"]}


@app.get("/api/trace/shop/{shop_id}")
def trace_shop(shop_id: str, days: int = 7, force: int = 0, start: str | None = None, end: str | None = None):
    """单店核算: 遍历消息轨迹统计核算采纳率(带按日缓存)

    start/end 可选: 自定义起止日期(YYYY-MM-DD), 不传则用近 N 天
    按日增量缓存: 已抓取的天零请求, 只补抓缺失的天
    """
    shops = {s["thirdShopId"]: s for s in load_all_shops()}
    shop = shops.get(shop_id)
    if not shop:
        raise HTTPException(404, "店铺不存在")
    start, end = date_range(days, end) if not start else (start, end or (datetime.date.today() - datetime.timedelta(days=1)).isoformat())
    if start and not _valid_datetime_iso(start):
        raise HTTPException(400, f"start 格式非法: {start}")
    if end and not _valid_datetime_iso(end):
        raise HTTPException(400, f"end 格式非法: {end}")
    if start and end and start > end:
        raise HTTPException(400, f"开始时间不能晚于结束时间: {start} > {end}")
    _check_span_limit(start, end)
    _sms, _ems, _sday, _eday, _ht = _split_bounds(start, end)
    today_str = datetime.date.today().isoformat()
    # 导入平台(天猫1/2)店: 纯 DB 天级聚合, 无原始消息, 绝不发 tanyu 请求
    if _is_import_shop(shop):
        stat = _import_shop_stat(shop_id, start, end)
        return {"shop": shop, "startDate": start, "endDate": end,
                "fetchedAt": time.time(), "stat": stat, "live": False}
    # 区间含今天: 历史(库) + 今天(实时) 合并, 保留实时数据(day 边界判定)
    if (not force and _use_sqlite_trace()
            and today_str >= _sday and today_str <= _eday):
        try:
            merged = _trace_shop_merged(shop_id, start, end, _sms, _ems, _ht)
            if merged and merged[0]:
                stat, has_today = merged
                return {"shop": shop, "startDate": start, "endDate": end,
                        "fetchedAt": time.time(), "stat": stat,
                        "live": has_today}
        except RiskTriggered:
            raise HTTPException(502, "风控/登录失效, 请重新登录后重试")
        except Exception as e:
            print(f"[trace] SQLite+实时 合并路径失败, 回退在线抓取: {e}")
    # SQLite 快路径: 数据库已覆盖该区间(纯历史) => 纯本地聚合(核算提速主因)
    # 覆盖判定限该平台该店: 导入平台(10/11)无限期数据会撑大全库窗口, 全库口径
    # 会让抓取平台早于其窗口的自定义区间误判已覆盖、全零短路在线抓取。
    if (not force and _use_sqlite_trace()
            and trace_store.db_window_covers(_sday, _eday,
                                             platform=shop.get("platform"),
                                             shop_filter={shop_id})):
        try:
            stat = _trace_shop_from_db(shop_id, start, end)
            if stat:
                return {"shop": shop, "startDate": start, "endDate": end,
                        "fetchedAt": time.time(), "stat": stat}
        except Exception as e:
            print(f"[trace] SQLite 单店快路径失败, 回退在线抓取: {e}")
    # 新式按日缓存: 全部天都命中(且未过期)则零请求
    cache = _load_trace_days_cache(shop_id) if not force else None
    if cache and cache.get("days") and not force:
        start_d = datetime.date.fromisoformat(_sday)
        end_d = datetime.date.fromisoformat(_eday)
        need = {(start_d + datetime.timedelta(days=i)).isoformat()
                for i in range((end_d - start_d).days + 1)}
        if all(_cached_days_usable(cache, ds) for ds in need):
            stat = stat_trace_daily(shop_id, start, end, trim_ms=(_sms, _ems))  # 全命中: 纯聚合零请求
            return {"shop": shop, "startDate": start, "endDate": end,
                    "fetchedAt": time.time(), "stat": stat}
    try:
        stat = stat_trace_daily(shop_id, start, end, force=bool(force), trim_ms=(_sms, _ems))
    except BusyQueueError as e:
        raise HTTPException(503, str(e))  # 夜间抓取进行中
    except RuntimeError as e:
        raise HTTPException(502, str(e))
    return {"shop": shop, "startDate": start, "endDate": end, "fetchedAt": time.time(), "stat": stat}


# 核算总览状态(异步任务)
TRACE_OVERVIEW_CACHE_FILE = DATA_DIR / "trace_overview_result.json"
_trace_state = {
    "running": False,
    "progress": {"done": 0, "total": 0, "current": ""},
    "result": None,  # 最近一次完成的全量核算结果
    "subset_result": None,  # 最近一次店铺子集核算结果(不覆盖全量视图, 不落盘)
    "start_date": None,
    "end_date": None,
    "platform": None,
    "last_run": None,
    "error": None,
    "paused": False,       # 核算暂停标志(worker 在每店边界阻塞等待恢复)
    "canceled": False,     # 取消核算标志(worker 中断后续店铺)
    "partial_list": [],    # 运行中已完成店铺的统计(浅拷贝), 供前端边算边展示
    "shop_filter": None,   # 本次核算的店铺子集(账号池勾选), None=全部店铺
    "triggered_by": None,  # 谁发起的核算(操作者用户名/昵称), 供状态展示
}
_trace_resume_evt = threading.Event()

# 今日实时抓取状态(独立于核算: 抓今天不写库, 供核算/人工客服板块「抓取今日数据」)
_today_state = {
    "running": False,
    "progress": {"done": 0, "total": 0, "current": ""},
    "result": None,
    "start_date": None,
    "end_date": None,
    "start_ts": "00:00",
    "end_ts": "",
    "platform": None,
    "shop_filter": None,
    "last_run": None,
    "error": None,
    "triggered_by": None,  # 谁发起的今日抓取(操作者用户名/昵称), 供状态展示
}


TRACE_OVERVIEW_CACHE_TTL = 6 * 3600   # 总览磁盘缓存有效期(探域数据每日更新)


def load_trace_overview_cache(start, end, platform=None):
    """从磁盘读核算总览结果(同时间段重复核算零请求, 且未过期)

    秒级区间(带 'T')不走盘: 磁盘是单文件只存一份 startDate/endDate,
    秒级区间几乎必 miss 且会覆盖掉"近7天"等日级缓存 → 一律返回 None。
    """
    if "T" in start or "T" in end:
        return None
    if not TRACE_OVERVIEW_CACHE_FILE.exists():
        return None
    try:
        c = json.loads(TRACE_OVERVIEW_CACHE_FILE.read_text(encoding="utf-8"))
    except Exception:
        return None
    if c.get("startDate") == start and c.get("endDate") == end and c.get("platform") == platform:
        cached_at = c.get("fetched_at") or c.get("last_run")
        if not cached_at:
            # 旧格式缓存没有时间戳: 用文件 mtime 兜底, 避免永久缓存
            try:
                cached_at = TRACE_OVERVIEW_CACHE_FILE.stat().st_mtime
            except Exception:
                cached_at = None
        if cached_at and time.time() - cached_at > TRACE_OVERVIEW_CACHE_TTL:
            return None
        return c  # 保存的本身就是 result 字典(含 startDate/endDate/platform 字段)
    return None


def load_latest_trace_overview_cache():
    """读磁盘里最近一次核算结果(不分时间段)"""
    if not TRACE_OVERVIEW_CACHE_FILE.exists():
        return None
    try:
        c = json.loads(TRACE_OVERVIEW_CACHE_FILE.read_text(encoding="utf-8"))
    except Exception:
        return None
    return c if "total" in c else None


def save_trace_overview_cache(result):
    # 秒级区间不落盘(单文件会被覆盖), 只存日级 presets
    if "T" in (result.get("startDate") or "") or "T" in (result.get("endDate") or ""):
        return
    try:
        _atomic_write_text(TRACE_OVERVIEW_CACHE_FILE,
                           json.dumps(result, ensure_ascii=False, separators=(",", ":")))
    except Exception:
        pass


def _trace_shop_from_db(shop_id, start, end):
    """从 SQLite 聚合单店 stat(与 stat_trace_daily 输出结构完全一致)

    支持秒级 start/end: end_ms 由 _split_bounds 统一给出(纯日期=整天 23:59:59.999,
    带时间=该时刻+999ms), 不再无脑 +1day(否则秒级 end 会多算一天)。
    """
    sms, ems, sday, eday, _ = _split_bounds(start, end)
    rows = trace_store.query_shop_aggregate(shop_id, sday, eday, sms, ems)
    if not rows and not trace_store.query_daily(shop_id, sday, eday):
        return None
    # 复用 stat_trace_daily 的聚合逻辑(保证口径一致)
    return _aggregate_stat_rows(rows, start, end)


def _trace_shop_merged(shop_id, start, end, start_ms=None, end_ms=None, has_time=False):
    """历史天走 SQLite + 今天实时抓取, 合并出 [start, end] 的单店 stat

    口径: 完整历史天(含昨天)读库聚合(与探域逐日口径一致);
      "今天"未定型, 实时抓取当天消息(不走 6h 缓存), 只读不写库。
    秒级区间(has_time=True): 历史库按 msg_time 毫秒精确过滤, 今天实时抓取整天后
      再按 [start_ms, end_ms] 裁剪(今天 09:00 起时只算 09:00 后的消息)。
    返回 (stat_dict, has_today_flag); 实时抓取抛 RiskTriggered 时向上传播。
    """
    today_str = datetime.date.today().isoformat()
    yesterday_str = (datetime.date.today() - datetime.timedelta(days=1)).isoformat()
    if has_time:
        # 已由外层 _split_bounds 归一化; 未传则兜底
        if start_ms is None:
            start_ms, end_ms, sday, eday, _ = _split_bounds(start, end)
        else:
            sday, eday = start[:10], end[:10]
        # 今天判定用 day 边界(字符串字典序对带 'T' 的 start 会恒 False)
        has_today = today_str >= sday and today_str <= eday
        hist_end = end if end < today_str else yesterday_str  # 语义: end 带时间仍按 day 比
        hist_end_day = hist_end[:10] if has_time else hist_end
        rows = []
        if start[:10] <= hist_end_day:
            rows = trace_store.query_shop_aggregate(
                shop_id, start[:10], hist_end_day, start_ms, end_ms)
    else:
        has_today = start <= today_str <= end
        # 历史部分: start .. min(end, 昨天)(ISO 日期字符串字典序可比较)
        hist_end = end if end < today_str else yesterday_str
        rows = []
        if start <= hist_end:
            rows = trace_store.query_shop_aggregate(
                shop_id, start, hist_end,
                int(datetime.datetime.fromisoformat(start).timestamp() * 1000),
                int((datetime.datetime.fromisoformat(hist_end) + datetime.timedelta(days=1)).timestamp() * 1000) - 1,
            )
    # 今天实时(未定型日不进库, 只读合并)
    live = []
    if has_today:
        live_res = _fetch_trace_day(shop_id, today_str)
        if live_res is None:
            live_res = []  # 今天暂无数据/请求失败, 历史部分照常返回
        live = [_trim_trace_msg(m) for m in live_res]
        # 秒级区间: 今天实时抓的是整天, 按 ms 边界裁剪(只留 [start_ms, end_ms])
        if has_time and start_ms is not None:
            live = [
                m for m in live
                if start_ms <= (m.get("time") or m.get("createTime") or m.get("createAt") or 0) <= end_ms
            ]
        rows.extend(live)
    # 区间含今天时恒返回 stat(即使全空), 避免回退到会写库的 stat_trace_daily;
    # 纯历史且无数据才返回 None(由调用方走在线抓取兜底)
    if not has_today and not rows and not live:
        return None, False
    return _aggregate_stat_rows(rows, start, end), has_today


def _aggregate_stat_rows(results, start, end):
    """按 stat_trace_daily 完全相同的口径聚合(独立函数, 供 JSON/SQLite 两路共用)"""
    counts = {1: 0, 2: 0, 3: 0, None: 0}
    by_staff = {}
    by_type = {}
    daily = {}
    total = 0
    for r in results:
        st = r.get("sendType")
        counts[st] = counts.get(st, 0) + 1
        staff = r.get("sellerAccount") or "未知"
        entry = by_staff.setdefault(staff, {"total": 0, "adopted": 0})
        entry["total"] += 1
        if st in ADOPTED_SEND_TYPES:
            entry["adopted"] += 1
        t = r.get("type") or "OTHER"
        by_type[t] = by_type.get(t, 0) + 1
        ts = r.get("time")
        if isinstance(ts, (int, float)):
            day = datetime.datetime.fromtimestamp(ts / 1000).strftime("%Y-%m-%d")
            de = daily.setdefault(day, {"total": 0, "adopted": 0, "replies": 0})
            de["total"] += 1
            if st in ADOPTED_SEND_TYPES:
                de["adopted"] += 1
            if t == "CONSULT_REPLY":
                de["replies"] += 1
        total += 1
    adopted = sum(counts.get(s, 0) for s in ADOPTED_SEND_TYPES)
    rate = (adopted / total * 100) if total else 0
    daily_list = [
        {"date": d, "total": v["total"], "adopted": v["adopted"], "replies": v["replies"],
         "rate": round(v["adopted"] / v["total"] * 100, 2) if v["total"] else 0}
        for d, v in sorted(daily.items())
    ]
    staff_list = _staff_list_from_agg(by_staff)
    raw_messages = [
        {
            "time": r.get("createTime") or r.get("createAt") or r.get("time") or "",
            "buyer": (r.get("buyerNick") or r.get("customerName") or r.get("userName") or ""),
            "sendType": r.get("sendType"),
            "content": (r.get("content") or r.get("replyContent") or r.get("question") or "")[:200],
            "staff": r.get("sellerAccount") or r.get("staffName") or "",
            "type": r.get("type") or "OTHER",
            "traceId": r.get("traceId") or r.get("id") or "",
        }
        for r in reversed(results)
    ]
    return {
        "total": total,
        "counts": counts,
        "adopted": adopted,
        "rate": round(rate, 2),
        "byStaff": staff_list,
        "byType": by_type,
        "daily": daily_list,
        "messages": raw_messages,
    }


def _avg_daily_rates(daily):
    """每天采纳率/回复率的算术平均(有消息的天才计入); 返回 (avgAdoptRate, avgReplyRate)"""
    n = ar = rr = 0
    for d in daily or []:
        t = d.get("total") or 0
        if t <= 0:
            continue
        n += 1
        ar += (d.get("adopted") or 0) / t * 100
        rr += (d.get("replies") or 0) / t * 100
    return (round(ar / n, 2) if n else 0), (round(rr / n, 2) if n else 0)


def _import_shop_stat(shop_id, start, end):
    """导入平台(天猫1/2)单店 stat: 从 trace_daily 聚合(天级), 无原始消息

    start/end 用 day 边界字符串(秒级区间对导入店按整天算——Excel 是天级粒度)。
    byStaff 从 staff_aggregate_per_shop(按店铺客服聚合, 读 by_staff_json)取该店行;
    daily 用 query_daily 逐日; messages 恒空(导入数据无原始消息)。
    """
    sms, ems, sday, eday, _ = _split_bounds(start, end)
    days = trace_store.query_daily(shop_id, sday, eday)
    if not days:
        return None
    total = sum(d["total"] for d in days)
    adopted = sum(d["adopted"] for d in days)
    # 生成率: 区间内该店非空 generation_rate 的均值(Excel 生成率列); 无值返回 None
    gen_rate = trace_store.avg_generation_rate(sday, eday, platform=None,
                                               shop_filter={shop_id})
    per_shop = trace_store.staff_aggregate_per_shop(sday, eday, platform=None,
                                                    shop_filter={shop_id})
    staff_map = per_shop["by_shop"].get(shop_id, {})
    daily_list = [
        {"date": d["day"], "total": d["total"], "adopted": d["adopted"],
         "rate": round(d["adopted"] / d["total"] * 100, 2) if d["total"] else 0}
        for d in days
    ]
    staff_list = [{"account": acct, "total": v.get("total", 0), "adopted": v.get("adopted", 0),
                   "rate": round(v.get("adopted", 0) / v.get("total", 0) * 100, 2) if v.get("total", 0) else 0}
                  for acct, v in sorted(staff_map.items())]
    return {
        "total": total,
        "counts": {1: 0, 2: 0, 3: 0, None: 0},
        "adopted": adopted,
        "rate": round(adopted / total * 100, 2) if total else 0,
        "generation_rate": round(gen_rate * 100, 2) if gen_rate is not None else None,
        "byStaff": staff_list,
        "byType": {},
        "daily": daily_list,
        "messages": [],
    }


def _is_import_shop(shop):
    """店铺是否导入平台(天猫1/2)"""
    return bool(shop and shop.get("platform") in IMPORT_PLATFORMS)


def _filtered_platforms(plats, shop_filter):
    """子集核算时把覆盖判定平台收窄到勾选店铺所属的平台(避免未勾选平台拖垮判定)"""
    if not shop_filter:
        return plats
    try:
        have = {s.get("platform") for s in trace_store.get_shops()
                if s.get("thirdShopId") in shop_filter}
    except Exception:
        return plats
    sub = [p for p in plats if p in have]
    return sub or plats


def _trace_overview_from_db_multi(start, end, plats, shop_filter=None, has_time=False):
    """跨平台 DB 聚合: 逐平台调用单平台聚合后合并(店铺列表拼接, 汇总累加)"""
    merged = None
    for p in plats:
        part = _trace_overview_from_db(start, end, p, shop_filter, has_time=has_time)
        if not part:
            continue
        if merged is None:
            merged = part
        else:
            merged["shopList"] += part["shopList"]
            merged["total"] += part["total"]
            merged["adopted"] += part["adopted"]
            merged["byStaff"] += part["byStaff"]
    if merged is None:
        return None
    merged["rate"] = round(merged["adopted"] / merged["total"] * 100, 2) if merged["total"] else 0
    merged["platform"] = ",".join(str(p) for p in sorted(plats))
    return merged


def _trace_overview_from_db(start, end, platform=None, shop_filter=None, has_time=False):
    """从 SQLite 聚合核算总览(与在线 worker 输出结构完全一致)

    shop_filter: 可选店铺子集(集合), 只聚合勾选的店铺。
    店铺池 = 跨集团店铺表(全部店铺)按 platform/shop_filter 过滤后的全集:
    零消息店铺(total=0)也保留在 shopList 里, 与在线 worker 口径一致。
    has_time=True 时区间带秒级边界: 改从 messages 表按 msg_time 毫秒过滤聚合
      (trace_daily 只有整天, 秒级边界会把整天算进来); 纯日期仍走 trace_daily day 版。
    """
    if has_time:
        sms, ems, sday, eday, _ = _split_bounds(start, end)
        agg = trace_store.overview_aggregate_ms(sms, ems, platform, shop_filter)
    else:
        agg = trace_store.overview_aggregate(start, end, platform, shop_filter)
    # 用跨集团店铺表(全部 118 家)解析店铺元数据并枚举店铺全集, 而非当前激活
    # 集团的 shops.json(否则非当前集团店铺 platform 为 None、无法展示平台标签)
    shop_map = {s["thirdShopId"]: s for s in trace_store.get_shops()}
    # 抓取平台 + 导入平台都进 DB 快路径(导入平台靠 Excel 数据, 无 tanyu 抓取能力)
    scope = [s for s in shop_map.values() if s.get("platform") in FETCH_PLATFORMS + IMPORT_PLATFORMS]
    if platform is not None:
        scope = [s for s in scope if s.get("platform") == platform]
    if shop_filter:
        scope = [s for s in scope if s["thirdShopId"] in shop_filter]
    if not scope:
        return None
    data = {sid: v for sid, v in agg["shop_list"]}
    shop_list = []
    agg_total = agg_adopted = 0
    for shop in scope:
        v = data.get(shop["thirdShopId"], {"total": 0, "adopted": 0, "replies": 0})
        shop_list.append({"shop": shop, "total": v["total"],
                          "adopted": v["adopted"],
                          "replies": v.get("replies", 0),
                          "avgAdoptRate": v.get("avgAdoptRate", 0),
                          "avgReplyRate": v.get("avgReplyRate", 0),
                          "rate": round(v["adopted"] / v["total"] * 100, 2) if v["total"] else 0,
                          "startDate": start, "endDate": end})
        agg_total += v["total"]
        agg_adopted += v["adopted"]
    # 客服按 (店铺,客服) 组合区分(不去重), 用同一区间/店铺集的按店聚合
    if has_time:
        per_shop = trace_store.staff_aggregate_per_shop_ms(sms, ems, platform, shop_filter)
    else:
        per_shop = trace_store.staff_aggregate_per_shop(start, end, platform, shop_filter)
    staff_list = _staff_list_per_shop(per_shop["by_shop"], platform, shop_filter)
    return {
        "startDate": start,
        "endDate": end,
        "platform": platform,
        "fetched_at": time.time(),
        "total": agg_total,
        "adopted": agg_adopted,
        "rate": round(agg_adopted / agg_total * 100, 2) if agg_total else 0,
        "shopList": shop_list,
        "byStaff": staff_list,
    }


@app.get("/api/trace/overview")
def trace_overview(request: Request, days: int = 7, platform: str | None = None, force: int = 0,
                   start: str | None = None, end: str | None = None,
                   from_cache: int = 0, shop_ids: str | None = None):
    """核算总览(异步): 遍历抓取店铺统计核算采纳率

    支持自定义时间段(start/end, YYYY-MM-DD 或 YYYY-MM-DDTHH:MM 秒级); 不传则用近 N 天。
    platform: 平台筛选, 支持逗号分隔多选(如 "1,5"), 空/省略=全部抓取平台。
    shop_ids: 逗号分隔的店铺 ID 子集(账号池勾选后只核算勾选店铺)。
    has_shop_filter: 勾选了店铺子集时强制走在线路径(不命中全平台缓存)。
    有缓存直接返回; 否则触发后台任务, 返回任务状态,
    前端轮询 /api/trace/overview/status 获取进度, 完成后取 result。
    from_cache=1: 只查内存+磁盘缓存, 未命中直接返回空(不触发任务)
    """
    global _trace_state
    # 店铺子集: 解析 shop_ids 参数(逗号分隔)
    shop_filter = None
    if shop_ids and shop_ids.strip():
        shop_filter = {s for s in shop_ids.split(",") if s.strip()}
    # 平台多选解析: "1,5" → [1,5]; 空/None → None(全部抓取平台)
    plats = None
    if platform and str(platform).strip():
        try:
            plats = sorted({int(x) for x in str(platform).split(",") if x.strip()})
        except (TypeError, ValueError):
            raise HTTPException(400, f"平台参数非法: {platform}")
        for p in plats:
            if p not in PLATFORM_NAMES:
                raise HTTPException(400, f"不支持的平台: {p}")
    # 缓存/状态键: 单平台用 int(兼容旧缓存), 多平台用 "1,5" 串, 全平台 None
    plat_key = plats[0] if plats and len(plats) == 1 else (",".join(str(p) for p in plats) if plats else None)
    single_plat = plats[0] if plats and len(plats) == 1 else None
    plat_label = "全平台" if not plats else "+".join(PLATFORM_NAMES.get(p, str(p)) for p in plats)
    # 并发互斥不在顶部拒绝: 缓存命中路径不需要任务槽; 真正要跑 worker 时才
    # 检查队列(有任务在跑则入队排队, 见启动点 _queue_if_busy)。
    start, end = date_range(days, end) if not start else (start, end or (datetime.date.today() - datetime.timedelta(days=1)).isoformat())
    if start and not _valid_datetime_iso(start):
        raise HTTPException(400, f"start 格式非法: {start}(应为 YYYY-MM-DD 或 YYYY-MM-DDTHH:MM)")
    if end and not _valid_datetime_iso(end):
        raise HTTPException(400, f"end 格式非法: {end}(应为 YYYY-MM-DD 或 YYYY-MM-DDTHH:MM)")
    if start and end and start > end:
        # ISO 串(含 'T')字典序即时间序; 否则 start>end 时 _split_bounds 负区间静默空结果
        raise HTTPException(400, f"开始时间不能晚于结束时间: {start} > {end}")
    _check_span_limit(start, end)
    # 区间统一归一化: 唯一事实源(ms 边界 + day 边界 + 是否带时间), 下游一律用它取值
    _sms, _ems, sday, eday, has_time = _split_bounds(start, end)
    # 店铺子集核算不复用全量缓存(勾选不同店铺结果不同)
    no_subset_cache = bool(shop_filter)
    # 内存缓存命中(需平台一致 + 未过期 + 结果自身时间段与请求一致, 防止跨范围误命中)
    mem_result = _trace_state.get("result")
    if (not force and not no_subset_cache and mem_result
            and _trace_state["start_date"] == start and _trace_state["end_date"] == end
            and _trace_state["platform"] == plat_key
            and mem_result.get("startDate") == start and mem_result.get("endDate") == end
            and _trace_state.get("last_run")
            and time.time() - _trace_state["last_run"] < TRACE_OVERVIEW_CACHE_TTL):
        return {"status": "done", "startDate": start, "endDate": end, "result": mem_result}
    # 磁盘缓存命中(服务重启后同时间段重复核算零请求); 秒级区间不走盘(单文件会互相覆盖)
    disk = load_trace_overview_cache(start, end, plat_key) if not force and not no_subset_cache else None
    if disk:
        _trace_state["result"] = disk
        _trace_state["start_date"] = start
        _trace_state["end_date"] = end
        _trace_state["platform"] = plat_key
        _trace_state["last_run"] = time.time()
        print(f"[trace] 命中磁盘缓存 {start}~{end}, 跳过抓取 (共{disk.get('total', 0)}条)")
        return {"status": "done", "startDate": start, "endDate": end, "result": disk}
    # 导入平台(天猫1/2)分支: 数据靠 Excel 文档导入, 无 tanyu 抓取能力。
    # 永不触发 worker、不落盘、不发任何 tanyu 请求; 秒级自定义区间也强制按整天
    # (Excel 是天级粒度, 秒级边界只对 messages 有意义, 导入店无 messages)。
    if single_plat is not None and single_plat in IMPORT_PLATFORMS:
        # day 边界字符串(sday/eday)传给 DB 聚合——start/end 带 'T' 的秒级串会
        # 让 trace_daily.day 的 BETWEEN 字典序误匹配; has_time 恒 False 用 day 版。
        result = _trace_overview_from_db(sday, eday, single_plat, shop_filter, has_time=False)
        if result is None:
            result = {"startDate": start, "endDate": end, "platform": plat_key,
                      "fetched_at": time.time(), "total": 0, "adopted": 0, "rate": 0,
                      "shopList": [], "byStaff": []}
        else:
            # 内层聚合用 day 边界算的, 但对外 startDate/endDate 保持请求原值(秒级区间)
            result["startDate"] = start
            result["endDate"] = end
        if shop_filter:
            _trace_state["subset_result"] = result
        else:
            _trace_state["result"] = result
        _trace_state["start_date"] = start
        _trace_state["end_date"] = end
        _trace_state["platform"] = plat_key
        _trace_state["shop_filter"] = shop_filter
        _trace_state["last_run"] = time.time()
        print(f"[trace] 导入平台{single_plat} 纯DB聚合 {start}~{end} (共{result['total']}条)")
        return {"status": "done", "startDate": start, "endDate": end, "result": result}
    # SQLite 快路径: 数据库已覆盖该区间 => 纯本地聚合, 零上游请求(核算提速主因)
    # 店铺子集/跨平台核算同样走 DB 快路径(预抓已把三集团数据都入库, 无需切集团)
    # 覆盖判定用 day 边界(datetime 串直接比会因 'T' 静默 miss)
    # 覆盖判定限该平台店铺集: 导入平台(10/11)无限期数据会撑大全库窗口, 全库口径
    # 会让抓取平台早于其窗口的自定义区间误判已覆盖、走 _trace_overview_from_db 全零短路
    # 在线 worker(抓取平台更早区间的真实数据被藏成 0)。多平台时逐平台判定, 全过才算覆盖。
    _db_plats = plats if plats else [p for p in FETCH_PLATFORMS]
    if shop_filter:
        # 子集核算只关心勾选店铺所属平台的覆盖情况, 避免未勾选平台拖垮判定
        _db_plats = _filtered_platforms(_db_plats, shop_filter)
    if (not force and _use_sqlite_trace() and _db_plats
            and all(trace_store.db_window_covers(sday, eday, platform=p,
                                                 shop_filter=shop_filter) for p in _db_plats)):
        try:
            result = _trace_overview_from_db_multi(start, end, _db_plats, shop_filter, has_time=has_time)
            if result:
                if shop_filter:
                    # 店铺子集结果单独存, 不覆盖全量核算视图
                    _trace_state["subset_result"] = result
                else:
                    _trace_state["result"] = result
                _trace_state["start_date"] = start
                _trace_state["end_date"] = end
                _trace_state["platform"] = plat_key
                _trace_state["shop_filter"] = shop_filter
                _trace_state["last_run"] = time.time()
                print(f"[trace] SQLite 快路径 {start}~{end} (共{result['total']}条)")
                return {"status": "done", "startDate": start, "endDate": end, "result": result}
        except Exception as e:
            print(f"[trace] SQLite 快路径失败, 回退在线抓取: {e}")
    if from_cache:
        # 只读模式: 未命中缓存不触发任务; 有最近结果则带回展示
        latest = load_latest_trace_overview_cache()
        return {"status": "idle", "startDate": start, "endDate": end,
                "result": None, "latest": latest}

    _task_label = f"核算采纳率 · {plat_label}"
    if start != end:
        _task_label += f" · {start[:10]}~{end[:10]}"
    _task_params = {"days": days, "platform": platform, "force": force,
                    "start": start, "end": end, "shop_ids": shop_ids}
    # 多用户队列: 有任意任务在跑 → 入队排队(前端显示"排队中"), 不 409 拒绝
    _queued = _queue_if_busy(request, "overview", _task_label, _task_params)
    if _queued:
        return _queued
    # check+set 原子化(锁内): 否则两次并发请求都能通过检查各起一个 worker,
    # 双 worker 会互切集团 cookie 造成跨集团数据污染(与 config 写侧 TOCTOU 同类)
    with _lock:
        if _any_task_running():
            # 队列检查到加锁间的竞态窗口被交互请求抢占: 锁内再查, 忙则入队
            if request.headers.get("x-scheduler-task") == "1":
                raise HTTPException(409, "任务进行中, 调度器稍后重试")
            _task = _enqueue_task("overview", _task_label, _operator_label(request), _task_params, _operator_role(request))
            return {"status": "queued", "taskId": _task["id"],
                    "queuePosition": _queue_position(_task["id"]), "task": _task_public(_task)}
        _trace_state["running"] = True
        _trace_state["start_date"] = start
        _trace_state["end_date"] = end
        _trace_state["platform"] = platform
        _trace_state["shop_filter"] = shop_filter
        _trace_state["error"] = None
        _trace_state["paused"] = False
        _trace_state["canceled"] = False
        _trace_state["partial_list"] = []
        _trace_state["triggered_by"] = _operator_label(request)
        # 新一轮核算开始: 内存中的上一次完成结果作废(取消/中断后不能再把旧结果当本次结果,
        # 且避免跨范围误命中); 磁盘缓存仅当同范围时才删除(不同范围旧结果保留供切换后加载)。
        _trace_state["result"] = None
        _trace_state["subset_result"] = None
        _trace_state["last_run"] = None
        _trace_state["progress"] = {"done": 0, "total": 0, "current": "准备中"}
        # 直接启动(非调度器)也登记任务列表: 前端显示"谁发起的 + 进度条"
        if request.headers.get("x-scheduler-task") != "1":
            _register_running_task(_trace_state, "overview", _task_label,
                                   _operator_label(request), _task_params, _operator_role(request))
    if not no_subset_cache and load_trace_overview_cache(start, end, platform):
        try:
            if TRACE_OVERVIEW_CACHE_FILE.exists():
                TRACE_OVERVIEW_CACHE_FILE.unlink()
        except Exception:
            pass

    def worker():
        orig_gid = None
        try:
            # 跨平台/跨集团店铺集: 用 SQLite 店铺表按平台解析(不依赖当前激活集团),
            # 逐集团切换抓取(trace API 以激活集团 cookie 为作用域)。与 today-fetch 同源。
            targets = []
            if plats:
                for p in plats:
                    targets.extend(_today_target_shops(p, shop_filter))
            else:
                targets = _today_target_shops(None, shop_filter)
            # 导入平台(天猫1/2)无 tanyu 抓取能力, 由 DB 路径处理; worker 只抓抓取平台
            targets = [(gid, g, [s for s in sl if s.get("platform") in FETCH_PLATFORMS])
                       for gid, g, sl in targets]
            targets = [t for t in targets if t[2]]
            shops = [s for _gid, _g, sl in targets for s in sl]
            total = len(shops)
            today_str = datetime.date.today().isoformat()
            # 区间含今天 => 单店走"历史库 + 今天实时"合并(今天未定型不进库)
            # 判定用 day 边界(datetime 串字典序对带 'T' 的 start 会恒 False)
            range_has_today = today_str >= sday and today_str <= eday
            _trace_state["progress"] = {"done": 0, "total": total, "current": ""}
            shop_list = []
            agg_total = agg_adopted = 0
            staff_shop_agg = {}
            cur_gid = load_config().get("cookies", {}).get("tanyu-group-id")
            done = 0
            for gid, g, gshops in targets:
                if _trace_state.get("canceled"):
                    log_line("trace", f"⏹ 已取消, 已核算 {done}/{total} 家")
                    break
                if _risk_state.get("triggered"):
                    log_line("trace", f"⛔ 风控/登录失效, 集团内停止: {_risk_state.get('reason')}")
                    _trace_state["error"] = _risk_state.get("reason") or "风控/登录失效"
                    break
                if cur_gid != gid:
                    try:
                        switch_group(gid)
                        if orig_gid is None:
                            orig_gid = cur_gid
                        cur_gid = gid
                        # switch_group 内部 sync_shops_from_tanyu 会"取消运行中的核算"
                        # (防旧 worker 串组)——但这里是本 worker 自己切集团, 取消的就是
                        # 自己, 毫无意义; 复位取消/暂停标志, 否则一进店铺循环就 canceled 退出。
                        with _lock:
                            _trace_state["canceled"] = False
                            _trace_state["paused"] = False
                    except RiskTriggered as e:
                        log_line("trace", f"⛔ 风控触发于切换集团, 停止: {e}")
                        _trace_state["error"] = f"风控/登录失效: {e}"
                        break
                    except Exception as e:
                        log_line("trace", f"⚠️ 切换集团「{g.get('groupName')}」失败, 跳过该集团: {e}")
                        continue
                for i, shop in enumerate(gshops, 1):
                    idx = done + i
                    # 暂停/取消检查(在店铺边界, 请求间隙)
                    while _trace_state["paused"] and not _trace_state["canceled"]:
                        _trace_state["progress"]["current"] = "已暂停, 等待恢复…"
                        _trace_resume_evt.wait(timeout=1.0)
                        _trace_resume_evt.clear()  # 事件用完即清, 防止残留置位导致 wait() 忙等
                    if _trace_state["canceled"]:
                        log_line("trace", f"⏹ 已取消, 已核算 {idx - 1}/{total} 家")
                        break
                    _trace_state["progress"]["current"] = f"{shop['platformName']} · {shop['shopName']} ({idx}/{total})"
                    # 该店在区间内天数全部已缓存 => 纯聚合零请求, 无需限速停顿
                    if idx > 1 and not days_all_cached(shop["thirdShopId"], sday, eday):
                        sleep_trace_shop()
                    try:
                        if range_has_today:
                            # 含今天: 历史(库)+今天(实时) 合并, 不写库
                            merged = _trace_shop_merged(shop["thirdShopId"], start, end,
                                                        _sms, _ems, has_time)
                            if not merged or not merged[0]:
                                # 零消息店铺也保留在池中(total=0), 展示"无消息"
                                shop_list.append(
                                    {"shop": shop, "total": 0, "adopted": 0, "rate": 0,
                                     "replies": 0, "avgAdoptRate": 0, "avgReplyRate": 0,
                                     "startDate": start, "endDate": end}
                                )
                                _trace_state["progress"]["done"] = idx
                                continue
                            stat = merged[0]
                        else:
                            # 不跨今天: 按天遍历抓取(历史已缓存的天零请求), 秒级区间聚合前按 ms 过滤
                            stat = stat_trace_daily(shop["thirdShopId"], start, end, trim_ms=(_sms, _ems))
                        a_rate, r_rate = _avg_daily_rates(stat.get("daily"))
                        shop_list.append(
                            {"shop": shop, "total": stat["total"], "adopted": stat["adopted"],
                             "rate": stat["rate"],
                             "replies": (stat.get("byType") or {}).get("CONSULT_REPLY", 0),
                             "avgAdoptRate": a_rate, "avgReplyRate": r_rate,
                             "startDate": start, "endDate": end}
                        )
                        agg_total += stat["total"]
                        agg_adopted += stat["adopted"]
                        for st in stat.get("byStaff", []):
                            shop_agg = staff_shop_agg.setdefault(shop["thirdShopId"], {})
                            entry = shop_agg.setdefault(st["account"], {"total": 0, "adopted": 0})
                            entry["total"] += st["total"]
                            entry["adopted"] += st["adopted"]
                    except RiskTriggered as e:
                        # 风控/登录失效: 立即停止, 不再请求剩余店铺
                        log_line("trace", f"⛔ 风控触发, 停止剩余 {total - idx} 家店铺: {e}")
                        _trace_state["error"] = f"风控/登录失效, 已停止: {e}"
                        _trace_state["progress"]["current"] = f"已停止: {e}"
                        break
                    except BusyQueueError as e:
                        # 夜间抓取占用激活集团 cookie: 停止, 不写零值缓存(避免跨集团串数据)
                        log_line("trace", f"⏸ 夜间抓取进行中, 停止核算: {e}")
                        _trace_state["error"] = f"夜间抓取进行中, 请稍后重试: {e}"
                        _trace_state["progress"]["current"] = f"已停止: {e}"
                        break
                    except Exception as e:
                        log_line("trace", f"{shop['shopName']} 失败: {e}")
                    _trace_state["progress"]["done"] = idx
                    # 边算边展示: 每完成一店就发布部分结果, 前端实时渲染
                    with _lock:
                        _trace_state["partial_list"] = list(shop_list)
                done += len(gshops)
                if _trace_state.get("canceled") or _trace_state.get("error"):
                    break
            # 被取消/风控中断时: 不写磁盘缓存, 不覆盖 result(保持部分结果可查)
            # 仅风控(非取消)允许 partial_list 作为部分结果供前端展示, 但绝不落盘
            if _trace_state["canceled"] or _trace_state["error"]:
                log_line("trace", f"中断不写缓存: canceled={_trace_state['canceled']} error={_trace_state['error']}")
                return
            staff_list = _staff_list_per_shop(staff_shop_agg, single_plat, shop_filter)
            completed = {
                "startDate": start,
                "endDate": end,
                "platform": plat_key,
                "fetched_at": time.time(),
                "total": agg_total,
                "adopted": agg_adopted,
                "rate": round(agg_adopted / agg_total * 100, 2) if agg_total else 0,
                "shopList": shop_list,
                "byStaff": staff_list,
                # 区间含今天时标记实时(前端显示"实时"徽标); day 边界比较
                "live": today_str >= sday and today_str <= eday,
            }
            if shop_filter:
                # 店铺子集结果单独存, 不覆盖全量核算视图, 也不落盘
                _trace_state["subset_result"] = completed
            else:
                _trace_state["result"] = completed
                # 全量核算完成: 清掉旧子集结果, 避免 status 接口残留暴露陈旧子集数据
                _trace_state["subset_result"] = None
            _trace_state["last_run"] = time.time()
            # 店铺子集核算结果不落盘: 磁盘缓存按(时间段,平台)为键、不分店铺子集,
            # 写入子集结果会让后续同时间段的全量查询命中错误的子集数据
            if not shop_filter:
                save_trace_overview_cache(completed)
        except Exception as e:
            _trace_state["error"] = str(e)
        finally:
            # 先恢复核算前激活集团(耗时数秒), 完成后再清 running —— 恢复窗口内
            # 保持 running=True, 防止新 worker 启动互踩集团 cookie
            try:
                if orig_gid and not _risk_state.get("triggered"):
                    try:
                        if load_config().get("cookies", {}).get("tanyu-group-id") != orig_gid:
                            # invalidate=False: 恢复集团不能清掉刚算好的核算结果
                            switch_group(orig_gid, invalidate=False)
                            log_line("trace", f"已恢复原激活集团 {orig_gid}")
                    except Exception as e:
                        log_line("trace", f"⚠️ 恢复原集团失败: {e}")
            finally:
                _trace_state["running"] = False
                # 运行已结束: 清掉暂停标记, 避免下次查询/再核算时遗留 paused=True
                _trace_state["paused"] = False
                _trace_resume_evt.clear()  # 清事件, 避免后续暂停 wait() 立即返回造成忙等

    threading.Thread(target=worker, daemon=True).start()
    return {"status": "running", "startDate": start, "endDate": end,
            "progress": _trace_state["progress"]}


def _iter_days(sday, eday):
    """按天生成 [sday, eday] 的日期列表(含两端)"""
    sd = datetime.date.fromisoformat(sday)
    ed = datetime.date.fromisoformat(eday)
    return [(sd + datetime.timedelta(days=i)).isoformat() for i in range((ed - sd).days + 1)]


def _staff_window_gaps(platform, shop_filter, sday, eday):
    """客服池窗口内疑似缺抓的天: {day: 该天缺抓的估计店铺数}

    用 DB 按天店铺数检测夜间抓取失败造成的整平台缺天: 某天"有消息的店铺数"
    骤降到窗口内最高店铺数的 50% 以下 → 当天疑似未抓全(缺抓夜是整平台同时
    缺, 典型如 08-13 凌晨 DNS 失败→08-12 抖音 21 店只剩 1 店有数据)。
    基准取"窗口内最高店铺数"而非当前在编数: 新店加入只补当前天, 不拉低历史
    天基线, 不会把"新店未补历史"误判成缺抓。比逐店读 trace_days 缓存文件快
    一个量级(单条 SQL), 且语义一致——缺抓天在 trace_daily 无行, 自然被检出。
    """
    if platform not in FETCH_PLATFORMS:
        return {}
    try:
        per_day = trace_store.count_shops_per_day(sday, eday, platform, shop_filter)
    except Exception:
        return {}
    if not per_day:
        return {}
    max_count = max(per_day.values())
    if max_count < 5:
        return {}  # 店少平台(如导入平台)波动大, 不判缺抓
    gaps = {}
    for d in _iter_days(sday, eday):
        n = per_day.get(d, 0)
        if n <= max_count * 0.5:
            gaps[d] = max_count - n
    return gaps


@app.get("/api/trace/staff")
def trace_staff(days: int = 7, start: str | None = None, end: str | None = None,
                platform: int | None = None, shop_ids: str | None = None):
    """客服账号池独立时间筛选: 返回该时间段内各客服的总消息/采纳条数/采纳率

    从 SQLite trace_daily 聚合(跨集团全量), 与核算口径一致(adopted=send_type IN 1,2,3)。
    任意平台/店铺子集/时间段均可查询, 不依赖当前激活集团。
    客服**按 (店铺,客服) 组合区分**(不去重/不跨店合并): 同一客服账号在不同店铺
    各自成行, 并补入该店在编但窗口内无消息的客服(来自 staff_names)。
    DB 未覆盖该区间时返回空 byStaff(前端提示数据未抓取); 部分覆盖(尾部缺天/
    区间内缺天)仍聚合已有天并标 coverage=partial + missingDays, 前端提示不全。
    """
    start, end = date_range(days, end) if not start else (start, end or (datetime.date.today() - datetime.timedelta(days=1)).isoformat())
    if start and not _valid_datetime_iso(start):
        raise HTTPException(400, f"start 格式非法: {start}")
    if end and not _valid_datetime_iso(end):
        raise HTTPException(400, f"end 格式非法: {end}")
    if start and end and start > end:
        raise HTTPException(400, f"开始时间不能晚于结束时间: {start} > {end}")
    _check_span_limit(start, end)
    _sms, _ems, sday, eday, has_time = _split_bounds(start, end)
    shop_filter = {s for s in shop_ids.split(",") if s.strip()} if shop_ids else None
    result = {"startDate": start, "endDate": end, "platform": platform,
              "total": 0, "adopted": 0, "rate": 0, "byStaff": [],
              "coverage": "full"}
    # 覆盖判定限该平台店铺集(理由同 overview SQLite 快路径注释:
    # 导入平台无限期数据会撑大全库窗口, 全库口径误判抓取平台早于窗口区间已覆盖)。
    # 导入平台(10/11)无预抓, 数据新鲜度取决于手动上传, 用宽松判定: 区间内
    # 至少有一天数据即可聚合(空天自然为 0), 否则默认"近7天"窗口末端没上传的
    # 昨天会把整个区间判为未覆盖、客服池一片空白。
    if platform in IMPORT_PLATFORMS:
        covered = trace_store.db_window_overlaps(sday, eday, platform=platform,
                                                 shop_filter=shop_filter)
    else:
        covered = trace_store.db_window_covers(sday, eday, platform=platform,
                                               shop_filter=shop_filter)
        # 抓取平台完整覆盖失败 → 部分覆盖兜底: 区间内部分天有数据就聚合已有天,
        # 标 coverage=partial 供前端提示"数据可能不全"(单天缺失整片空白更误导)。
        if not covered:
            rng = trace_store.db_window_range(sday, eday, platform=platform,
                                              shop_filter=shop_filter)
            if rng:
                covered = True
                result["coverage"] = "partial"
                result["coveredFrom"], result["coveredTo"] = rng
        # 即使 DB 覆盖判定通过(MIN/MAX 有最远天), 某平台/某店可能没抓到: 用按天
        # 店铺数骤降检测(夜间抓取失败是整平台同时缺, 某天店铺数会断崖), 精确暴露
        # "没抓到"而非"真没消息", 让前端警告数据不全。
        if covered:
            gaps = _staff_window_gaps(platform, shop_filter, sday, eday)
            if gaps:
                result["coverage"] = "partial"
                result["missingDays"] = gaps
                result["missingCount"] = sum(gaps.values())
    if covered:
        if has_time and platform not in IMPORT_PLATFORMS:
            per_shop = trace_store.staff_aggregate_per_shop_ms(_sms, _ems, platform, shop_filter)
        else:
            # 导入平台(10/11)或纯日期区间: 走天级 trace_daily(导入数据无 messages,
            # 秒级边界对导入店按整天聚合; Excel 就是天级粒度)。
            # 必须用 day 边界字符串 sday/eday: start/end 带 'T' 的秒级串会让
            # trace_daily.day 的 BETWEEN 字典序误匹配, 丢掉区间首日。
            per_shop = trace_store.staff_aggregate_per_shop(sday, eday, platform, shop_filter)
        staff_list = _staff_list_per_shop(per_shop["by_shop"], platform, shop_filter)
        result["byStaff"] = staff_list
        result["total"] = sum(s["total"] for s in staff_list)
        result["adopted"] = sum(s["adopted"] for s in staff_list)
        result["rate"] = round(result["adopted"] / result["total"] * 100, 2) if result["total"] else 0
    return result


@app.get("/api/trace/compare")
def trace_compare(mode: str = "week", platform: int | None = None):
    """周期对比: 本周 vs 上周 / 本月 vs 上月, 逐店消息量/采纳率变化

    mode=week: 本周一~昨天 vs 上周一~上周日; mode=month: 本月1日~昨天 vs 上月全月。
    返回每店 {name, platform, cur/prev 的 total/adopted/rate, deltaTotal%(消息量), deltaRatePP(采纳率百分点)}。
    数据源 trace_daily(本地聚合, 零上游请求)。
    """
    today = datetime.date.today()
    yesterday = today - datetime.timedelta(days=1)
    if mode == "month":
        cur_start = today.replace(day=1)
        prev_end = cur_start - datetime.timedelta(days=1)
        prev_start = prev_end.replace(day=1)
    else:  # week: 本周一 ~ 昨天(周一当天昨天在上周, 退化为近7天)
        cur_start = today - datetime.timedelta(days=today.weekday())
        if cur_start > yesterday:
            cur_start = yesterday - datetime.timedelta(days=6)
        prev_end = cur_start - datetime.timedelta(days=1)
        prev_start = prev_end - datetime.timedelta(days=6)
    cur_shops = trace_store.shop_totals(cur_start.isoformat(), yesterday.isoformat(), platform=platform)
    prev_shops = trace_store.shop_totals(prev_start.isoformat(), prev_end.isoformat(), platform=platform)
    rows = []
    for sid, cs in cur_shops.items():
        ps = prev_shops.get(sid)
        if not ps:
            continue
        delta_total = ((cs["total"] - ps["total"]) / ps["total"] * 100) if ps["total"] else 0
        cur_rate = (cs["adopted"] / cs["total"] * 100) if cs["total"] else 0
        prev_rate = (ps["adopted"] / ps["total"] * 100) if ps["total"] else 0
        rows.append({
            "id": sid, "name": cs["name"], "platform": cs["platform"],
            "curTotal": cs["total"], "curAdopted": cs["adopted"], "curRate": round(cur_rate, 2),
            "prevTotal": ps["total"], "prevAdopted": ps["adopted"], "prevRate": round(prev_rate, 2),
            "deltaTotal": round(delta_total, 1),
            "deltaRatePP": round(cur_rate - prev_rate, 2),
        })
    rows.sort(key=lambda x: x["curTotal"], reverse=True)
    return {"mode": mode,
            "curStart": cur_start.isoformat(), "curEnd": yesterday.isoformat(),
            "prevStart": prev_start.isoformat(), "prevEnd": prev_end.isoformat(),
            "rows": rows,
            "curTotal": sum(r["curTotal"] for r in rows),
            "prevTotal": sum(r["prevTotal"] for r in rows)}


# ---------- 数据洞察(哨兵日历 / 客服生物钟 / AI 洞察日报) ----------
@app.get("/api/trace/coverage-grid")
def trace_coverage_grid(days: int = 35, platform: int | None = None):
    """数据完整性哨兵: 完整日窗口 × 店铺 覆盖矩阵(数据洞察页)

    绿=该天已抓取(含 0 消息, upsert_shop_day 已写行); 红=缺抓; 今天未定型不展示。
    """
    days = max(1, min(int(days), 35))
    end = datetime.date.today() - datetime.timedelta(days=1)
    start = end - datetime.timedelta(days=days - 1)
    # 店铺清单: 跨集团全部(trace_store.get_shops), 可平台过滤
    try:
        shop_rows = trace_store.get_shops(platform=platform)
    except Exception:
        shop_rows = []
    shops = [{"id": r["thirdShopId"], "name": r["shopName"], "platform": r["platform"]}
             for r in shop_rows]
    cells = trace_store.coverage_grid(start.isoformat(), end.isoformat(), platform=platform)
    # 该平台店铺数(用于缺格统计)
    day_list = [(start + datetime.timedelta(days=i)).isoformat() for i in range(days)]
    missing = 0
    total_cells = len(shops) * days
    for sid in shops:
        s_cells = cells.get(sid["id"], {})
        for d in day_list:
            if d not in s_cells:
                missing += 1
    return {"start": start.isoformat(), "end": end.isoformat(),
            "days": day_list, "shops": shops, "cells": cells,
            "missingCells": missing, "totalCells": total_cells}


@app.get("/api/trace/hourly-heatmap")
def trace_hourly_heatmap(days: int = 7, platform: int | None = None):
    """客服生物钟: 近 N 个完整日 × 24 小时的消息量热力分布"""
    days = max(1, min(int(days), 30))
    end = datetime.date.today() - datetime.timedelta(days=1)
    start = end - datetime.timedelta(days=days - 1)
    start_ms = int(datetime.datetime(start.year, start.month, start.day).timestamp() * 1000)
    end_ms = int((datetime.datetime(end.year, end.month, end.day) + datetime.timedelta(days=1)).timestamp() * 1000) - 1
    rows = trace_store.hourly_counts(start_ms, end_ms, platform=platform)
    counts = {}  # day -> {hour: n}
    for d, h, n in rows:
        counts.setdefault(d, {})[h] = n
    day_list = [d for d in counts.keys()]
    day_list.sort()
    # 补全空天(缺抓的天也显示, 便于一眼看出哪天没数据)
    full = [(start + datetime.timedelta(days=i)).isoformat() for i in range(days)]
    for d in full:
        counts.setdefault(d, {})
    return {"start": start.isoformat(), "end": end.isoformat(),
            "days": full, "hours": list(range(24)), "counts": counts}


@app.get("/api/insights")
def insights(days: int = 1, platform: int | None = None):
    """AI 洞察日报(规则引擎, 非大模型): 基于 trace_daily 本地聚合自动生成
    自然语言洞察: 总量环比 / 平台分解 / 采纳率 / 客服 TOP / 异常店铺 / 缺抓警告。
    days=1 只对比昨天 vs 前天; days=3/7 对比区间 vs 前同等长度区间。
    """
    days = max(1, min(int(days), 30))
    today = datetime.date.today()
    # 区间 A: 最近 N 个完整日(不含今天); 区间 B: 更早的同等长度区间
    a_end = today - datetime.timedelta(days=1)
    a_start = a_end - datetime.timedelta(days=days - 1)
    b_end = a_start - datetime.timedelta(days=1)
    b_start = b_end - datetime.timedelta(days=days - 1)
    items = []
    # 1) 区间内每店聚合(对比用)
    a_shops = trace_store.shop_totals(a_start.isoformat(), a_end.isoformat(), platform=platform)
    b_shops = trace_store.shop_totals(b_start.isoformat(), b_end.isoformat(), platform=platform)
    a_total = sum(s["total"] for s in a_shops.values())
    a_adopted = sum(s["adopted"] for s in a_shops.values())
    b_total = sum(s["total"] for s in b_shops.values())
    b_adopted = sum(s["adopted"] for s in b_shops.values())
    a_rate = (a_adopted / a_total * 100) if a_total else 0
    b_rate = (b_adopted / b_total * 100) if b_total else 0
    label_a = "昨日" if days == 1 else f"近{days}天"
    label_b = "前日" if days == 1 else f"前{days}天"
    # 2) 总量洞察
    def _fmt(n):
        """千分位格式化(前端 fmt 的 Python 版)"""
        try:
            return f"{int(n):,}"
        except Exception:
            return str(n)

    import html as _html_mod
    _html = _html_mod.escape

    if a_total:
        diff = (a_total - b_total) / b_total * 100 if b_total else 0
        arrow = "▲" if diff >= 0 else "▼"
        items.append({
            "level": "success" if diff > 0 else ("info" if diff == 0 else "warning"),
            "text": f"{arrow} {label_a}消息量 <b>{_fmt(a_total)}</b> 条, 环比{label_b} "
                    f"{'+' if diff >= 0 else ''}{diff:.1f}%({_fmt(b_total)} 条)"})
    # 3) 采纳率洞察
    if a_total:
        rdiff = a_rate - b_rate
        items.append({
            "level": "success" if rdiff >= 0 else "warning",
            "text": f"采纳率 <b>{a_rate:.2f}%</b>({_fmt(a_adopted)}/{_fmt(a_total)}), "
                    f"环比{'+' if rdiff >= 0 else ''}{rdiff:.2f} 个百分点"})
    # 4) 平台分解(仅当未指定平台)
    if platform is None:
        pf_totals = {}
        for s in a_shops.values():
            pf = s["platform"]
            e = pf_totals.setdefault(pf, {"total": 0, "adopted": 0, "shops": 0})
            e["total"] += s["total"]
            e["adopted"] += s["adopted"]
            e["shops"] += 1
        if len(pf_totals) > 1:
            top_pf = max(pf_totals.items(), key=lambda x: x[1]["total"])
            share = (top_pf[1]["total"] / a_total * 100) if a_total else 0
            pname = PLATFORM_NAMES.get(top_pf[0], f"平台{top_pf[0]}")
            items.append({
                "level": "info",
                "text": f"平台占比: <b>{pname}</b> 居首({share:.1f}%), "
                        f"{top_pf[1]['total']} 条 / {top_pf[1]['shops']} 家店"})
    # 5) 客服 TOP(采纳数 / 采纳率, 来自 trace_daily.by_staff_json)
    try:
        shop_names = {}
        for s in load_all_shops():
            shop_names[s["thirdShopId"]] = s.get("shopName", "")
        staff_totals = {}
        cells_a = trace_store.coverage_grid(a_start.isoformat(), a_end.isoformat(), platform=platform)
        for sid, days_map in cells_a.items():
            for d in days_map:
                # 从 trace_daily 读 by_staff_json
                staff_rows = trace_store.shop_staff_json(sid, d)
                for acct, v in (staff_rows or []):
                    e = staff_totals.setdefault(f"{acct}@{shop_names.get(sid, '')}",
                                                {"total": 0, "adopted": 0, "acct": acct})
                    e["total"] += v.get("total", 0)
                    e["adopted"] += v.get("adopted", 0)
        if staff_totals:
            top = max(staff_totals.items(), key=lambda x: x[1]["total"])
            top_rate = max(
                (x for x in staff_totals.items() if x[1]["total"] >= 20),
                key=lambda x: (x[1]["adopted"] / x[1]["total"] if x[1]["total"] else 0),
                default=None)
            msg = f"客服处理量第一: <b>{_html(top[1]['acct'])}</b> 处理 {_fmt(top[1]['total'])} 条"
            if top_rate and top_rate[0] != top[0]:
                msg += (f" · 采纳率第一: <b>{_html(top_rate[1]['acct'])}</b> "
                        f"({top_rate[1]['adopted']/top_rate[1]['total']*100:.1f}%)")
            items.append({"level": "success", "text": msg})
    except Exception:
        pass
    # 6) 异常店铺(环比变化 > ±50% 且总量 ≥ 100)
    if a_shops and b_shops:
        swings = []
        for sid, sa in a_shops.items():
            sb = b_shops.get(sid)
            if not sb or sb["total"] < 50:
                continue
            if sa["total"] < 50:
                continue
            d = (sa["total"] - sb["total"]) / sb["total"] * 100
            if abs(d) >= 50:
                swings.append((d, sa["name"]))
        swings.sort(reverse=True)
        for d, name in swings[:3]:
            items.append({
                "level": "danger" if d < 0 else "info",
                "text": f"店铺「{_html(name)}」消息量 "
                        f"{'+' if d >= 0 else ''}{d:.0f}%"})
    # 7) 缺抓警告(覆盖矩阵缺格)
    try:
        cells_b = trace_store.coverage_grid(a_start.isoformat(), a_end.isoformat(), platform=platform)
        missing = 0
        for sid in a_shops:
            s_cells = cells_b.get(sid, {})
            for i in range(days):
                d = (a_start + datetime.timedelta(days=i)).isoformat()
                if d not in s_cells:
                    missing += 1
        if missing:
            items.append({
                "level": "danger",
                "text": f"数据缺口: 区间内有 <b>{missing}</b> 个 (店×天) 未抓到, "
                        f"可在「系统设置 → 手动补抓数据」补齐"})
        else:
            items.append({
                "level": "success",
                "text": f"数据完整性: 区间内 {len(a_shops)} 家店全部抓取完整, 无缺口"})
    except Exception:
        pass
    return {"items": items, "start": a_start.isoformat(), "end": a_end.isoformat(),
            "label": label_a, "compare": label_b, "total": a_total, "adopted": a_adopted,
            "rate": round(a_rate, 2)}


def _shop_name(shop_id):
    """店铺名(洞察客服行展示用)"""
    try:
        shops = {s["thirdShopId"]: s["shopName"] for s in load_all_shops()}
        return shops.get(shop_id, "")
    except Exception:
        return ""


# ---------- 今日实时抓取(核算/人工客服「抓取今日数据」) ----------
def _parse_hhmm(ts, default="00:00"):
    """解析 HH:MM 或 HH:MM:SS(秒级今日抓取用), 非法或越界返回 default"""
    try:
        parts = str(ts).strip().split(":")
        h, m = int(parts[0]), int(parts[1])
        s = int(parts[2]) if len(parts) > 2 else 0
        if not (0 <= h <= 23 and 0 <= m <= 59 and 0 <= s <= 59):
            return default
        if s:
            return f"{h:02d}:{m:02d}:{s:02d}"
        return f"{h:02d}:{m:02d}"
    except Exception:
        return default


def _ts_seconds(ts):
    """HH:MM 或 HH:MM:SS → 当日秒数(用于秒级起止比较), 非法返回 None"""
    try:
        p = str(ts).split(":")
        h, m = int(p[0]), int(p[1])
        s = int(p[2]) if len(p) > 2 else 0
        if not (0 <= h <= 23 and 0 <= m <= 59 and 0 <= s <= 59):
            return None
        return h * 3600 + m * 60 + s
    except Exception:
        return None


def _full_today_ts(today, ts, end=False):
    """'YYYY-MM-DD' + 'HH:MM[:SS]' → 完整时间串

    end=True 且 ts 只有分钟时补 :59(含整分钟, 与旧行为一致);
    带秒时原样用(秒级精确到该秒, tanyu 终点含该秒)。
    """
    p = str(ts).split(":")
    h, m = int(p[0]), int(p[1])
    s = int(p[2]) if len(p) > 2 else (59 if end else 0)
    return f"{today} {h:02d}:{m:02d}:{s:02d}"


def _today_target_shops(platform=None, shop_filter=None):
    """按平台/店铺子集解析目标店铺集合(跨全部 config.groups)

    返回 [(groupId, group, [shops...]), ...]: 每个集团一组店铺, 便于 worker 逐集团切换抓取。
    platform 为 None 时含全部抓取平台集团; shop_filter 只保留集合内的店铺。
    """
    cfg = load_config()
    platform_order = cfg.get("prefetch_platforms") or [1, 5, 7]
    plat_rank = {p: i for i, p in enumerate(platform_order)}
    ordered = sorted(
        [g for g in cfg.get("groups", []) if g.get("groupId") and g.get("accountId")],
        key=lambda g: plat_rank.get(g.get("platform"), 99),
    )
    out = []
    for g in ordered:
        if platform is not None and g.get("platform") != platform:
            continue
        # 每家集团切过去后 load_shops 才读对(switch_group 内部会 sync_shops_from_tanyu),
        # 这里用 SQLite shops 表(全部集团店铺已入库)解析, 不依赖当前激活集团
        import trace_store as _ts
        try:
            shops = _ts.get_shops()
        except Exception:
            shops = []
        if not shops:
            continue
        # SQLite shops 表含全部集团店铺; 每家集团只取本平台店铺, 否则 scoped
        # shop_filter 会被错误地塞进每个集团重复抓取(单店被抓 3 次的 bug)
        shops = [s for s in shops if s.get("platform") == g.get("platform")]
        if shop_filter is not None:
            shops = [s for s in shops if s["thirdShopId"] in shop_filter]
        if shops:
            out.append((g["groupId"], g, shops))
    return out


@app.post("/api/trace/today")
def trace_today(request: Request, platform: int | None = None, shop_ids: str | None = None,
                start_ts: str = "00:00", end_ts: str | None = None, mode: str = ""):
    """抓取今日数据(实时, 不写库): 按集团切换轮询, 支持平台/店铺子集筛选

    mode: 'staff'=抓当日客服实时采纳率(进客服池) / 'audit'=抓当日核算数据 / 空=通用。
    仅用于任务列表标签展示(区分"抓取当日客服实时采纳率"与"抓取当日核算数据")。

    - start_ts/end_ts: 今天内的起止时刻(HH:MM), 默认全天 00:00 ~ 当前时刻
    - shop_ids: 逗号分隔店铺子集, 空=全部店铺(含平台筛选)
    - 逐集团 switch_group → 抓今天 → 恢复原集团; 风控立即停止
    """
    today_str = datetime.date.today().isoformat()
    start_ts = _parse_hhmm(start_ts, "00:00")
    end_ts = _parse_hhmm(end_ts, datetime.datetime.now().strftime("%H:%M")) if end_ts else datetime.datetime.now().strftime("%H:%M")
    # 秒级比较: 字符串比较对 "14:35" vs "14:35:30" 会误判(前缀短者小), 用当日秒数
    s_sec, e_sec = _ts_seconds(start_ts), _ts_seconds(end_ts)
    if s_sec is None or e_sec is None or e_sec < s_sec:
        raise HTTPException(400, "结束时刻不能早于开始时刻")
    shop_filter = {s for s in shop_ids.split(",") if s.strip()} if shop_ids else None
    groups_target = _today_target_shops(platform, shop_filter)
    if not groups_target:
        return {"status": "done", "startDate": today_str, "endDate": today_str,
                "startTs": start_ts, "endTs": end_ts,
                "result": {"startDate": today_str, "endDate": today_str,
                           "startTs": start_ts, "endTs": end_ts, "platform": platform,
                           "total": 0, "adopted": 0, "rate": 0, "byStaff": [],
                           "shopList": [], "live": True}}
    _today_action = "抓取当日客服实时采纳率" if mode == "staff" else ("抓取当日核算数据" if mode == "audit" else "抓取今日实时数据")
    _task_label = f"{_today_action} · {PLATFORM_NAMES.get(platform, '全平台')} · {start_ts}~{end_ts}"
    _task_params = {"platform": platform, "shop_ids": shop_ids, "start_ts": start_ts, "end_ts": end_ts, "mode": mode}
    # 多用户队列: 有任意任务在跑 → 入队排队(前端显示"排队中"), 不 409 拒绝
    _queued = _queue_if_busy(request, "today", _task_label, _task_params)
    if _queued:
        return _queued
    # check+set 原子化(锁内先重查再置位): 并发双开今日抓取会互切集团 cookie 串数据
    with _lock:
        if _any_task_running():
            if request.headers.get("x-scheduler-task") == "1":
                raise HTTPException(409, "任务进行中, 调度器稍后重试")
            _task = _enqueue_task("today", _task_label, _operator_label(request), _task_params, _operator_role(request))
            return {"status": "queued", "taskId": _task["id"],
                    "queuePosition": _queue_position(_task["id"]), "task": _task_public(_task)}
        _today_state["running"] = True
        _today_state["start_date"] = today_str
        _today_state["end_date"] = today_str
        _today_state["start_ts"] = start_ts
        _today_state["end_ts"] = end_ts
        _today_state["platform"] = platform
        _today_state["shop_filter"] = shop_filter
        _today_state["result"] = None
        _today_state["last_run"] = None
        _today_state["error"] = None
        _today_state["progress"] = {"done": 0, "total": 0, "current": "准备中"}
        _today_state["triggered_by"] = _operator_label(request)
        # 直接启动(非调度器)也登记任务列表: 前端显示"谁发起的 + 进度条"
        if request.headers.get("x-scheduler-task") != "1":
            _register_running_task(_today_state, "today", _task_label,
                                   _operator_label(request), _task_params, _operator_role(request))
    total_plan = sum(len(shops) for _, _, shops in groups_target)
    _today_state["progress"]["total"] = total_plan
    orig_gid = load_config().get("cookies", {}).get("tanyu-group-id")

    def worker():
        try:
            shop_list = []
            staff_shop_agg = {}
            agg_total = agg_adopted = 0
            by_plat = {}   # {platform: {"total","adopted","shops"}} 按平台拆分, 供前端"今日各平台"展示
            done = 0
            for gid, g, shops in groups_target:
                if _risk_state.get("triggered"):
                    log_line("today", "⛔ 风控/登录失效, 整体停止")
                    _today_state["error"] = _risk_state.get("reason") or "风控/登录失效"
                    return
                try:
                    switch_group(gid)
                except RiskTriggered:
                    log_line("today", "⛔ 风控触发于切换集团, 整体停止")
                    _today_state["error"] = _risk_state.get("reason") or "风控/登录失效"
                    return
                except Exception as e:
                    log_line("today", f"⚠️ 集团「{g.get('groupName')}」切换失败, 跳过: {e}")
                    continue
                for i, shop in enumerate(shops, 1):
                    if _risk_state.get("triggered"):
                        log_line("today", "⛔ 风控/登录失效, 停止")
                        _today_state["error"] = _risk_state.get("reason") or "风控/登录失效"
                        return
                    _today_state["progress"]["current"] = (
                        f"集团[{g.get('groupName')}] {shop.get('platformName', '')} · "
                        f"{shop.get('shopName', '')} ({done + i}/{total_plan})")
                    if done > 0 or i > 1:
                        sleep_trace_shop()
                    try:
                        res = _fetch_trace_range(
                            shop["thirdShopId"],
                            _full_today_ts(today_str, start_ts),
                            _full_today_ts(today_str, end_ts, end=True),
                        )
                        msgs = [_trim_trace_msg(m) for m in (res or [])]
                        stat = _aggregate_stat_rows(msgs, today_str, today_str)
                        shop_list.append(
                            {"shop": shop, "total": stat["total"], "adopted": stat["adopted"],
                             "rate": stat["rate"], "startDate": today_str, "endDate": today_str}
                        )
                        agg_total += stat["total"]
                        agg_adopted += stat["adopted"]
                        p = shop.get("platform", 0)
                        agg_p = by_plat.setdefault(p, {"total": 0, "adopted": 0, "shops": 0})
                        agg_p["total"] += stat["total"]
                        agg_p["adopted"] += stat["adopted"]
                        agg_p["shops"] += 1
                        for st in stat.get("byStaff", []):
                            shop_agg = staff_shop_agg.setdefault(shop["thirdShopId"], {})
                            entry = shop_agg.setdefault(st["account"], {"total": 0, "adopted": 0})
                            entry["total"] += st["total"]
                            entry["adopted"] += st["adopted"]
                        if stat["total"]:
                            log_line("today", f"{shop.get('shopName', '')} 今日 {stat['total']} 条 / 采纳 {stat['adopted']}")
                    except RiskTriggered as e:
                        log_line("today", f"⛔ 风控触发, 停止: {e}")
                        _today_state["error"] = f"风控/登录失效, 已停止: {e}"
                        return
                    except BusyQueueError as e:
                        # 夜间抓取占用激活集团 cookie: 停止, 不写零值缓存
                        log_line("today", f"⏸ 夜间抓取进行中, 停止今日抓取: {e}")
                        _today_state["error"] = f"夜间抓取进行中, 请稍后重试: {e}"
                        return
                    except Exception as e:
                        log_line("today", f"{shop.get('shopName', '')} 今日抓取失败: {e}")
                    done += 1
                    _today_state["progress"]["done"] = done
                    with _lock:
                        _today_state["result"] = {
                            "startDate": today_str, "endDate": today_str,
                            "startTs": start_ts, "endTs": end_ts, "platform": platform,
                            "total": agg_total, "adopted": agg_adopted,
                            "rate": round(agg_adopted / agg_total * 100, 2) if agg_total else 0,
                            "byPlatform": {
                                str(p): {"total": v["total"], "adopted": v["adopted"],
                                         "rate": round(v["adopted"] / v["total"] * 100, 2) if v["total"] else 0,
                                         "shops": v["shops"]}
                                for p, v in by_plat.items()
                            },
                            "shopList": list(shop_list),
                            "byStaff": _staff_list_per_shop(staff_shop_agg, platform, shop_filter),
                            "live": True,
                        }
            _today_state["last_run"] = time.time()
            log_line("today", f"今日抓取完成: {agg_total} 条 / 采纳 {agg_adopted} / {total_plan} 家店铺")
        except Exception as e:
            _today_state["error"] = str(e)
        finally:
            # 先恢复原激活集团(耗时数秒), 完成后再清 running(防止恢复窗口内双启动)
            try:
                if orig_gid and not _risk_state.get("triggered"):
                    try:
                        if orig_gid != load_config().get("cookies", {}).get("tanyu-group-id"):
                            # invalidate=False: 恢复集团不清核算缓存(今日抓取不干扰已算好的结果)
                            switch_group(orig_gid, invalidate=False)
                            log_line("today", f"已恢复原激活集团 {orig_gid}")
                    except Exception as e:
                        log_line("today", f"⚠️ 恢复原集团失败: {e}")
            finally:
                _today_state["running"] = False

    threading.Thread(target=worker, daemon=True).start()
    return {"status": "running", "startDate": today_str, "endDate": today_str,
            "startTs": start_ts, "endTs": end_ts, "progress": _today_state["progress"]}


@app.get("/api/trace/today/status")
def trace_today_status(request: Request):
    """今日实时抓取任务状态/结果(附带夜间抓取标志: 前端展示"谁在抓取/为何阻塞")"""
    with _lock:
        st = dict(_today_state)
    st["nightly"] = _nightly_flag_info()
    if st.get("triggered_by"):
        st["triggered_by"] = _mask_admin_label(request, st["triggered_by"])
    return st


@app.get("/api/trace/overview/status")
def trace_overview_status(request: Request):
    """核算总览任务状态/结果(附带夜间抓取标志: 前端展示"谁在抓取/为何阻塞")"""
    with _lock:
        st = {
            "running": _trace_state["running"],
            "progress": _trace_state["progress"],
            "startDate": _trace_state["start_date"],
            "endDate": _trace_state["end_date"],
            "platform": _trace_state["platform"],
            "lastRun": _trace_state["last_run"],
            "taskShopFilter": _trace_state["shop_filter"],
            "error": _trace_state["error"],
            "result": _trace_state["result"],
            "subsetResult": _trace_state["subset_result"],
            "paused": _trace_state["paused"],
            "canceled": _trace_state["canceled"],
            "shopFilter": _trace_state["shop_filter"],
            "shopList": _trace_state["partial_list"],  # 运行中已完成店铺(边算边展示)
            "triggeredBy": _mask_admin_label(request, _trace_state.get("triggered_by")),
        }
    st["nightly"] = _nightly_flag_info()
    return st


@app.post("/api/trace/overview/pause")
def trace_overview_pause():
    """暂停核算: 置暂停标志, worker 在当前店铺完成后阻塞, 不再请求剩余店铺"""
    with _lock:
        if not _trace_state["running"]:
            return {"ok": False, "reason": "无进行中的核算任务"}
        if _trace_state["canceled"]:
            # 已取消的运行不响应暂停(worker 即将退出); 避免遗留 paused=True 误导前端
            return {"ok": False, "reason": "核算已取消"}
        _trace_state["paused"] = True
        _trace_state["progress"]["current"] = "已暂停, 等待恢复…"
    print("[trace] paused")
    return {"ok": True, "paused": True}


@app.post("/api/trace/overview/resume")
def trace_overview_resume():
    """恢复核算: 清除暂停标志并唤醒 worker"""
    with _lock:
        if not _trace_state["running"]:
            return {"ok": False, "reason": "无进行中的核算任务"}
        _trace_state["paused"] = False
    _trace_resume_evt.set()  # 唤醒阻塞中的 worker
    print("[trace] resumed")
    return {"ok": True, "paused": False}


@app.post("/api/trace/overview/cancel")
def trace_overview_cancel():
    """取消核算: 停止后续店铺, 保留已完成店铺的部分结果(不写磁盘缓存)"""
    with _lock:
        if not _trace_state["running"]:
            return {"ok": False, "reason": "无进行中的核算任务"}
        _trace_state["canceled"] = True
        _trace_state["paused"] = False
    _trace_resume_evt.set()  # 若 worker 正阻塞在暂停等待中, 唤醒使其退出
    print("[trace] canceled")
    return {"ok": True, "canceled": True}


@app.get("/api/trace/messages/{shop_id}")
def trace_messages(shop_id: str, days: int = 7, force: int = 0,
                   start: str | None = None, end: str | None = None):
    """单店原始消息记录(用于与探域后台核对)

    返回逐条消息: 时间/买家/发送状态/人工客服/内容
    按日增量缓存: 已抓取的天零请求, 只补抓缺失的天
    """
    shops = {s["thirdShopId"]: s for s in load_all_shops()}
    shop = shops.get(shop_id)
    if not shop:
        raise HTTPException(404, "店铺不存在")
    start, end = date_range(days, end) if not start else (start, end or (datetime.date.today() - datetime.timedelta(days=1)).isoformat())
    if start and not _valid_datetime_iso(start):
        raise HTTPException(400, f"start 格式非法: {start}")
    if end and not _valid_datetime_iso(end):
        raise HTTPException(400, f"end 格式非法: {end}")
    if start and end and start > end:
        raise HTTPException(400, f"开始时间不能晚于结束时间: {start} > {end}")
    _check_span_limit(start, end)
    _sms, _ems, sday, eday, has_time = _split_bounds(start, end)
    today_str = datetime.date.today().isoformat()
    # 导入平台(天猫1/2)店: 无原始消息, 返回天级 daily + 空 messages, 绝不请求 tanyu
    if _is_import_shop(shop):
        stat = _import_shop_stat(shop_id, start, end)
        return {"shop": shop, "startDate": start, "endDate": end,
                "total": (stat or {}).get("total", 0),
                "daily": (stat or {}).get("daily", []),
                "messages": [], "live": False, "imported": True}
    # 区间含今天: 历史(库) + 今天(实时) 合并(day 边界判定)
    if (not force and _use_sqlite_trace()
            and today_str >= sday and today_str <= eday):
        try:
            merged = _trace_shop_merged(shop_id, start, end, _sms, _ems, has_time)
            if merged and merged[0]:
                stat, has_today = merged
                return {"shop": shop, "startDate": start, "endDate": end,
                        "total": stat["total"], "daily": stat.get("daily", []),
                        "messages": stat.get("messages", []), "live": has_today}
        except RiskTriggered:
            raise HTTPException(502, "风控/登录失效, 请重新登录后重试")
        except Exception as e:
            print(f"[trace] 原始消息 合并路径失败, 回退在线抓取: {e}")
    cache = _load_trace_days_cache(shop_id) if not force else None
    if cache and cache.get("days") and not force:
        start_d = datetime.date.fromisoformat(sday)
        end_d = datetime.date.fromisoformat(eday)
        need = {(start_d + datetime.timedelta(days=i)).isoformat()
                for i in range((end_d - start_d).days + 1)}
        if all(_cached_days_usable(cache, ds) for ds in need):
            stat = stat_trace_daily(shop_id, start, end, trim_ms=(_sms, _ems))
            return {"shop": shop, "startDate": start, "endDate": end,
                    "total": stat["total"], "daily": stat.get("daily", []),
                    "messages": stat.get("messages", [])}
    try:
        stat = stat_trace_daily(shop_id, start, end, force=bool(force), trim_ms=(_sms, _ems))
    except BusyQueueError as e:
        raise HTTPException(503, str(e))  # 夜间抓取进行中
    except RuntimeError as e:
        raise HTTPException(502, str(e))
    return {"shop": shop, "startDate": start, "endDate": end,
            "total": stat["total"], "daily": stat.get("daily", []),
            "messages": stat.get("messages", [])}


@app.get("/api/risk")
def risk_status():
    """风控状态: triggered/reason/at, 供前端展示"""
    with _lock:
        return {
            "triggered": _risk_state["triggered"],
            "reason": _risk_state["reason"],
            "at": _risk_state["at"],
            "lastCode": _risk_state["last_code"],
        }


# ---------- 系统通知(风控/登录失效提醒, 页面铃铛展示) ----------
@app.get("/api/notifications")
def notifications_get(limit: int = 50):
    """系统通知列表(时间倒序) + 未读数"""
    with _notify_lock:
        items = _notify_load()
    items = sorted(items, key=lambda x: x.get("at", 0), reverse=True)[: max(1, min(int(limit), 100))]
    unread = sum(1 for it in items if not it.get("read"))
    return {"items": items, "unread": unread}


class NotifyReadBody(BaseModel):
    ids: list[str] | None = None  # None = 全部已读


@app.post("/api/notifications/read")
def notifications_read(body: NotifyReadBody, user: dict = Depends(auth.get_current_user)):
    """标记通知已读(ids 缺省/空 = 全部已读)"""
    with _notify_lock:
        items = _notify_load()
        if body.ids:
            id_set = set(body.ids)
            for it in items:
                if it.get("id") in id_set:
                    it["read"] = True
        else:
            for it in items:
                it["read"] = True
        _notify_save(items)
    return {"ok": True}


@app.post("/api/notifications/clear")
def notifications_clear(user: dict = Depends(auth.require_admin)):
    """清空全部通知(管理员)"""
    with _notify_lock:
        _notify_save([])
    return {"ok": True}


@app.get("/api/tasks/refresh")
def task_status(request: Request):
    with _lock:
        st = dict(_refresh_state)
    st["nightly"] = _nightly_flag_info()
    if st.get("triggered_by"):
        st["triggered_by"] = _mask_admin_label(request, st["triggered_by"])
    return st


# 任务 6 小时自动清理: 懒触发(访问任务列表时, 距上次清理 >1 小时才执行一次)
_last_task_prune_ts = [0.0]


def _maybe_prune_tasks():
    now = time.time()
    if now - _last_task_prune_ts[0] > 3600:
        _last_task_prune_ts[0] = now
        try:
            log_store.prune_tasks(days=log_store.TASK_RETENTION_HOURS / 24)
        except Exception:
            pass
        # 操作日志保留 90 天(每条非轮询请求一行, 无限膨胀拖慢查询/写放大)
        try:
            trace_store.prune_operation_log(keep_days=90)
        except Exception:
            pass


@app.get("/api/tasks/list")
def tasks_list(request: Request, limit: int = 10, offset: int = 0, status: str | None = None,
               task_type: str | None = None, since_hours: float | None = 6):
    """任务队列列表: 最近 limit 条(默认 10, 用户界面只保留 6 小时窗口)。

    - limit/offset: 池子分页(前端"查看全部"翻历史)
    - since_hours: 时间窗口, 默认 6; 传 0 表示不限时间(看全部历史, 池子用)
    - status/task_type: 分类筛选
    数据源为独立日志库 logs.db(任务持久化, 跨重启可查)。先懒同步 running 任务进度。
    非管理员看管理员发起的任务时, 发起人显示为 'admin'(隐藏真实账号)。
    """
    _maybe_prune_tasks()
    _sync_task_progress()
    limit = max(1, min(int(limit), 100))
    offset = max(0, int(offset))
    since = None
    if since_hours:
        since = time.time() - float(since_hours) * 3600
    tasks = log_store.query_tasks(limit=limit, offset=offset, status=status,
                                  task_type=task_type, since=since)
    total = log_store.count_tasks(status=status, since=since)
    with _tasks_lock:
        queue_pos = {qid: i + 1 for i, qid in enumerate(_task_queue)}
        live_ids = {tid for tid, t in _tasks.items() if t.get("status") in ("running", "queued")}
        # live 任务进度用内存最新值覆盖(DB 可能落后一个轮询); 并标记是否在队首
        for t in tasks:
            t["queuePosition"] = queue_pos.get(t["id"])
            m = _tasks.get(t["id"])
            if m:
                t["progress"] = m.get("progress", t.get("progress"))
            # 非管理员视角: 管理员发起的任务发起人显示为 'admin'
            t["requested_by"] = _mask_admin_label(request, t.get("requested_by"), t.get("requested_role"))
    return {"tasks": tasks, "live": list(live_ids),
            "running": _running_task_id,
            "anyRunning": _any_task_running(), "total": total,
            "nightly": _nightly_flag_info()}


@app.get("/api/logs")
def logs_list(tag: str | None = None, limit: int = 200):
    """抽屉日志窗口: 返回进程内环形缓冲的最近日志(不含任何 cookie/凭据)"""
    limit = max(1, min(int(limit), 500))
    with _log_lock:
        entries = list(_log_ring)
    if tag and tag.strip():
        wanted = {t for t in tag.split(",") if t.strip()}
        entries = [e for e in entries if e["tag"] in wanted]
    return {"logs": entries[-limit:], "total": len(entries), "limit": limit, "tag": tag or ""}


@app.get("/api/oplog")
def oplog_list(client_id: str | None = None, client_name: str | None = None,
               limit: int = 100, user_id: int | None = None,
               _: dict = Depends(auth.require_admin)):
    """操作日志查询(仅管理员可见): 按浏览器身份/昵称/用户筛选, 返回最近记录 + 去重操作者"""
    limit = max(1, min(int(limit), 500))
    logs = trace_store.query_operation_log(limit, client_id, client_name, user_id)
    return {"logs": logs, "clients": trace_store.list_operation_clients()}


# ---------- 管理员统一日志中心(操作日志 / 任务记录 / 系统日志) ----------
@app.get("/api/logs/center")
def logs_center(op_limit: int = 100, op_user_id: int | None = None, op_client_id: str | None = None,
                task_limit: int = 10, task_status: str | None = None, task_type: str | None = None,
                task_since_hours: float | None = 6,
                sys_limit: int = 100, sys_tag: str | None = None,
                _: dict = Depends(auth.require_admin)):
    """管理员日志中心: 三个分类一次返回(仅管理员)。

      - 操作日志(谁做了什么, trace.db)
      - 任务记录(任务队列历史: 谁请求/状态/进度, logs.db)
      - 系统日志(进程日志持久化, logs.db)
    各分类带 total 供前端"共 N 条"展示; 分页/筛选参数独立。
    """
    oplog = trace_store.query_operation_log(op_limit, op_client_id, None, op_user_id)
    oplog_total = trace_store.count_operation_log(op_client_id, op_user_id)
    task_since = None
    if task_since_hours:
        task_since = time.time() - float(task_since_hours) * 3600
    tasks = log_store.query_tasks(limit=task_limit, status=task_status,
                                  task_type=task_type, since=task_since)
    task_total = log_store.count_tasks(status=task_status, since=task_since)
    sys_logs = log_store.query_system_logs(limit=sys_limit, tag=sys_tag)
    sys_total = log_store.count_system_logs(tag=sys_tag)
    return {
        "oplog": {"logs": oplog, "total": oplog_total,
                  "clients": trace_store.list_operation_clients()},
        "tasks": {"logs": tasks, "total": task_total},
        "system": {"logs": sys_logs, "total": sys_total,
                   "tags": log_store.list_system_tags()},
    }


@app.delete("/api/logs/tasks")
def logs_tasks_delete(task_ids: str | None = None, task_status: str | None = None,
                      _: dict = Depends(auth.require_admin)):
    """删除任务记录: 传 task_ids(逗号分隔)按 id 删; 传 task_status 按状态清; 都不传清空全部。"""
    if task_ids:
        ids = [int(x) for x in task_ids.split(",") if x.strip().isdigit()]
        deleted = sum(log_store.delete_task(i) for i in ids)
    else:
        deleted = log_store.clear_tasks(task_status)
    return {"ok": True, "deleted": deleted}


@app.delete("/api/logs/system")
def logs_system_delete(sys_tag: str | None = None,
                       _: dict = Depends(auth.require_admin)):
    """删除系统日志: 传 sys_tag 删该标签; 不传清空全部。"""
    deleted = log_store.delete_system_logs(tag=sys_tag)
    return {"ok": True, "deleted": deleted}


@app.delete("/api/oplog/records")
def oplog_records_delete(ids: str | None = None, older_than_id: int | None = None,
                         client_id: str | None = None,
                         _: dict = Depends(auth.require_admin)):
    """删除操作日志: ids(逗号分隔)按 id 删; older_than_id 删该 id 之前; client_id 删该客户端。"""
    idset = [int(x) for x in (ids or "").split(",") if x.strip().isdigit()] if ids else None
    deleted = trace_store.delete_operation_log(ids=idset, older_than=older_than_id, client_id=client_id)
    return {"ok": True, "deleted": deleted}


@app.post("/api/logs/prune")
def logs_prune(_: dict = Depends(auth.require_admin)):
    """手动执行保留期裁剪(任务 14 天 / 系统日志 7 天)。"""
    r = log_store.prune_all()
    return {"ok": True, **r}


# ---------- 账号系统: 注册 / 登录 / 登出 / 当前用户 ----------
class AuthRegisterBody(BaseModel):
    username: str
    password: str


class AuthLoginBody(BaseModel):
    username: str
    password: str


class AdminCreateUserBody(BaseModel):
    username: str
    password: str
    role: str = "user"
    expire_date: str | None = None
    note: str | None = None


class AdminSetPasswordBody(BaseModel):
    new_password: str


class AdminSetStatusBody(BaseModel):
    status: str


class AdminSetExpireBody(BaseModel):
    expire_date: str | None = None


def _public_user(user: dict) -> dict:
    """对外用户信息(不含 password_hash)"""
    return {k: user.get(k) for k in ("id", "username", "role", "status",
                                     "expire_date", "created_at", "last_login_at", "note")}


@app.post("/api/auth/register")
def auth_register(body: AuthRegisterBody, request: Request):
    """开放注册(受 config.auth.register_enabled 开关控制); 注册用户角色=user
    限流: 同一来源 IP 在窗口内(默认 5 分钟)最多注册 max_per_ip(默认 3)次,
    超限返回 429, 窗口/次数可改 config.auth.register_rate_limit 实时调整。
    """
    if not auth.get_register_enabled():
        raise HTTPException(status_code=403, detail="当前未开放注册, 请联系管理员创建账号")
    ip = request.client.host if request.client else ""
    client_id = request.headers.get("x-client-id") or ""
    ua = request.headers.get("user-agent") or ""
    if not auth.check_register_allowed(ip, client_id, ua):
        raise HTTPException(status_code=429, detail="注册过于频繁, 请稍后再试")
    username = (body.username or "").strip()
    if not auth.valid_username(username):
        raise HTTPException(status_code=400, detail="用户名需为 3~64 位字母/数字/下划线/@/.(支持邮箱格式)")
    if not auth.valid_password_strong(body.password):
        raise HTTPException(status_code=400, detail="密码需 8 位以上, 且同时包含字母和数字")
    try:
        user = trace_store.create_user(username, auth.hash_password(body.password), role="user")
    except ValueError as e:
        raise HTTPException(status_code=409, detail=str(e))
    return {"ok": True, "user": _public_user(user)}


@app.post("/api/auth/login")
def auth_login(body: AuthLoginBody):
    """登录: 校验密码 + 封禁/到期, 返回签名 token + 用户信息"""
    username = (body.username or "").strip()
    user = trace_store.get_user_by_username(username)
    if not user or not auth.verify_password(body.password or "", user["password_hash"]):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    if user["status"] != "active":
        raise HTTPException(status_code=403, detail="账号已被封禁, 请联系管理员")
    if user.get("expire_date"):
        try:
            if datetime.date.fromisoformat(user["expire_date"]) < datetime.date.today():
                raise HTTPException(status_code=403, detail="账号已到期, 请联系管理员续期")
        except ValueError:
            pass
    token = auth.make_token(user["id"])
    trace_store.touch_last_login(user["id"])
    return {"token": token, "user": _public_user(user)}


@app.post("/api/auth/logout")
def auth_logout(_: dict = Depends(auth.get_current_user)):
    """登出(无状态 token, 前端丢弃即可; 此端点用于记录操作日志)"""
    return {"ok": True}


@app.get("/api/auth/me")
def auth_me(user: dict = Depends(auth.get_current_user)):
    """当前登录用户信息(前端首屏校验登录态用)"""
    return {"user": _public_user(user)}


@app.get("/api/auth/config")
def auth_config():
    """登录页配置: 是否开放注册"""
    return {"registerEnabled": auth.get_register_enabled()}


# ---------- 管理员后台 ----------
@app.get("/api/admin/users")
def admin_users(_: dict = Depends(auth.require_admin)):
    """用户列表(注册时间升序), 不含密码哈希"""
    return {"users": trace_store.list_users()}


@app.post("/api/admin/users")
def admin_create_user(body: AdminCreateUserBody, _: dict = Depends(auth.require_admin)):
    """管理员创建账号(注册关闭时由管理员开号的通道)"""
    username = (body.username or "").strip()
    if not auth.valid_username(username):
        raise HTTPException(status_code=400, detail="用户名需为 3~32 位字母/数字/下划线")
    if not auth.valid_password(body.password):
        raise HTTPException(status_code=400, detail="密码至少 6 位")
    role = body.role if body.role in ("admin", "user") else "user"
    expire = body.expire_date or None
    if expire:
        try:
            datetime.date.fromisoformat(expire)
        except ValueError:
            raise HTTPException(status_code=400, detail="到期时间格式应为 YYYY-MM-DD")
    try:
        user = trace_store.create_user(username, auth.hash_password(body.password),
                                       role=role, expire_date=expire, note=body.note or None)
    except ValueError as e:
        raise HTTPException(status_code=409, detail=str(e))
    return {"ok": True, "user": user}


@app.post("/api/admin/users/{uid}/password")
def admin_set_password(uid: int, body: AdminSetPasswordBody, _: dict = Depends(auth.require_admin)):
    """管理员重置用户密码"""
    if not auth.valid_password(body.new_password):
        raise HTTPException(status_code=400, detail="新密码至少 6 位")
    if not trace_store.get_user_by_id(uid):
        raise HTTPException(status_code=404, detail="用户不存在")
    trace_store.update_user_password(uid, auth.hash_password(body.new_password))
    return {"ok": True}


@app.post("/api/admin/users/{uid}/status")
def admin_set_status(uid: int, body: AdminSetStatusBody, admin: dict = Depends(auth.require_admin)):
    """封禁 / 解封用户(封禁后其 token 立即失效)
    禁止封禁自己(否则管理员把自己锁死, 且无人可解)。
    """
    if body.status not in ("active", "banned"):
        raise HTTPException(status_code=400, detail="状态须为 active 或 banned")
    if body.status == "banned" and uid == admin["id"]:
        raise HTTPException(status_code=400, detail="不能封禁自己的账号(会导致管理员失联)")
    if not trace_store.get_user_by_id(uid):
        raise HTTPException(status_code=404, detail="用户不存在")
    trace_store.set_user_status(uid, body.status)
    return {"ok": True}


@app.post("/api/admin/users/{uid}/expire")
def admin_set_expire(uid: int, body: AdminSetExpireBody, admin: dict = Depends(auth.require_admin)):
    """设置到期时间(YYYY-MM-DD; 传空/null 清除=永久)
    禁止设置自己的到期时间(否则把自己锁在门外且无人能解)。
    """
    if not trace_store.get_user_by_id(uid):
        raise HTTPException(status_code=404, detail="用户不存在")
    expire = body.expire_date or None
    if expire and uid == admin["id"]:
        raise HTTPException(status_code=400, detail="不能设置自己的到期时间(会导致管理员失联)")
    if expire:
        try:
            datetime.date.fromisoformat(expire)
        except ValueError:
            raise HTTPException(status_code=400, detail="到期时间格式应为 YYYY-MM-DD")
    trace_store.set_user_expire(uid, expire)
    return {"ok": True}


@app.post("/api/admin/users/{uid}/note")
def admin_set_user_note(uid: int, body: dict = Body(...), _: dict = Depends(auth.require_admin)):
    """管理员更新用户备注"""
    if not trace_store.get_user_by_id(uid):
        raise HTTPException(status_code=404, detail="用户不存在")
    note = str(body.get("note") or "").strip()
    trace_store.set_user_note(uid, note)
    return {"ok": True, "note": note}


@app.post("/api/admin/users/{uid}/legacy")
def admin_set_user_legacy(uid: int, body: dict = Body(...), _: dict = Depends(auth.require_admin)):
    """设置用户是否开放旧版看板(allow: bool; 管理员自身恒有权限, 不受此开关影响)"""
    if not trace_store.get_user_by_id(uid):
        raise HTTPException(status_code=404, detail="用户不存在")
    allow = bool(body.get("allow"))
    trace_store.set_user_allow_legacy(uid, allow)
    return {"ok": True, "allow_legacy": allow}


@app.post("/api/admin/settings/register")
def admin_set_register(enabled: bool = Body(True), _: dict = Depends(auth.require_admin)):
    """管理员开关开放注册"""
    auth.set_register_enabled(enabled)
    return {"ok": True, "registerEnabled": enabled}


@app.post("/api/cookies")
def update_cookies(body: CookieUpdate, _: dict = Depends(auth.require_admin)):
    mutate_config(lambda c: c.__setitem__("cookies", body.cookies))
    return {"ok": True}


@app.get("/api/auth/cookie-expiry")
def auth_cookie_expiry():
    """各登录 cookie 的到期日 + 最近到期剩余天数(仅日期, 无 cookie 值)"""
    cfg = load_config()
    expires_map = dict(cfg.get("cookie_expires") or {})
    today = datetime.date.today()
    days_left = None
    for day in expires_map.values():
        try:
            d = datetime.date.fromisoformat(day)
        except Exception:
            continue
        left = (d - today).days
        if days_left is None or left < days_left:
            days_left = left
    return {"expires": expires_map, "daysLeft": days_left}


# ---------- 静态页面 ----------
# ---------- 浏览器登录获取 Cookie ----------
_login_state = {
    "running": False,
    "phase": "idle",  # idle / opening / waiting_login / got_cookie / failed
    "message": "",
    "started_at": None,
    "last_error": None,
}
_login_lock = threading.Lock()

# 需要捕获的 cookie 名
LOGIN_COOKIE_NAMES = [
    "tanyu-account-id",
    "tanyu-agent-account",
    "tanyu-group-account",
    "tanyu-group-id",
]


def _login_worker():
    """后台线程: 启动浏览器让用户登录, 检测到 cookie 后写入 config"""
    try:
        from playwright.sync_api import sync_playwright

        with _login_lock:
            _login_state["running"] = True
            _login_state["phase"] = "opening"
            _login_state["message"] = "正在打开浏览器…"
            _login_state["started_at"] = time.time()

        with sync_playwright() as p:
            browser = p.chromium.launch(headless=False)
            ctx = browser.new_context()
            page = ctx.new_page()
            page.goto("https://agent.tanyuai.com/", wait_until="domcontentloaded", timeout=60000)
            with _login_lock:
                _login_state["phase"] = "waiting_login"
                _login_state["message"] = "请在浏览器中完成登录(推荐扫码登录以获取全部集团)…"

            # 轮询等待登录 cookie
            deadline = time.time() + 15 * 60  # 最多等15分钟
            while time.time() < deadline:
                page.wait_for_timeout(3000)
                cks = ctx.cookies()
                found = {c["name"]: c["value"] for c in cks if c["name"] in LOGIN_COOKIE_NAMES}
                if len(found) >= 3:  # 至少3个关键 cookie
                    # 记录各 cookie 到期日(playwright expires 是 epoch 秒; -1=SESSION 不记录)
                    try:
                        expires_map = {}
                        for c in cks:
                            if c["name"] in LOGIN_COOKIE_NAMES and c.get("expires", -1) not in (-1, None):
                                day = datetime.datetime.fromtimestamp(c["expires"]).strftime("%Y-%m-%d")
                                expires_map[c["name"]] = day
                    except Exception:
                        expires_map = {}
                    # 尝试读取扫码登录时写入 localStorage 的集团列表
                    try:
                        group_list = page.evaluate("() => localStorage.getItem('groupList')")
                        new_groups = json.loads(group_list) if group_list else None
                    except Exception:
                        new_groups = None

                    def _apply_login_cfg(c):
                        c["cookies"].update(found)
                        if expires_map:
                            c["cookie_expires"] = expires_map
                        # 集团列表: 只保留与现有 config.groups 匹配的集团(groupId 相同),
                        # 按 groupId 合并更新 —— 与扫码登录一致, 防止换成其他账号后
                        # groupList 全量覆盖导致"突然多出其他平台"
                        if isinstance(new_groups, list) and new_groups:
                            scanned = _scan_groups_allowed(new_groups)
                            _merge_scan_groups(c, scanned)

                    # 锁内读改写: cookie + 到期日 + 集团列表一次落盘, 不覆盖并发写者
                    mutate_config(_apply_login_cfg)
                    cfg = load_config()
                    # 新凭证已保存, 自动解除风控停止状态
                    _reset_risk()
                    group_note = ""
                    if cfg.get("groups"):
                        group_note = f", 发现 {len(cfg['groups'])} 个集团"
                    with _login_lock:
                        _login_state["phase"] = "got_cookie"
                        _login_state["message"] = (
                            "已获取 Cookie 并保存 " + ", ".join(found.keys()) + group_note
                        )
                    browser.close()
                    return
            with _login_lock:
                _login_state["phase"] = "failed"
                _login_state["message"] = "等待登录超时(15分钟), 未检测到 Cookie"
            browser.close()
    except Exception as e:
        with _login_lock:
            _login_state["phase"] = "failed"
            _login_state["message"] = f"登录流程出错: {e}"
    finally:
        with _login_lock:
            _login_state["running"] = False


@app.post("/api/login/start")
def login_start():
    """弹出浏览器窗口, 用户人工登录, 自动抓取 Cookie"""
    with _login_lock:
        if _login_state["running"]:
            return {"ok": False, "message": "登录流程已在运行中"}
        _login_state["running"] = True
        _login_state["phase"] = "opening"
        _login_state["message"] = "正在打开浏览器…"
    threading.Thread(target=_login_worker, daemon=True).start()
    return {"ok": True, "message": "浏览器已启动, 请完成登录"}


@app.get("/api/login/status")
def login_status():
    with _login_lock:
        return dict(_login_state)


@app.get("/api/login/close")
def login_close():
    """关闭浏览器窗口(可选)"""
    with _login_lock:
        _login_state["phase"] = "idle"
        _login_state["running"] = False
        _login_state["message"] = ""
    return {"ok": True}


# ---------- 微信扫码登录(二维码, 免浏览器) ----------
# 流程照搬探域前端(account 登录页 chunk), 与 tanyu 后台扫码完全一致:
#   1) GET  /api/gc/wx/get-auth-qr-code                  → {qrCodeUrl, scene}
#      二维码图片即微信官方 showqrcode 图(tanyu 后台原图), 前端直接展示
#   2) POST /api/gc/wx/query-login-res-by-scene?scene=x  (前端 1s 轮询)
#      扫码并确认后 data = {sceneStr, accountInfos: [{accountId, accountType, ...}]}
#   3) POST /api/gc/agent/auth/login-by-wx  {accountId, sceneStr, accountType}
#      → 响应 Set-Cookie 下发 tanyu-* cookie, 本地自动捕获保存
WX_QR_API = "https://agent.tanyuai.com/api/gc/wx"
WX_QR_TTL = 150          # 二维码有效期(秒), 超时前端提示重新生成
_wx_qr_state = {
    "scene": None,
    "info": None,        # 扫码确认后的完整 data(sceneStr/accountInfos)
    "created": 0.0,
    "last_poll": 0.0,
    "phase": "idle",     # idle / waiting / confirmed / done / expired / error
    "error": None,
}
_wx_qr_lock = threading.Lock()


def _wx_headers():
    """微信登录接口请求头: 与探域登录页一致, 不带 tanyu cookie"""
    return {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36",
        "Accept": "application/json, text/plain, */*",
        "Content-Type": "application/json",
        "Referer": "https://agent.tanyuai.com/",
        "Origin": "https://agent.tanyuai.com",
    }


def _capture_wx_cookies(resp):
    """从微信登录相关响应的 Set-Cookie 捕获 tanyu-* cookie 并持久化(仅登录链路)"""
    try:
        got = {}
        for name, value in resp.cookies.items():
            if name.startswith("tanyu-"):
                got[name] = value
        if not got:
            return False
        # 锁内读改写, 不覆盖并发写者
        mutate_config(lambda c: c.setdefault("cookies", {}).update(got))
        _capture_cookie_expiry(resp)
        log_line("auth", f"二维码登录捕获 cookie: {', '.join(sorted(got))}")
        return True
    except Exception as e:
        log_line("auth", f"捕获 cookie 失败: {e}")
        return False


def _scan_groups_allowed(infos):
    """扫码返回的账号列表 → 仅保留与现有 config.groups 匹配的集团(groupId 相同)

    需求: 换微信扫码登录时保留当前这几个集团; 新微信上出现的其他平台/新集团
    一律不加载, 防止 config.groups 被整体替换成微信账号全量绑定导致"突然多出
    其他平台"。匹配键用集团维度 id(groupId, 同一集团无论谁登录都不变)。
    返回构造好的集团列表(字段与 wx_qr_select 原逻辑一致)。
    """
    cfg = load_config()
    existing = {g.get("groupId") for g in (cfg.get("groups") or []) if g.get("groupId")}
    groups = []
    for g in infos or []:
        if not g.get("accountId"):
            continue
        gid = g.get("groupId") or g.get("chatbotGroupId") or ""
        if not gid or gid not in existing:
            continue  # 新集团/其他平台: 不加载
        ng = {
            "groupId": gid,
            "groupName": g.get("groupName") or g.get("accountName") or "未知集团",
            "accountId": g.get("accountId"),
            "accountType": g.get("accountType", 3),
            "ifEnable": g.get("ifEnable", True),
            "platform": g.get("platform"),
        }
        if not ng["platform"]:
            ng["platform"] = _infer_group_platform(ng)
        groups.append(ng)
    return groups


def _merge_scan_groups(cfg, scanned):
    """把扫码/登录得到的集团合并进 config.groups

    - 现有 groups 为空: 全量采用 scanned(首次初始化, 无既有集团可保留);
    - 非空: 仅更新 groupId 匹配的条目, 其余现有条目原样保留; 新出现的集团
      (含其他平台)一律不加入 —— 换微信后集团列表既不丢也不"多出平台"。
    """
    old = cfg.get("groups") or []
    if not old:
        cfg["groups"] = list(scanned)
        return
    by_id = {g.get("groupId"): g for g in scanned if g.get("groupId")}
    merged = []
    for g in old:
        upd = by_id.pop(g.get("groupId"), None)
        merged.append(upd if upd else g)
    cfg["groups"] = merged


@app.post("/api/wx-qr/start")
def wx_qr_start():
    """生成微信扫码登录二维码(照搬 tanyu 后台原图, 无需打开浏览器)"""
    try:
        r = requests.get(f"{WX_QR_API}/get-auth-qr-code", headers=_wx_headers(), timeout=20)
        d = r.json()
    except Exception as e:
        raise HTTPException(502, f"获取二维码失败: {e}")
    data = (d or {}).get("data") or {}
    if d.get("success") is False or not data.get("qrCodeUrl"):
        raise HTTPException(502, (d or {}).get("msg") or "获取二维码失败")
    # 拉取微信官方二维码原图转 base64(免前端跨域); 失败则回退直链
    qr_b64 = None
    try:
        ir = requests.get(data["qrCodeUrl"], headers=_wx_headers(), timeout=20)
        if ir.status_code == 200 and ir.content:
            import base64 as _b64
            mime = (ir.headers.get("Content-Type") or "image/png").split(";")[0].strip() or "image/png"
            qr_b64 = f"data:{mime};base64," + _b64.b64encode(ir.content).decode()
    except Exception:
        pass
    with _wx_qr_lock:
        _wx_qr_state.update({
            "scene": data.get("scene"),
            "info": None,
            "created": time.time(),
            "last_poll": 0.0,
            "phase": "waiting",
            "error": None,
        })
    log_line("auth", f"已生成扫码二维码 scene={data.get('scene')}")
    return {"ok": True, "scene": data.get("scene"), "qrCodeUrl": data["qrCodeUrl"],
            "qrImage": qr_b64, "expiresIn": WX_QR_TTL}


@app.get("/api/wx-qr/status")
def wx_qr_status():
    """轮询扫码状态(前端每 2s 调; 上游轮询限频 ≥1s)"""
    with _wx_qr_lock:
        st = dict(_wx_qr_state)
    if st["phase"] == "confirmed":
        # 只返回与现有 config.groups 匹配的账号(集团), 新平台/新集团不展示,
        # 前端不会出现"其他平台"可选项; 全被过滤时前端显示"未返回可进入的集团"
        infos = (st["info"] or {}).get("accountInfos") or []
        allowed = _scan_groups_allowed(infos)
        return {"phase": "confirmed", "accountInfos": allowed}
    if st["phase"] not in ("waiting",):
        return {"phase": st["phase"], "error": st["error"]}
    if time.time() - st["created"] > WX_QR_TTL:
        with _wx_qr_lock:
            if _wx_qr_state["phase"] == "waiting":
                _wx_qr_state["phase"] = "expired"
                _wx_qr_state["error"] = "二维码已过期, 请重新生成"
        return {"phase": "expired", "error": "二维码已过期, 请重新生成"}
    if st["last_poll"] and time.time() - st["last_poll"] < 1.0:
        return {"phase": "waiting"}
    try:
        r = requests.post(f"{WX_QR_API}/query-login-res-by-scene",
                          params={"scene": st["scene"]}, headers=_wx_headers(), timeout=15)
        d = r.json()
    except Exception as e:
        log_line("auth", f"扫码轮询异常: {e}")
        return {"phase": "waiting", "error": str(e)}
    if d.get("success") is False:
        with _wx_qr_lock:
            _wx_qr_state["phase"] = "error"
            _wx_qr_state["error"] = d.get("msg") or "轮询失败"
        return {"phase": "error", "error": _wx_qr_state["error"]}
    data = d.get("data") or {}
    if data.get("accountInfos"):
        # 扫码并确认成功: 记录完整 data(sceneStr + 集团列表), 顺带捕获轮询 Set-Cookie
        with _wx_qr_lock:
            _wx_qr_state["info"] = data
            _wx_qr_state["phase"] = "confirmed"
        _capture_wx_cookies(r)
        return {"phase": "confirmed", "accountInfos": data.get("accountInfos") or []}
    with _wx_qr_lock:
        _wx_qr_state["last_poll"] = time.time()
    return {"phase": "waiting"}


class WxQrSelectBody(BaseModel):
    account_id: str
    account_type: int = 3


@app.post("/api/wx-qr/select")
def wx_qr_select(body: WxQrSelectBody):
    """扫码确认后选择集团: 调 login-by-wx 获取 tanyu-* cookie 并自动保存

    完成后: 合并保存集团列表(仅保留现有集团, 新平台不加载) → 捕获 cookie →
    强制切到选中账号的集团并同步店铺 → 解除风控停止。
    强制切换: 无论 login-by-wx 是否下发 tanyu-group-id, 都以选中账号所属集团
    为准切换, 防止 config 里残留旧微信的 tanyu-group-id 导致串集团。
    """
    with _wx_qr_lock:
        st = dict(_wx_qr_state)
    info = st.get("info") or {}
    scene_str = info.get("sceneStr")
    if not scene_str:
        raise HTTPException(400, "请先扫码并确认登录")
    payload = {"accountId": body.account_id, "sceneStr": scene_str,
               "accountType": body.account_type}
    try:
        r = requests.post(f"{GC_API}/agent/auth/login-by-wx", json=payload,
                          headers=_wx_headers(), timeout=20)
        d = r.json()
    except Exception as e:
        raise HTTPException(502, f"登录失败: {e}")
    if d.get("success") is False:
        raise HTTPException(502, d.get("msg") or "登录失败")
    _capture_wx_cookies(r)
    # 集团列表: 只保留与现有 config.groups 匹配的集团(新平台/新集团不加载),
    # 并按 groupId 合并更新(未匹配的现有集团原样保留, 不整体替换)
    scanned = _scan_groups_allowed(info.get("accountInfos") or [])

    def _apply_login_groups(c):
        _merge_scan_groups(c, scanned)

    mutate_config(_apply_login_groups)
    cfg = load_config()
    # 选中的账号必须属于保留集团(前端只展示这些, 此处兜底拦截误选)
    sel = next((g for g in (cfg.get("groups") or [])
                if g.get("accountId") == body.account_id), None)
    if not sel or not sel.get("groupId"):
        _reset_risk()
        with _wx_qr_lock:
            _wx_qr_state["phase"] = "done"
            _wx_qr_state["error"] = None
        log_line("auth", f"二维码登录: 账号 {body.account_id} 不属于看板保留集团, 已保存 Cookie 未切换集团")
        return {"ok": True, "group": None,
                "warning": "该账号所属集团不在看板支持的集团内, 已保存 Cookie 但未切换集团"}
    # 强制把激活集团切到选中账号的集团(拉取该集团 cookie), 而非仅当 config
    # 缺少 tanyu-group-id 才切 —— 覆盖"新微信 cookie + 残留旧微信 group-id"的串集团场景
    switched = False
    try:
        if cfg.get("cookies", {}).get("tanyu-group-id") != sel["groupId"]:
            switch_group(sel["groupId"])
        else:
            sync_shops_from_tanyu()
        switched = True
    except Exception as e:
        log_line("auth", f"登录后切换集团失败: {e}")
    _reset_risk()
    with _wx_qr_lock:
        _wx_qr_state["phase"] = "done"
        _wx_qr_state["error"] = None
    cur = get_current_group()
    log_line("auth", f"二维码登录成功: {cur.get('name') if cur else ''} ({body.account_id})")
    resp = {"ok": True, "group": cur}
    if not switched:
        resp["warning"] = "Cookie 已保存, 但切换集团失败, 请检查该集团账号在新微信下是否可用"
    return resp


_NO_CACHE_HEADERS = {"Cache-Control": "no-store, no-cache, must-revalidate", "Pragma": "no-cache"}


@app.get("/")
def index():
    """入口: 重定向到独立登录页(登录成功后进入 /app 新后台)"""
    return RedirectResponse("/login", headers=_NO_CACHE_HEADERS)


@app.get("/legacy")
def legacy_index(user: dict = Depends(auth.get_current_user)):
    """旧版看板: 管理员恒可访问; 普通用户需管理员在「更多」板块勾选开放, 否则跳回新版"""
    if user["role"] != "admin" and not user.get("allow_legacy"):
        return RedirectResponse("/app", headers=_NO_CACHE_HEADERS)
    return FileResponse(BASE_DIR / "static" / "index.html", headers=_NO_CACHE_HEADERS)


@app.get("/login")
def login_page():
    """独立登录页: 登录/注册后写 token 跳转 /app"""
    return FileResponse(BASE_DIR / "static" / "login.html", headers=_NO_CACHE_HEADERS)


@app.get("/app")
def app_page():
    """新版后台: 侧边栏多模块, 前端鉴权, 未登录自动跳 /login"""
    return FileResponse(BASE_DIR / "static" / "app" / "index.html", headers=_NO_CACHE_HEADERS)


@app.get("/preview")
def tabler_preview():
    """Tabler 套壳原型页(评估用): 展示现有看板套用 Tabler 后的视觉"""
    return FileResponse(BASE_DIR / "static" / "tabler_preview.html")


# 静态资源(vendored Tabler 等): /static/vendor/... 与 /static/tabler_preview.html
from fastapi.staticfiles import StaticFiles

_static_dir = BASE_DIR / "static"
_static_dir.mkdir(exist_ok=True)
app.mount("/static", StaticFiles(directory=_static_dir), name="static")


# ---------- SQLite 预抓 / 回填 ----------
_last_prefetch_duration = None  # 最近一次夜间预抓耗时(秒), 供 /api/db/status 展示


def _use_sqlite_trace():
    """config.json 开关: use_sqlite_trace(默认 true)"""
    try:
        cfg = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
        return cfg.get("use_sqlite_trace", True)
    except Exception:
        return True


def backfill_trace_db():
    """一次性回填: 把现有 data/trace_days 逐日 JSON 全量投影进 SQLite(幂等可重跑)"""
    import glob

    trace_store.init_db()
    shops = load_shops()
    trace_store.upsert_shops(shops)
    shop_map = {s.get("thirdShopId"): s.get("platform", 0) for s in shops}
    files = sorted(glob.glob(str(DATA_DIR / "trace_days" / "*.json")))
    total_rows = total_shops = 0
    t0 = time.time()
    for i, f in enumerate(files, 1):
        shop_id = Path(f).stem
        try:
            cache = json.loads(Path(f).read_text(encoding="utf-8"))
        except Exception as e:
            print(f"[backfill] 读取失败 {f}: {e}")
            continue
        days = cache.get("days") or {}
        platform = shop_map.get(shop_id, 0)
        for day_str, msgs in days.items():
            try:
                trace_store.upsert_shop_day(shop_id, platform, day_str, msgs)
            except Exception as e:
                print(f"[backfill] {shop_id} {day_str} 失败: {e}")
        n = sum(len(v) for v in days.values())
        total_rows += n
        total_shops += 1
        if i % 20 == 0 or i == len(files):
            print(f"[backfill] {i}/{len(files)} 家, 已入库 {total_rows} 条, "
                  f"{time.time() - t0:.0f}s")
    print(f"[backfill] 完成: {total_shops} 家 / {total_rows} 条 / {time.time() - t0:.0f}s")


def _prefetch_refresh_week_summary(shop, cfg):
    """预抓时按需刷新店铺 natural_week summary 缓存(生成率聚合数据源)。

    trace 预抓只写 trace_daily, 不碰 summary 缓存; 平台级生成率聚合依赖
    店铺 summary 缓存(ai_consult_response_rate), 若长期不刷, 跨周后旧值
    过期, 周一早高峰推送/概览生成率会回落 None。这里在 trace 抓取后顺带
    补刷: 仅当 缓存缺失或已过期(本期 TTL 外) 才请求, 命中缓存零请求,
    且受 _rate_limit 约束, 不放大风控风险。cfg 用于检查 prefetch 开关。
    """
    try:
        if not (cfg.get("prefetch_refresh_summary", True)):
            return
        platform = shop.get("platform", 0)
        if platform not in FETCH_PLATFORMS:
            return  # 导入平台不走 tanyu, 无 summary 缓存
        sid = shop["thirdShopId"]
        ws, we = _stat_type_range("natural_week")
        key = f"{sid}__natural_week"
        cache = load_cache("summary", key, max_age=_period_max_age("natural_week"))
        if cache and cache.get("data"):
            return  # 本周缓存有效, 无需刷新
        _assert_no_risk()
        _rate_limit()
        payload = {
            "statType": "natural_week",
            "startDate": ws,
            "endDate": we,
            "platform": platform,
            "dimension": "shop",
            "targetId": sid,
        }
        items = fetch_summary_interactive(payload)
        save_cache("summary", key, {"fetched_at": time.time(), "data": items})
        print(f"[prefetch]     ↻ 刷新 {shop.get('shopName','')} summary(natural_week)")
    except RiskTriggered:
        raise
    except BusyQueueError as e:
        # 限速繁忙/夜间门控: 单店 summary 刷新跳过(下轮/夜间任务会补), 非错误不刷日志
        pass
    except Exception as e:
        # 单店 summary 刷新失败不阻塞 trace 预抓(数据缺失只是生成率 '—')
        log_line("prefetch", f"刷新 summary 失败 {shop.get('shopName','')}: {e}")


def _prefetch_group(gid, start, end, force_days=None, progress_cb=None, done_shops=0,
                    cancel_check=None):
    """切换集团→同步店铺→抓该集团全部店铺 trace, 返回 (店铺总数, 失败数)

    依赖 switch_group 内部自动 sync_shops_from_tanyu(写 shops.json + SQLite 店铺表),
    切换后 load_shops 即读当前集团店铺。任一店铺 RiskTriggered 向上传播(整个预抓停止)。
    force_days: 窗口内这些天强制重抓(默认最近一天=昨天, 防 tanyu 回溯更新 sendType
    造成的采纳口径漂移), 由 config.prefetch_force_days 控制(0=关闭)。
    progress_cb: 每完成一家店铺回调 {"done", "total", "current"}; done_shops 为调用方
      已累计完成数, 供跨集团累计进度。
    cancel_check: 可选回调 fn() -> bool, 每店循环开头调用; 返回 True 则立即停止本集团
      (用于手动补抓取消, 不用等集团跑完)。
    """
    cfg = load_config()
    g = next((x for x in cfg.get("groups", []) if x.get("groupId") == gid), None)
    if not g:
        print(f"[prefetch] ⚠️ 集团 {gid} 不在 config.groups 中, 跳过")
        return 0, 0
    if not g.get("accountId"):
        print(f"[prefetch] ⚠️ 集团「{g.get('groupName')}」缺少 accountId, 无法切换, 跳过")
        return 0, 0
    # switch_group 是每集团第一跳(切激活集团 cookie + 同步店铺), 一旦失败整组跳过、
    # 该集团全部店铺当夜丢数据。DNS 解析失败/SSL 断连等瞬断通常几秒~几十秒自愈,
    # 用退避重试(3 次)扛过瞬时网络抖动; 风控/业务错误由 _with_network_retry 原样抛出不重试。
    _with_network_retry(lambda: switch_group(gid),
                        f"切换集团「{g.get('groupName')}」", tag="prefetch")
    shops = load_shops()
    print(f"[prefetch] 集团「{g.get('groupName')}」({gid}) 店铺 {len(shops)} 家, "
          f"窗口 {start} ~ {end}" + (f", 强制重抓最近 {len(force_days)} 天" if force_days else ""))
    failed = []
    for i, shop in enumerate(shops, 1):
        if cancel_check and cancel_check():
            print(f"[prefetch] ⏹ 收到取消请求, 停止于 {i - 1}/{len(shops)} 家")
            break
        if _risk_state.get("triggered"):
            print(f"[prefetch] ⛔ 风控/登录失效, 集团内停止于 {i - 1}/{len(shops)} 家")
            raise RiskTriggered(_risk_state.get("reason") or "风控触发")
        sid = shop["thirdShopId"]
        try:
            # history_mode=True: 历史天缓存已抓就复用(不按 7 天 TTL 判过期),
            # 使每晚只增量抓昨天/缺格天, 不周期性重抓整窗旧天
            stat = stat_trace_daily(sid, start, end, force_days=force_days, history_mode=True)
            print(f"[prefetch]   {i}/{len(shops)} {shop.get('platformName','')}·{shop.get('shopName','')} "
                  f"total={stat['total']} adopted={stat['adopted']}")
            if progress_cb:
                done = done_shops + i
                progress_cb({"done": done, "total": None,
                             "current": f"{g.get('groupName','')}·{shop.get('shopName','')}"})
            _prefetch_refresh_week_summary(shop, cfg)
        except RiskTriggered:
            print(f"[prefetch] ⛔ 风控触发于 {shop.get('shopName','')}")
            raise
        except Exception as e:
            print(f"[prefetch]   {shop.get('shopName','')} 失败: {e}")
            failed.append(sid)
    return len(shops), len(failed)


def repair_platform_attribution():
    """按 SQLite 店铺表把历史 platform=0 的消息纠正为正确平台(按 shop_id 映射)

    早期全量回填时无平台信息, 359K 条消息 platform=0。三集团同步后 shops 表已含
    各集团店铺的正确平台, 这里做一次兜底修复: 逐店 UPDATE 该店仍为 0 的行,
    并重建受影响店的按日聚合。shop_id 不在 shops 表的孤儿消息保持原样。
    """
    import sqlite3
    try:
        c = sqlite3.connect(trace_store.db_path())
        rows = c.execute(
            "SELECT third_shop_id, platform FROM shops WHERE third_shop_id IN "
            "(SELECT DISTINCT third_shop_id FROM messages WHERE platform = 0)"
        ).fetchall()
        c.close()
    except Exception as e:
        print(f"[repair] 读取映射失败: {e}")
        return 0
    if not rows:
        return 0
    fixed = 0
    for sid, platform in rows:
        try:
            import trace_store as _ts
            _ts.repair_shop_platform(sid, platform)
            fixed += 1
        except Exception as e:
            print(f"[repair] {sid} 修复失败: {e}")
    print(f"[repair] 平台归属修复完成: {fixed} 家店铺(shop 表映射)已归位")
    return fixed


def prefetch_trace_window(days=None, prune=True, progress_cb=None):
    """夜间预抓: 轮询拼多多→京东→抖音三集团, 按集团窗口抓缺失/过期天, 并滚动裁剪

    每天 00:00 由计划任务 TanyuDashboardTracePrefetch 触发:
      零点一过, 昨天才真正定型, 这里把"昨天"整体抓取入库(当天数据当天实时展示,
      不进库); 完成后 prune_window 按整日边界裁剪。
    trace API 以激活集团(cookie tanyu-group-id)为作用域, 所以多平台预抓逐集团:
      switch_group → sync_shops_from_tanyu → stat_trace_daily(该集团店铺) → 下一个。
    窗口: 全平台保留 35 天(窗口上限, 每晚滚动: 新增昨天、超窗口的最老完整日被裁剪)。
      集团窗口取 config.prefetch_windows{platform: days}, 未配置的集团用
      config.prefetch_days(默认 7)。已抓的天命中缓存零请求(增量)。
    强制重抓: config.prefetch_force_days(默认 1)指定窗口最近 N 天即使缓存有效也
      强制重抓(默认最近一天=昨天)。tanyu 会对已抓消息回溯更新 sendType(草稿→已发送),
      不重抓则昨天采纳数停在预抓时刻、看板采纳率与 tanyu 后台不一致。
    结束后恢复原激活集团, 让常驻看板继续服务原集团。风控/登录失效立即整体停止。

    prune=False: 跳过最后的 prune_window 滚动裁剪(由调用方 nightly_fetch.py 决定
      何时进入"满 35 天连续无缺口才裁最老一天"的滚动模式)。默认 True 保持既有行为。
    progress_cb: 可选回调 fn(progress_dict), 每完成一家店铺调用一次
      {"done": 累计完成店铺数, "total": 总店铺数, "current": "集团·店名"}, 供前端进度条。
    """
    t_all = time.time()
    trace_store.init_db()
    cfg = load_config()
    days = days or int(cfg.get("prefetch_days", 7))
    # prefetch_windows 的 key 可能是 JSON 字符串("5")或数字(5), 统一归一化为 int
    window_map = {int(k): int(v) for k, v in (cfg.get("prefetch_windows") or {}).items()}
    force_days = int(cfg.get("prefetch_force_days", 1) or 0)
    platform_order = cfg.get("prefetch_platforms") or [1, 5, 7]
    groups = cfg.get("groups") or []
    # 按平台排序优先级(1 拼多多, 5 抖音, 7 京东), 未标注平台的集团放最后
    plat_rank = {p: i for i, p in enumerate(platform_order)}
    ordered = sorted(
        [g for g in groups if g.get("groupId")],
        key=lambda g: plat_rank.get(g.get("platform"), 99),
    )
    if not ordered:
        print("[prefetch] ⚠️ config.groups 为空, 无可预抓集团")
        return
    # 记录进入前的激活集团, 预抓结束后切回
    orig_gid = cfg.get("cookies", {}).get("tanyu-group-id")
    kept_days = days  # 最终裁剪窗口天数(取各集团窗口最大值)
    # 预计算全量店铺数, 使进度条 done/total 有确定分母
    try:
        total_shops = len([s for s in load_all_shops() if s.get("platform") in FETCH_PLATFORMS])
    except Exception:
        total_shops = 0
    done_shops = 0
    # 写跨进程标志: 8080 常驻检测到夜间抓取占用激活集团 cookie 时, 直连请求返回
    # 繁忙不向 tanyu 出网, 防止带错集团 cookie 串数据
    set_nightly_fetch_flag()
    try:
        g_total = g_failed = 0
        for g in ordered:
            if _risk_state.get("triggered"):
                print("[prefetch] ⛔ 风控/登录失效, 整体停止")
                raise RiskTriggered(_risk_state.get("reason") or "风控触发")
            # 集团窗口: 按平台映射取保留天数(PDD/JD/DY 均 35 天), 未映射的集团用默认 days
            gid = g["groupId"]
            g_days = int(window_map.get(g.get("platform"), days))
            kept_days = max(kept_days, g_days)
            start, end = date_range(g_days)
            # 强制重抓"昨天"等最近 N 天(防 tanyu 回溯更新 sendType 造成的采纳漂移)
            fset = set()
            if force_days > 0:
                from datetime import timedelta as _td
                end_d = datetime.date.fromisoformat(end)
                fset = {(end_d - _td(days=i)).isoformat() for i in range(min(force_days, g_days))}
            print(f"[prefetch] ===== 集团「{g.get('groupName')}」窗口 {g_days} 天 "
                  f"({start} ~ {end}) =====" + (f" 强制重抓: {sorted(fset)}" if fset else ""))
            try:
                # 包装 progress_cb: _prefetch_group 内部不知道全量总店数, 在此填真实 total
                def _cb(p):
                    if progress_cb:
                        progress_cb({"done": p["done"], "total": total_shops or p["done"],
                                     "current": p["current"]})
                n, f = _prefetch_group(gid, start, end, force_days=fset,
                                       progress_cb=_cb, done_shops=done_shops)
                done_shops += n
                g_total += n
                g_failed += f
            except RiskTriggered:
                print("[prefetch] ⛔ 风控触发, 整体停止")
                raise
        print(f"[prefetch] 全部集团完成: {g_total} 家店铺, {g_failed} 家失败, "
              f"耗时 {time.time() - t_all:.0f}s")
    except RiskTriggered:
        # 风控触发: 向上 re-raise, 使计划任务以非零退出码结束(LastTaskResult≠0),
        # 运维可见"夜间抓取失败"而非静默记成功。已抓的数据照常落盘(逐天 save_cache)。
        print("[prefetch] ⛔ 风控/登录失效, 停止轮询并向上抛出(任务将记为失败)")
        raise
    finally:
        # 兜底修复历史 platform=0 消息(依赖 shops 表, 各集团已同步)
        try:
            repair_platform_attribution()
        except Exception as e:
            print(f"[repair] 兜底修复异常: {e}")
        # 裁剪窗口到已抓数据的最大集团窗口(全平台 35 天)
        # nightly_fetch.py 以 prune=False 调用时跳过裁剪, 由其满窗守卫决定何时滚动
        if prune:
            try:
                deleted = trace_store.prune_window(keep_days=kept_days)
                print(f"[prefetch] 裁剪完成(保留 {kept_days} 天): 删除 {deleted} 条过期消息")
            except Exception as e:
                print(f"[prefetch] 裁剪失败: {e}")
        else:
            print("[prefetch] prune=False: 跳过滚动裁剪(由 nightly_fetch 满窗守卫接管)")
        # 恢复原激活集团(常驻看板继续服务原集团)。
        # 风控触发时 switch_group 会被 _assert_no_risk 拦截(恢复必然失败),
        # 无法再出网切换; 仅告警说明 cookie 可能停在切换中途的集团, 需重新登录。
        if orig_gid and not _risk_state.get("triggered"):
            try:
                if orig_gid != load_config().get("cookies", {}).get("tanyu-group-id"):
                    # invalidate=False: 恢复集团不清核算缓存(刷新任务不影响已算好的结果)
                    switch_group(orig_gid, invalidate=False)
                    print(f"[prefetch] 已恢复原激活集团 {orig_gid}")
            except Exception as e:
                print(f"[prefetch] ⚠️ 恢复原集团失败: {e}")
        elif _risk_state.get("triggered"):
            print("[prefetch] ⚠️ 风控触发, cookie 可能停在切换中途的集团; 风控下无法切换, 请重新登录")
        # 清除跨进程标志(独立进程抓完, 8080 直连恢复)
        clear_nightly_fetch_flag()
    _last_prefetch_duration = time.time() - t_all
    print(f"[prefetch] 完成: 窗口 {kept_days} 天, 总耗时 {_last_prefetch_duration:.0f}s")


@app.get("/api/db/status")
def db_status():
    """SQLite 存储状态(供前端/运维确认预抓进度)"""
    import sqlite3

    info = {"enabled": _use_sqlite_trace(), "path": trace_store.db_path()}
    try:
        c = sqlite3.connect(trace_store.db_path())
        info["messages"] = c.execute("SELECT COUNT(*) FROM messages").fetchone()[0]
        info["shops"] = c.execute("SELECT COUNT(DISTINCT third_shop_id) FROM messages").fetchone()[0]
        info["daily"] = c.execute("SELECT COUNT(*) FROM trace_daily").fetchone()[0]
        info["platform0_count"] = c.execute(
            "SELECT COUNT(*) FROM messages WHERE platform = 0"
        ).fetchone()[0]
        lo, hi = c.execute("SELECT MIN(day), MAX(day) FROM trace_daily").fetchone()
        info["window"] = [lo, hi]
        c.close()
    except Exception as e:
        info["error"] = str(e)
    # 预抓配置与最近一次耗时
    try:
        cfg = load_config()
        info["prefetch_days"] = cfg.get("prefetch_days", 7)
        info["prefetch_platforms"] = cfg.get("prefetch_platforms", [1, 5, 7])
        info["prefetch_windows"] = cfg.get("prefetch_windows", {5: 30})
        info["prefetch_force_days"] = cfg.get("prefetch_force_days", 1)
    except Exception:
        pass
    info["prefetch_duration"] = _last_prefetch_duration
    return info


# ---------- 钉钉机器人接入 ----------
# 两种方式: A=自定义机器人Webhook推送, B=Stream长连接交互。
# 独立模块 dingtalk_bot.py + 独立配置 dingtalk_config.json(已 gitignore)。
# 未配置 webhook/stream 凭据时所有端点安全降级(返回未配置), 不影响看板主流程。
try:
    import dingtalk_bot
    HAVE_DINGTALK = True
except Exception as e:
    HAVE_DINGTALK = False
    print(f"[dingtalk] 模块加载失败(钉钉功能禁用): {e}")


class DingTalkConfigBody(BaseModel):
    webhook_url: str | None = None
    webhook_keyword: str | None = None
    stream_client_id: str | None = None
    stream_client_secret: str | None = None


@app.get("/api/dingtalk/status")
def dingtalk_status():
    """钉钉接入状态(是否配置/连接是否建立), 不暴露任何凭据值"""
    if not HAVE_DINGTALK:
        return {"enabled": False, "error": "模块未加载"}
    cfg = dingtalk_bot.load_config()
    web = bool((cfg.get("webhook") or {}).get("url"))
    st = cfg.get("stream") or {}
    stream_ok = bool(st.get("client_id") and st.get("client_secret"))
    return {
        "enabled": True,
        "webhook_configured": web,
        "stream_configured": stream_ok,
        "stream_running": dingtalk_bot._stream_state.get("running", False),
        "stream_error": dingtalk_bot._stream_state.get("error"),
        "stream_conn_at": dingtalk_bot._stream_state.get("conn_at"),
        "note": "凭据只存于 dingtalk_config.json(不入 git)",
    }


@app.get("/api/dingtalk/push")
def dingtalk_push(platform: str | None = None, stat_type: str = "natural_week", week: str | None = None):
    """方式A: 把平台概览推送到钉钉群(自定义机器人 webhook)

    platform: 平台号(如 10)或 all(全部); 留空默认 all。
    """
    if not HAVE_DINGTALK:
        raise HTTPException(503, "钉钉模块未加载")
    plat = _parse_dingtalk_platform(platform)
    return dingtalk_bot.push_overview_to_group(plat, stat_type, week)


@app.post("/api/dingtalk/config")
def dingtalk_set_config(body: DingTalkConfigBody):
    """保存钉钉配置(webhook / stream 凭据); 会写 dingtalk_config.json(不入 git)"""
    if not HAVE_DINGTALK:
        raise HTTPException(503, "钉钉模块未加载")
    cfg = dingtalk_bot.load_config()
    cfg.setdefault("webhook", {})
    cfg.setdefault("stream", {})
    if body.webhook_url is not None:
        cfg["webhook"]["url"] = body.webhook_url.strip() or None
    if body.webhook_keyword is not None:
        cfg["webhook"]["keyword"] = body.webhook_keyword.strip() or None
    if body.stream_client_id is not None:
        cfg["stream"]["client_id"] = body.stream_client_id.strip() or None
    if body.stream_client_secret is not None:
        cfg["stream"]["client_secret"] = body.stream_client_secret.strip() or None
    dingtalk_bot.save_config(cfg)
    return {"ok": True, "saved": "dingtalk_config.json"}


@app.post("/api/dingtalk/stream/start")
def dingtalk_stream_start():
    """方式B: 启动 Stream 长连接(后台线程)"""
    if not HAVE_DINGTALK:
        raise HTTPException(503, "钉钉模块未加载")
    return dingtalk_bot.ensure_stream()


@app.post("/api/dingtalk/stream/stop")
def dingtalk_stream_stop():
    """方式B: 停止 Stream 长连接"""
    if not HAVE_DINGTALK:
        raise HTTPException(503, "钉钉模块未加载")
    dingtalk_bot.stop_stream()
    return {"ok": True, "running": False}


@app.get("/api/dingtalk/preview")
def dingtalk_preview(platform: str | None = None, stat_type: str = "natural_week", week: str | None = None):
    """预览将要推送到钉钉的 markdown 文本(不实际发送)"""
    if not HAVE_DINGTALK:
        raise HTTPException(503, "钉钉模块未加载")
    plat = _parse_dingtalk_platform(platform)
    txt, detail = dingtalk_bot._overview_cards(plat, stat_type, week)
    return {"markdown": txt, "platforms": detail}


def _parse_dingtalk_platform(p):
    """钉钉推送的 platform 参数: 数字平台号或 all/None -> int 或 'all'"""
    if p is None or str(p).strip().lower() in ("", "all", "全部"):
        return "all"
    try:
        return int(str(p).strip())
    except (TypeError, ValueError):
        return "all"


if __name__ == "__main__":
    if "--backfill" in sys.argv:
        backfill_trace_db()
        sys.exit(0)
    if "--prefetch" in sys.argv:
        try:
            prefetch_trace_window()
        except RiskTriggered as e:
            print(f"[prefetch] ⛔ 风控/登录失效: {e} — 任务失败(LastTaskResult≠0)")
            sys.exit(2)
        except Exception as e:
            print(f"[prefetch] ⛔ 异常: {e}")
            sys.exit(1)
        sys.exit(0)
    # 常驻服务启动时初始化 SQLite 表结构
    try:
        trace_store.init_db()
        # 导入平台(天猫1/2)店铺只进 SQLite shops 表(不进 shops.json, 避免被切集团覆写)
        trace_store.upsert_shops(IMPORT_SHOPS)
    except Exception as e:
        print(f"[db] 初始化失败: {e}")
    # 独立日志库: 建表 + 恢复任务历史(任务列表跨重启可查) + 裁剪过期记录
    try:
        log_store.init_db()
        _restore_task_history()
        log_store.prune_all()
    except Exception as e:
        print(f"[logs] 初始化失败: {e}")
    # 账号系统: 首次生成 token 签名密钥 + 按 config.json 的 auth.admin_username/admin_password
    # 创建初始管理员(config 指定; 创建后 admin_password 明文自动清除)
    try:
        auth.get_secret()
        auth.ensure_admin()
    except Exception as e:
        print(f"[auth] 初始化失败: {e}")
    port = int(os.environ.get("PORT", "8080"))
    uvicorn.run(app, host="127.0.0.1", port=port)
